from dotenv import load_dotenv
from pathlib import Path
import os

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Request, Depends, Response, Form, File, UploadFile
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import logging
from pydantic import BaseModel, Field, EmailStr, ConfigDict, BeforeValidator
from typing import List, Optional, Annotated
from datetime import datetime, timezone, timedelta, date as DateClass
from bson import ObjectId
from io import BytesIO
from xml.sax.saxutils import quoteattr
import base64
import bcrypt
import resend
import jwt
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_ALGORITHM = "HS256"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
PyObjectId = Annotated[str, BeforeValidator(str)]


def to_object_id(value: str) -> ObjectId:
    try:
        return ObjectId(value)
    except Exception:
        raise HTTPException(status_code=404, detail="Hittades inte")


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))


def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]


def create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "access",
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        token = request.query_params.get("token")
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user["_id"] = str(user["_id"])
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class BookingCreate(BaseModel):
    name: str = Field(..., min_length=1)
    email: EmailStr
    phone: str = Field(..., min_length=1)
    address: Optional[str] = None
    kvm: Optional[str] = None
    services: List[str] = Field(default_factory=list)
    preferred_date: Optional[str] = None
    other_description: Optional[str] = None


class Booking(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    id: Optional[PyObjectId] = Field(default=None, alias="_id")
    name: str
    email: str
    phone: str
    address: Optional[str] = None
    kvm: Optional[str] = None
    services: List[str] = Field(default_factory=list)
    preferred_date: Optional[str] = None
    other_description: Optional[str] = None
    status: str = "new"
    created_at: str


class StatusUpdate(BaseModel):
    status: str


class ReviewCreate(BaseModel):
    name: str = Field(..., min_length=1)
    rating: int = Field(..., ge=1, le=5)
    text: str = Field(..., min_length=1)


class Review(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    id: Optional[PyObjectId] = Field(default=None, alias="_id")
    name: str
    rating: int
    text: str
    approved: bool = False
    created_at: str


class ApproveUpdate(BaseModel):
    approved: bool


# ---- Employees ----
class EmployeeCreate(BaseModel):
    name: str = Field(..., min_length=1)
    phone: Optional[str] = None
    color: str = "#166534"
    hourly_rate: float = 0
    personnummer: Optional[str] = None
    employment_type: str = "fastanstalld"  # "fastanstalld" or "vikarie"
    skatt_typ: Optional[str] = "A-skatt"  # A-skatt | F-skatt
    fskatt_nr: Optional[str] = None
    fskatt_verified: Optional[bool] = False


class Employee(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    id: Optional[PyObjectId] = Field(default=None, alias="_id")
    name: str
    phone: Optional[str] = None
    color: str = "#166534"
    skatt_typ: Optional[str] = "A-skatt"
    fskatt_nr: Optional[str] = None
    fskatt_verified: Optional[bool] = False
    hourly_rate: float = 0
    personnummer: Optional[str] = None
    employment_type: str = "fastanstalld"
    created_at: str


class EmployeeUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    color: Optional[str] = None
    skatt_typ: Optional[str] = None
    fskatt_nr: Optional[str] = None
    fskatt_verified: Optional[bool] = None
    hourly_rate: Optional[float] = None
    personnummer: Optional[str] = None
    employment_type: Optional[str] = None


# ---- Shifts (Schema) ----
class ShiftCreate(BaseModel):
    employee_id: str
    date: str
    start_time: str
    end_time: str
    title: str = Field(..., min_length=1)
    note: Optional[str] = None
    booking_id: Optional[str] = None


class ShiftUpdate(BaseModel):
    employee_id: Optional[str] = None
    date: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    title: Optional[str] = None
    note: Optional[str] = None
    booking_id: Optional[str] = None


class Shift(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    id: Optional[PyObjectId] = Field(default=None, alias="_id")
    employee_id: str
    date: str
    start_time: str
    end_time: str
    title: str
    note: Optional[str] = None
    booking_id: Optional[str] = None
    created_at: str


# ---- Frånvaro (Absence) ----
class AbsenceCreate(BaseModel):
    employee_id: str
    type: str = Field(..., min_length=1)
    start_date: str
    end_date: str
    note: Optional[str] = None


class AbsenceUpdate(BaseModel):
    employee_id: Optional[str] = None
    type: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    note: Optional[str] = None


class Absence(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    id: Optional[PyObjectId] = Field(default=None, alias="_id")
    employee_id: str
    type: str
    start_date: str
    end_date: str
    note: Optional[str] = None
    created_at: str


# ---- Utlägg (Expenses) ----
class ExpenseCreate(BaseModel):
    employee_id: str
    date: str
    amount: float = Field(..., ge=0)
    antal: Optional[int] = 1
    unit_price: Optional[float] = None
    moms_rate: float = 0.0
    category: str = "Övrigt"
    description: Optional[str] = None
    receipt_image: Optional[str] = None


class ExpenseUpdate(BaseModel):
    employee_id: Optional[str] = None
    date: Optional[str] = None
    amount: Optional[float] = None
    antal: Optional[int] = None
    unit_price: Optional[float] = None
    category: Optional[str] = None
    description: Optional[str] = None
    status: Optional[str] = None
    moms_rate: Optional[int] = None
    receipt_image: Optional[str] = None


class Expense(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    id: Optional[PyObjectId] = Field(default=None, alias="_id")
    employee_id: str
    date: str
    amount: float
    category: str = "Övrigt"
    description: Optional[str] = None
    status: str = "pending"
    created_at: str
    moms_rate: Optional[float] = 0.0
    antal: Optional[int] = 1
    unit_price: Optional[float] = None
    receipt_image: Optional[str] = None


# ---- Payroll settings (OB-tillägg + salary codes for PAXML) ----
class PayrollSettings(BaseModel):
    # OB1: Kväll mån-lör 18:00-24:00 (25.69 kr/tim per Serviceentreprenadavtalet 2025)
    ob1_label: str = "Kväll (mån–lör 18–24)"
    ob1_extra: float = 25.69
    ob1_days: List[int] = Field(default_factory=lambda: [0, 1, 2, 3, 4, 5])
    ob1_start: str = "18:00"
    ob1_end: str = "23:59"
    # OB2: Helg lördag-söndag hela dagen (54.08 kr/tim per Serviceentreprenadavtalet 2025)
    ob2_label: str = "Helg (lör–sön)"
    ob2_extra: float = 54.08
    ob2_days: List[int] = Field(default_factory=lambda: [5, 6])
    ob2_start: str = "00:00"
    ob2_end: str = "23:59"
    code_normal: str = "100"
    code_ob1: str = "210"
    code_ob2: str = "220"
    code_expense: str = "710"
    company_orgnr: Optional[str] = None


async def get_payroll_settings_obj() -> PayrollSettings:
    doc = await db.settings.find_one({"_key": "payroll"})
    if not doc:
        return PayrollSettings()
    doc.pop("_id", None)
    doc.pop("_key", None)
    return PayrollSettings(**doc)


# ---- Invoice settings ----
class InvoiceSettings(BaseModel):
    company_name: str = "PureNorth Städ"
    company_orgnr: Optional[str] = None
    company_address: Optional[str] = None
    company_email: Optional[str] = None
    company_phone: Optional[str] = None
    bankgiro: Optional[str] = None
    plusgiro: Optional[str] = None
    iban: Optional[str] = None
    vat_rate: float = 25.0
    payment_terms_days: int = 30
    next_invoice_number: int = 1001
    company_logo: Optional[str] = None  # base64 encoded image


async def get_invoice_settings_obj() -> InvoiceSettings:
    doc = await db.settings.find_one({"_key": "invoice"})
    if not doc:
        return InvoiceSettings()
    doc.pop("_id", None)
    doc.pop("_key", None)
    return InvoiceSettings(**doc)


async def get_next_invoice_number() -> int:
    doc = await db.settings.find_one({"_key": "invoice"})
    if not doc:
        defaults = InvoiceSettings().model_dump()
        defaults["_key"] = "invoice"
        await db.settings.insert_one(defaults)
        doc = defaults
    number = doc.get("next_invoice_number", 1001)
    await db.settings.update_one({"_key": "invoice"}, {"$set": {"next_invoice_number": number + 1}})
    return number


# ---- Invoices (Fakturering) ----
class InvoiceItemModel(BaseModel):
    description: str = Field(..., min_length=1)
    quantity: float = 1
    unit_price: float = 0
    is_material: bool = False


class InvoiceCreate(BaseModel):
    booking_id: Optional[str] = None
    customer_name: str = Field(..., min_length=1)
    customer_email: Optional[str] = None
    customer_phone: Optional[str] = None
    customer_address: Optional[str] = None
    customer_personnummer: Optional[str] = None
    customer_type: str = "private"
    rut_eligible: bool = True
    items: List[InvoiceItemModel]
    note: Optional[str] = None
    due_date: Optional[str] = None
    due_days: Optional[int] = None
    notes: Optional[str] = None
    subtotal: Optional[float] = None
    rut_deduction: Optional[float] = None
    vat_amount: Optional[float] = None
    total_amount: Optional[float] = None
    customer_pays: Optional[float] = None


class InvoiceStatusUpdate(BaseModel):
    status: str


class Invoice(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    id: Optional[PyObjectId] = Field(default=None, alias="_id")
    invoice_number: int
    booking_id: Optional[str] = None
    customer_name: str
    customer_email: Optional[str] = None
    customer_phone: Optional[str] = None
    customer_address: Optional[str] = None
    customer_personnummer: Optional[str] = None
    customer_type: str = "private"
    rut_eligible: bool = True
    items: List[InvoiceItemModel]
    note: Optional[str] = None
    due_date: str
    status: str = "draft"
    labor_total: float = 0.0
    material_total: float = 0.0
    subtotal: float
    vat_amount: float
    total_amount: float
    rut_deduction: float
    customer_pays: float
    created_at: str
    paid_at: Optional[str] = None
    reminder_count: Optional[int] = 0
    last_reminder_at: Optional[str] = None


def calc_invoice_amounts(items: list, rut_eligible: bool, customer_type: str, vat_rate: float):
    labor_total = sum(i["quantity"] * i["unit_price"] for i in items if not i.get("is_material"))
    material_total = sum(i["quantity"] * i["unit_price"] for i in items if i.get("is_material"))
    subtotal = labor_total + material_total
    vat_amount = subtotal * vat_rate / 100.0
    total_amount = subtotal + vat_amount
    rut_deduction = 0.0
    if rut_eligible and customer_type == "private":
        rut_deduction = round(labor_total * 0.5, 2)
    customer_pays = total_amount - rut_deduction
    return {
        "labor_total": round(labor_total, 2),
        "material_total": round(material_total, 2),
        "subtotal": round(subtotal, 2),
        "vat_amount": round(vat_amount, 2),
        "total_amount": round(total_amount, 2),
        "rut_deduction": rut_deduction,
        "customer_pays": round(customer_pays, 2),
    }


def build_invoice_pdf(inv: dict, settings: InvoiceSettings) -> bytes:
    from reportlab.platypus import Image as RLImage
    from reportlab.lib.utils import ImageReader

    buf = BytesIO()
    page_w, page_h = A4
    inv_num = inv.get("invoice_number", "")
    customer = inv.get("customer_name", "")
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=15*mm, bottomMargin=20*mm, leftMargin=20*mm, rightMargin=20*mm,
        title=f"Faktura #{inv_num} - {customer}",
        author=settings.company_name or "PureNorth Städ",
        subject="Faktura")
    styles = getSampleStyleSheet()

    def ps(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    label_style  = ps("lbl",  fontSize=7,  textColor=colors.HexColor("#888888"), leading=10, spaceAfter=1)
    value_style  = ps("val",  fontSize=9,  textColor=colors.HexColor("#1a1a1a"), leading=12)
    bold_style   = ps("bld",  fontSize=9,  textColor=colors.HexColor("#1a1a1a"), leading=12, fontName="Helvetica-Bold")
    small_style  = ps("sml",  fontSize=7.5,textColor=colors.HexColor("#555555"), leading=10)
    footer_style = ps("ftr",  fontSize=7,  textColor=colors.HexColor("#999999"), leading=9)

    elements = []

    # ── HEADER: logo left, FAKTURA right ────────────────────────────────────
    logo_cell = ""
    if settings.company_logo:
        try:
            logo_data = base64.b64decode(settings.company_logo.split(",")[-1])
            logo_buf = BytesIO(logo_data)
            ir = ImageReader(logo_buf)
            iw, ih = ir.getSize()
            max_h = 16*mm
            ratio = max_h / ih
            logo_buf.seek(0)
            logo_cell = RLImage(logo_buf, width=iw*ratio, height=max_h)
        except:
            logo_cell = Paragraph(settings.company_name or "", ps("cn", fontSize=14, fontName="Helvetica-Bold"))
    else:
        logo_cell = Paragraph(settings.company_name or "", ps("cn", fontSize=14, fontName="Helvetica-Bold"))

    faktura_title = Paragraph(
        '<font size="28" color="#141414"><b>FAKTURA</b></font>',
        ps("ft", alignment=2)
    )
    faktura_nr = Paragraph(
        f'<font size="9" color="#888888">Nr {inv["invoice_number"]}</font>',
        ps("fnr", alignment=2)
    )

    hdr = Table([[logo_cell, [faktura_title, faktura_nr]]], colWidths=[90*mm, None])
    hdr.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("ALIGN",(1,0),(1,0),"RIGHT"),
        ("LEFTPADDING",(0,0),(-1,-1),0),
        ("RIGHTPADDING",(0,0),(-1,-1),0),
        ("TOPPADDING",(0,0),(-1,-1),0),
        ("BOTTOMPADDING",(0,0),(-1,-1),0),
    ]))
    elements.append(hdr)
    elements.append(Spacer(1, 4*mm))

    # ── DIVIDER ──────────────────────────────────────────────────────────────
    elements.append(Table([[""]], colWidths=[page_w - 40*mm]))
    elements[-1].setStyle(TableStyle([
        ("LINEBELOW",(0,0),(-1,-1),1.5,colors.HexColor("#141414")),
        ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0),
    ]))
    elements.append(Spacer(1, 5*mm))

    # ── SENDER + RECIPIENT + INVOICE DETAILS ────────────────────────────────
    def info_block(label, lines):
        block = [Paragraph(label, label_style)]
        for l in lines:
            if l: block.append(Paragraph(str(l), value_style))
        return block

    sender = info_block("FRÅN", [
        settings.company_name,
        settings.company_address,
        settings.company_orgnr and f"Org.nr {settings.company_orgnr}",
        settings.company_email,
        settings.company_phone,
    ])
    recipient = info_block("TILL", [
        inv["customer_name"],
        inv.get("customer_address"),
        inv.get("customer_email"),
        inv.get("customer_phone"),
        inv.get("customer_personnummer") and f"Personnr. {inv['customer_personnummer']}",
    ])
    details = [
        [Paragraph("FAKTURADATUM", label_style), Paragraph(inv["created_at"][:10], value_style)],
        [Paragraph("FÖRFALLODATUM", label_style), Paragraph(inv["due_date"], bold_style)],
        [Paragraph("BETALNINGSVILLKOR", label_style), Paragraph(f"{inv.get('due_days',30)} dagar netto", value_style)],
    ]
    if settings.bankgiro:
        details.append([Paragraph("BANKGIRO", label_style), Paragraph(settings.bankgiro, value_style)])
    if settings.iban:
        details.append([Paragraph("IBAN", label_style), Paragraph(settings.iban, small_style)])

    details_tbl = Table(details, colWidths=[28*mm, 37*mm])
    details_tbl.setStyle(TableStyle([
        ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0),
        ("TOPPADDING",(0,0),(-1,-1),1),("BOTTOMPADDING",(0,0),(-1,-1),1),
    ]))

    top_row = Table([[sender, recipient, details_tbl]], colWidths=[65*mm, 65*mm, 65*mm])
    top_row.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),4),
        ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0),
    ]))
    elements.append(top_row)
    elements.append(Spacer(1, 8*mm))

    # ── ITEMS TABLE ──────────────────────────────────────────────────────────
    HOURLY = ("hemstädning","kontorsstädning","storstädning","byggstädning","trappstädning",
               "hemstadning","kontorstadning","storstadning","byggstadning","trappstadning")
    def is_hourly(desc): return desc.strip().lower() in HOURLY

    col_hdr_style = ParagraphStyle("ch", parent=styles["Normal"], fontSize=8,
        textColor=colors.white, fontName="Helvetica-Bold")
    col_val_style = ParagraphStyle("cv", parent=styles["Normal"], fontSize=9,
        textColor=colors.HexColor("#1a1a1a"))
    col_val_r = ParagraphStyle("cvr", parent=col_val_style, alignment=2)

    data = [[
        Paragraph("BESKRIVNING", col_hdr_style),
        Paragraph("ANT/TIM", col_hdr_style),
        Paragraph("À-PRIS", col_hdr_style),
        Paragraph("BELOPP", col_hdr_style),
    ]]
    for item in inv["items"]:
        line_total = item["quantity"] * item["unit_price"]
        qty_label = f'{item["quantity"]:g} tim' if is_hourly(item["description"]) else f'{item["quantity"]:g}'
        data.append([
            Paragraph(item["description"], col_val_style),
            Paragraph(qty_label, col_val_r),
            Paragraph(f'{item["unit_price"]:,.2f}', col_val_r),
            Paragraph(f'{line_total:,.2f}', col_val_r),
        ])

    items_table = Table(data, colWidths=[85*mm, 22*mm, 30*mm, 30*mm])
    row_count = len(data)
    items_table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0), colors.HexColor("#141414")),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#F8F9FA")]),
        ("FONTSIZE",(0,0),(-1,-1),9),
        ("ALIGN",(1,0),(-1,-1),"RIGHT"),
        ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
        ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LINEBELOW",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
        ("LINEBELOW",(0,0),(-1,0),0,colors.white),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 6*mm))

    # ── TOTALS ───────────────────────────────────────────────────────────────
    totals = [
        ["Delsumma (exkl. moms)", f"{inv['subtotal']:,.2f} kr"],
        [f"Moms ({settings.vat_rate:g}%)", f"{inv['vat_amount']:,.2f} kr"],
        ["Totalt inkl. moms", f"{inv['total_amount']:,.2f} kr"],
    ]
    if inv.get("rut_eligible") and inv.get("rut_deduction", 0) > 0:
        totals.append([f"RUT-avdrag (50% av arbetskostnad)", f"-{inv['rut_deduction']:,.2f} kr"])
        totals.append(["ATT BETALA", f"{inv['customer_pays']:,.2f} kr"])

    tot_styles = []
    last = len(totals) - 1
    for i, row in enumerate(totals):
        is_last = i == last
        totals[i] = [
            Paragraph(row[0], bold_style if is_last else small_style),
            Paragraph(row[1], bold_style if is_last else small_style),
        ]
        if is_last:
            tot_styles += [
                ("LINEABOVE",(0,i),(-1,i),1.2,colors.HexColor("#141414")),
                ("BACKGROUND",(0,i),(-1,i),colors.HexColor("#F0F0F0")),
                ("TOPPADDING",(0,i),(-1,i),6),("BOTTOMPADDING",(0,i),(-1,i),6),
            ]

    totals_table = Table(totals, colWidths=[None, 40*mm], hAlign="RIGHT")
    totals_table.setStyle(TableStyle([
        ("ALIGN",(1,0),(-1,-1),"RIGHT"),
        ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LINEBELOW",(0,0),(-1,-2),0.3,colors.HexColor("#E2E8F0")),
    ] + tot_styles))
    elements.append(totals_table)

    # ── NOTES ────────────────────────────────────────────────────────────────
    if inv.get("notes"):
        elements.append(Spacer(1, 6*mm))
        elements.append(Paragraph("Meddelande", label_style))
        elements.append(Paragraph(inv["notes"], small_style))

    # ── FOOTER ───────────────────────────────────────────────────────────────
    elements.append(Spacer(1, 10*mm))
    footer_parts = []
    if settings.company_name: footer_parts.append(settings.company_name)
    if settings.company_orgnr: footer_parts.append(f"Org.nr {settings.company_orgnr}")
    if settings.bankgiro: footer_parts.append(f"Bankgiro {settings.bankgiro}")
    if settings.company_email: footer_parts.append(settings.company_email)
    footer_line = "  ·  ".join(p for p in footer_parts if p)
    if footer_line:
        ft = Table([[Paragraph(footer_line, footer_style)]], colWidths=[page_w-40*mm])
        ft.setStyle(TableStyle([
            ("LINEABOVE",(0,0),(-1,-1),0.5,colors.HexColor("#CCCCCC")),
            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),0),
            ("LEFTPADDING",(0,0),(-1,-1),0),
        ]))
        elements.append(ft)

    if inv.get("note"):
        elements.append(Spacer(1, 6 * mm))
        elements.append(Paragraph(inv["note"], styles["Normal"]))

    doc.build(elements)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Payroll calculation helpers
# ---------------------------------------------------------------------------
def _time_to_minutes(t: str) -> int:
    h, m = t.split(":")
    h = int(h)
    m = int(m)
    if h >= 24:
        return 24 * 60
    # treat 23:59 as end of day (same as 24:00)
    if h == 23 and m == 59:
        return 24 * 60
    return h * 60 + m


def _overlap_minutes(s1, e1, s2, e2) -> int:
    return max(0, min(e1, e2) - max(s1, s2))


def split_shift_hours(date_str: str, start_str: str, end_str: str, settings: PayrollSettings):
    y, m, d = map(int, date_str.split("-"))
    weekday = DateClass(y, m, d).weekday()

    start_m = _time_to_minutes(start_str)
    end_m = _time_to_minutes(end_str)
    if end_m <= start_m:
        end_m += 24 * 60
    total_minutes = end_m - start_m

    ob1_minutes = 0
    ob2_minutes = 0
    if weekday in settings.ob1_days:
        ob1_minutes = _overlap_minutes(start_m, end_m, _time_to_minutes(settings.ob1_start), _time_to_minutes(settings.ob1_end))
    if weekday in settings.ob2_days:
        ob2_minutes = _overlap_minutes(start_m, end_m, _time_to_minutes(settings.ob2_start), _time_to_minutes(settings.ob2_end))

    combined = ob1_minutes + ob2_minutes
    if combined > total_minutes:
        excess = combined - total_minutes
        reduce1 = min(ob1_minutes, excess)
        ob1_minutes -= reduce1

    normal_minutes = max(0, total_minutes - ob1_minutes - ob2_minutes)
    return normal_minutes / 60.0, ob1_minutes / 60.0, ob2_minutes / 60.0


async def build_payroll_summary(start: str, end: str):
    settings = await get_payroll_settings_obj()
    employees = await db.employees.find().to_list(1000)
    shifts = await db.shifts.find({"date": {"$gte": start, "$lte": end}}).to_list(5000)
    absences = await db.absences.find({"start_date": {"$lte": end}, "end_date": {"$gte": start}}).to_list(2000)
    expenses = await db.expenses.find({"date": {"$gte": start, "$lte": end}}).to_list(2000)

    range_start = DateClass(*map(int, start.split("-")))
    range_end = DateClass(*map(int, end.split("-")))

    summary = {}
    absence_dates_by_employee = {}
    for emp in employees:
        eid = str(emp["_id"])
        summary[eid] = {
            "name": emp["name"],
            "personnummer": emp.get("personnummer") or "",
            "hourly_rate": emp.get("hourly_rate", 0) or 0,
            "employment_type": emp.get("employment_type", "fastanstalld"),
            "normal_h": 0.0,
            "ob1_h": 0.0,
            "ob2_h": 0.0,
            "expense_total": 0.0,
            "absence_days": 0,
            "absence_scheduled_days": 0,
            "absence_lost_amount": 0.0,
        }
        absence_dates_by_employee[eid] = set()

    for a in absences:
        eid = a["employee_id"]
        if eid not in summary:
            continue
        sd = DateClass(*map(int, a["start_date"].split("-")))
        ed = DateClass(*map(int, a["end_date"].split("-")))
        clip_s = max(sd, range_start)
        clip_e = min(ed, range_end)
        if clip_e >= clip_s:
            summary[eid]["absence_days"] += (clip_e - clip_s).days + 1
            d = clip_s
            while d <= clip_e:
                absence_dates_by_employee[eid].add(d.isoformat())
                d += timedelta(days=1)

    overlap_dates_by_employee = {eid: set() for eid in summary}

    for s in shifts:
        eid = s["employee_id"]
        if eid not in summary:
            continue
        n, o1, o2 = split_shift_hours(s["date"], s["start_time"], s["end_time"], settings)
        row = summary[eid]
        if s["date"] in absence_dates_by_employee.get(eid, set()):
            rate = row["hourly_rate"]
            row["absence_lost_amount"] += n * rate + o1 * settings.ob1_extra + o2 * settings.ob2_extra
            overlap_dates_by_employee[eid].add(s["date"])
        else:
            row["normal_h"] += n
            row["ob1_h"] += o1
            row["ob2_h"] += o2

    for eid, dates in overlap_dates_by_employee.items():
        summary[eid]["absence_scheduled_days"] = len(dates)

    for e in expenses:
        eid = e["employee_id"]
        if eid in summary:
            summary[eid]["expense_total"] += e["amount"]

    # ── Sjuklön calculation (only for "Sjuk" absences) ──────────────────
    # Rules: sjuklön = 80% of scheduled shift pay
    #        karensavdrag = 20% of one average week's sjuklön
    for eid, row in summary.items():
        row["sick_shifts_h"] = 0.0
        row["sjuklon_gross"] = 0.0
        row["karensavdrag"] = 0.0
        row["sjuklon_net"] = 0.0

    # Get sick absences only
    sick_absences = [a for a in absences if a.get("type", "").lower() == "sjuk"]
    sick_dates_by_employee = {}
    for a in sick_absences:
        eid = a["employee_id"]
        if eid not in summary:
            continue
        sd = DateClass(*map(int, a["start_date"].split("-")))
        ed = DateClass(*map(int, a["end_date"].split("-")))
        clip_s = max(sd, range_start)
        clip_e = min(ed, range_end)
        if eid not in sick_dates_by_employee:
            sick_dates_by_employee[eid] = set()
        d = clip_s
        while d <= clip_e:
            sick_dates_by_employee[eid].add(d.isoformat())
            d += timedelta(days=1)

    # Calculate sjuklön for sick shifts
    for s in shifts:
        eid = s["employee_id"]
        if eid not in summary:
            continue
        if s["date"] not in sick_dates_by_employee.get(eid, set()):
            continue
        n, o1, o2 = split_shift_hours(s["date"], s["start_time"], s["end_time"], settings)
        total_h = n + o1 + o2
        rate = summary[eid]["hourly_rate"]
        summary[eid]["sick_shifts_h"] += total_h
        summary[eid]["sjuklon_gross"] += total_h * rate * 0.8

    for eid, row in summary.items():
        if row["sjuklon_gross"] > 0:
            # karensavdrag = 20% of one week's sjuklön
            # one week avg = sjuklon_gross / sick_weeks (min 1)
            sick_days = len(sick_dates_by_employee.get(eid, set()))
            sick_weeks = max(1, sick_days / 5.0)
            weekly_sjuklon = row["sjuklon_gross"] / sick_weeks
            row["karensavdrag"] = round(weekly_sjuklon * 0.20, 2)
            row["sjuklon_net"] = round(max(0, row["sjuklon_gross"] - row["karensavdrag"]), 2)

    for eid, row in summary.items():
        rate = row["hourly_rate"]
        base_pay = row["normal_h"] * rate
        ob1_pay = row["ob1_h"] * settings.ob1_extra
        ob2_pay = row["ob2_h"] * settings.ob2_extra
        row["base_pay"] = base_pay
        row["ob1_pay"] = ob1_pay
        row["ob2_pay"] = ob2_pay
        row["total_pay"] = base_pay + ob1_pay + ob2_pay + row["expense_total"] + row["sjuklon_net"]
        # Nettolön = bruttolön - prelskatt (30%)
        brutto = base_pay + ob1_pay + ob2_pay + row["sjuklon_net"]
        row["brutto_pay"] = round(brutto, 2)
        row["prelskatt"] = round(brutto * 0.30, 2)
        row["netto_pay"] = round(brutto - brutto * 0.30 + row["expense_total"], 2)

    return summary, settings


def build_xlsx_bytes(summary: dict) -> bytes:
    PRELSKATT = 0.30
    ARBETSGIVARAVGIFT = 0.3142
    SEMESTER = 0.12
    wb = Workbook()
    ws = wb.active
    ws.title = "Löneunderlag"
    headers = [
        "Anställd", "Typ", "Personnummer", "Lön/h (kr)",
        "Normaltid (h)", "OB1 (h)", "OB2 (h)", "OB-tillägg (kr)",
        "Grundlön (kr)", "Semesterersättning 12% (kr)", "Arbetsgivaravgift 31.42% (kr)",
        "Bruttolön (kr)", "Prelskatt 30% (kr)", "Nettolön (kr)",
        "Utlägg (kr)", "Frånvarodagar", "Förlorad lön (kr)",
        "Sjuklön netto (kr)", "Total kostnad arbetsgivare (kr)",
    ]
    ws.append(headers)
    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="141414", end_color="141414", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    total_row = ["TOTALT", "", "", "", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    for row in summary.values():
        type_label = "Vikarie" if row.get("employment_type") == "vikarie" else "Fast anställd"
        brutto = round(row["normal_h"]*row["hourly_rate"] + row["ob1_h"]*row.get("ob1_extra",0) + row["ob2_h"]*row.get("ob2_extra",0) + row.get("sjuklon_net",0), 2)
        ob_kr = round(row["ob1_h"]*row.get("ob1_extra",0) + row["ob2_h"]*row.get("ob2_extra",0), 2)
        grundlon = round(row["normal_h"]*row["hourly_rate"], 2)
        semester = round(brutto * SEMESTER, 2)
        ag_avg = round(brutto * ARBETSGIVARAVGIFT, 2)
        prelskatt = round(brutto * PRELSKATT, 2)
        netto = round(brutto - prelskatt, 2)
        utlagg = round(row.get("expense_total", 0), 2)
        total_cost = round(brutto + ag_avg + semester, 2)

        ws.append([
            row["name"], type_label, row.get("personnummer",""),
            row["hourly_rate"],
            round(row["normal_h"], 2), round(row["ob1_h"], 2), round(row["ob2_h"], 2), ob_kr,
            grundlon, semester, ag_avg, brutto, -prelskatt, netto,
            utlagg, row["absence_days"], round(row.get("absence_lost_amount",0), 2),
            round(row.get("sjuklon_net",0), 2), total_cost,
        ])
        for j, val in enumerate([4,5,6,7,8,9,10,11,12,13,14,16,17,18]):
            total_row[val] = round(total_row[val] + [grundlon,round(row["normal_h"],2),round(row["ob1_h"],2),round(row["ob2_h"],2),ob_kr,semester,ag_avg,brutto,-prelskatt,netto,utlagg,round(row.get("absence_lost_amount",0),2),round(row.get("sjuklon_net",0),2),total_cost][j], 2)

    ws.append(total_row)
    # Style totals row
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, min(max_len + 2, 25))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_paxml_str(summary: dict, settings: PayrollSettings, start: str, end: str) -> str:
    lines = ['<?xml version="1.0" encoding="UTF-8"?>']
    lines.append('<paxml version="2.0" xmlns="http://www.paxml.se/2.0">')
    lines.append('  <header>')
    lines.append('    <format>LONETRANS</format>')
    if settings.company_orgnr:
        lines.append(f'    <orgnr>{quoteattr(settings.company_orgnr)[1:-1]}</orgnr>')
    lines.append(f'    <skapad>{datetime.now(timezone.utc).isoformat()}</skapad>')
    lines.append(f'    <period start="{start}" slut="{end}"/>')
    lines.append('  </header>')
    lines.append('  <lonetransaktioner>')
    for eid, row in summary.items():
        persnr = quoteattr(row["personnummer"] or "")
        if row["normal_h"]:
            lines.append(f'    <lonetrans anstid={quoteattr(eid)} persnr={persnr} kod="{settings.code_normal}" antal="{round(row["normal_h"], 2)}" enhet="TIM"/>')
        if row["ob1_h"]:
            lines.append(f'    <lonetrans anstid={quoteattr(eid)} persnr={persnr} kod="{settings.code_ob1}" antal="{round(row["ob1_h"], 2)}" enhet="TIM"/>')
        if row["ob2_h"]:
            lines.append(f'    <lonetrans anstid={quoteattr(eid)} persnr={persnr} kod="{settings.code_ob2}" antal="{round(row["ob2_h"], 2)}" enhet="TIM"/>')
        if row["expense_total"]:
            lines.append(f'    <lonetrans anstid={quoteattr(eid)} persnr={persnr} kod="{settings.code_expense}" belopp="{round(row["expense_total"], 2)}" enhet="KR"/>')
    lines.append('  </lonetransaktioner>')
    lines.append('</paxml>')
    return "\n".join(lines)



# ---- Price List ----
class PriceItem(BaseModel):
    id: str
    service: str
    description: Optional[str] = None
    unit: str = "tim"  # "tim", "kvm", "st", "fast"
    price: float = 0
    is_rut_eligible: bool = True
    is_active: bool = True


class PriceListSettings(BaseModel):
    items: List[PriceItem] = Field(default_factory=list)

# ---------------------------------------------------------------------------
# App / Router
# ---------------------------------------------------------------------------
app = FastAPI()
api_router = APIRouter(prefix="/api")


@api_router.get("/")
async def root():
    return {"message": "PureNorth Städ API"}


# ---- Auth ----
@api_router.post("/auth/login")
async def login(payload: LoginRequest):
    email = payload.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Fel e-post eller lösenord")
    token = create_access_token(str(user["_id"]), email)
    return {
        "token": token,
        "user": {"email": user["email"], "name": user.get("name", "Admin"), "role": user.get("role", "admin")},
    }


@api_router.get("/auth/me")
async def me(current=Depends(get_current_user)):
    return {"email": current["email"], "name": current.get("name", "Admin"), "role": current.get("role", "admin")}


# ---- Bookings ----
@api_router.post("/bookings", response_model=Booking, response_model_by_alias=False)
async def create_booking(payload: BookingCreate):
    doc = payload.model_dump()
    doc["status"] = "new"
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.bookings.insert_one(doc)
    doc["_id"] = str(result.inserted_id)
    return Booking(**doc)


@api_router.get("/bookings", response_model=List[Booking], response_model_by_alias=False)
async def list_bookings(current=Depends(get_current_user)):
    docs = await db.bookings.find().sort("created_at", -1).to_list(1000)
    return [Booking(**{**d, "_id": str(d["_id"])}) for d in docs]


@api_router.patch("/bookings/{booking_id}/status")
async def update_status(booking_id: str, payload: StatusUpdate, current=Depends(get_current_user)):
    result = await db.bookings.update_one(
        {"_id": to_object_id(booking_id)}, {"$set": {"status": payload.status}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Bokning hittades inte")
    # Send confirmation email when status changes to "contacted"
    if payload.status == "contacted":
        doc = await db.bookings.find_one({"_id": to_object_id(booking_id)})
        if doc:
            doc["_id"] = str(doc["_id"])
            inv_settings = await get_invoice_settings_obj()
            await send_booking_confirmation(doc, inv_settings)
    return {"success": True}


@api_router.patch("/bookings/{booking_id}")
async def update_booking(booking_id: str, payload: dict, current=Depends(get_current_user)):
    updates = {k: v for k, v in payload.items() if k not in ["_id", "id", "created_at"]}
    await db.bookings.update_one({"_id": to_object_id(booking_id)}, {"$set": updates})
    doc = await db.bookings.find_one({"_id": to_object_id(booking_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Bokning hittades inte")
    doc["_id"] = str(doc["_id"])
    return Booking(**doc)


@api_router.delete("/bookings/{booking_id}")
async def delete_booking(booking_id: str, current=Depends(get_current_user)):
    result = await db.bookings.delete_one({"_id": to_object_id(booking_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bokning hittades inte")
    return {"success": True}


# ---- Reviews ----
@api_router.post("/reviews", response_model=Review, response_model_by_alias=False)
async def create_review(payload: ReviewCreate):
    doc = payload.model_dump()
    doc["approved"] = False
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.reviews.insert_one(doc)
    doc["_id"] = str(result.inserted_id)
    return Review(**doc)


@api_router.get("/reviews/approved", response_model=List[Review], response_model_by_alias=False)
async def list_approved_reviews():
    docs = await db.reviews.find({"approved": True}).sort("created_at", -1).to_list(1000)
    return [Review(**{**d, "_id": str(d["_id"])}) for d in docs]


@api_router.get("/reviews", response_model=List[Review], response_model_by_alias=False)
async def list_reviews(current=Depends(get_current_user)):
    docs = await db.reviews.find().sort("created_at", -1).to_list(1000)
    return [Review(**{**d, "_id": str(d["_id"])}) for d in docs]


@api_router.patch("/reviews/{review_id}/approve")
async def approve_review(review_id: str, payload: ApproveUpdate, current=Depends(get_current_user)):
    result = await db.reviews.update_one(
        {"_id": to_object_id(review_id)}, {"$set": {"approved": payload.approved}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Omdöme hittades inte")
    return {"success": True}


@api_router.delete("/reviews/{review_id}")
async def delete_review(review_id: str, current=Depends(get_current_user)):
    result = await db.reviews.delete_one({"_id": to_object_id(review_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Omdöme hittades inte")
    return {"success": True}


# ---- Employees ----
@api_router.post("/employees", response_model=Employee, response_model_by_alias=False)
async def create_employee(payload: EmployeeCreate, current=Depends(get_current_user)):
    doc = payload.model_dump()
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.employees.insert_one(doc)
    doc["_id"] = str(result.inserted_id)
    return Employee(**doc)


@api_router.get("/employees", response_model=List[Employee], response_model_by_alias=False)
async def list_employees(current=Depends(get_current_user)):
    docs = await db.employees.find().sort("created_at", 1).to_list(1000)
    return [Employee(**{**d, "_id": str(d["_id"])}) for d in docs]


@api_router.patch("/employees/{employee_id}", response_model=Employee, response_model_by_alias=False)
async def update_employee(employee_id: str, payload: EmployeeUpdate, current=Depends(get_current_user)):
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    if updates:
        result = await db.employees.update_one({"_id": to_object_id(employee_id)}, {"$set": updates})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Anställd hittades inte")
    doc = await db.employees.find_one({"_id": to_object_id(employee_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Anställd hittades inte")
    doc["_id"] = str(doc["_id"])
    return Employee(**doc)


@api_router.delete("/employees/{employee_id}")
async def delete_employee(employee_id: str, current=Depends(get_current_user)):
    result = await db.employees.delete_one({"_id": to_object_id(employee_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Anställd hittades inte")
    await db.shifts.delete_many({"employee_id": employee_id})
    return {"success": True}


# ---- Shifts (Schema) ----
@api_router.post("/shifts", response_model=Shift, response_model_by_alias=False)
async def create_shift(payload: ShiftCreate, current=Depends(get_current_user)):
    doc = payload.model_dump()
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.shifts.insert_one(doc)
    doc["_id"] = str(result.inserted_id)
    return Shift(**doc)


@api_router.get("/shifts", response_model=List[Shift], response_model_by_alias=False)
async def list_shifts(start: Optional[str] = None, end: Optional[str] = None, current=Depends(get_current_user)):
    query = {}
    if start and end:
        query["date"] = {"$gte": start, "$lte": end}
    docs = await db.shifts.find(query).sort("date", 1).to_list(2000)
    return [Shift(**{**d, "_id": str(d["_id"])}) for d in docs]


@api_router.patch("/shifts/{shift_id}", response_model=Shift, response_model_by_alias=False)
async def update_shift(shift_id: str, payload: ShiftUpdate, current=Depends(get_current_user)):
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    if updates:
        result = await db.shifts.update_one({"_id": to_object_id(shift_id)}, {"$set": updates})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Pass hittades inte")
    doc = await db.shifts.find_one({"_id": to_object_id(shift_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Pass hittades inte")
    doc["_id"] = str(doc["_id"])
    return Shift(**doc)


@api_router.delete("/shifts/{shift_id}")
async def delete_shift(shift_id: str, current=Depends(get_current_user)):
    result = await db.shifts.delete_one({"_id": to_object_id(shift_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pass hittades inte")
    return {"success": True}


# ---- Frånvaro (Absence) ----
@api_router.post("/absences", response_model=Absence, response_model_by_alias=False)
async def create_absence(payload: AbsenceCreate, current=Depends(get_current_user)):
    doc = payload.model_dump()
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.absences.insert_one(doc)
    doc["_id"] = str(result.inserted_id)
    return Absence(**doc)


@api_router.get("/absences", response_model=List[Absence], response_model_by_alias=False)
async def list_absences(start: Optional[str] = None, end: Optional[str] = None, current=Depends(get_current_user)):
    query = {}
    if start and end:
        query["start_date"] = {"$lte": end}
        query["end_date"] = {"$gte": start}
    docs = await db.absences.find(query).sort("start_date", -1).to_list(2000)
    return [Absence(**{**d, "_id": str(d["_id"])}) for d in docs]


@api_router.patch("/absences/{absence_id}", response_model=Absence, response_model_by_alias=False)
async def update_absence(absence_id: str, payload: AbsenceUpdate, current=Depends(get_current_user)):
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    if updates:
        result = await db.absences.update_one({"_id": to_object_id(absence_id)}, {"$set": updates})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Frånvaro hittades inte")
    doc = await db.absences.find_one({"_id": to_object_id(absence_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Frånvaro hittades inte")
    doc["_id"] = str(doc["_id"])
    return Absence(**doc)


@api_router.delete("/absences/{absence_id}")
async def delete_absence(absence_id: str, current=Depends(get_current_user)):
    result = await db.absences.delete_one({"_id": to_object_id(absence_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Frånvaro hittades inte")
    return {"success": True}


# ---- Utlägg (Expenses) ----
@api_router.post("/expenses", response_model=Expense, response_model_by_alias=False)
async def create_expense(payload: ExpenseCreate, current=Depends(get_current_user)):
    doc = payload.model_dump()
    doc["status"] = "pending"
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.expenses.insert_one(doc)
    doc["_id"] = str(result.inserted_id)
    return Expense(**doc)


@api_router.get("/expenses", response_model=List[Expense], response_model_by_alias=False)
async def list_expenses(start: Optional[str] = None, end: Optional[str] = None, current=Depends(get_current_user)):
    query = {}
    if start and end:
        query["date"] = {"$gte": start, "$lte": end}
    docs = await db.expenses.find(query).sort("date", -1).to_list(2000)
    return [Expense(**{**d, "_id": str(d["_id"])}) for d in docs]


@api_router.patch("/expenses/{expense_id}", response_model=Expense, response_model_by_alias=False)
async def update_expense(expense_id: str, payload: ExpenseUpdate, current=Depends(get_current_user)):
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    if updates:
        result = await db.expenses.update_one({"_id": to_object_id(expense_id)}, {"$set": updates})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Utlägg hittades inte")
    doc = await db.expenses.find_one({"_id": to_object_id(expense_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Utlägg hittades inte")
    doc["_id"] = str(doc["_id"])
    return Expense(**doc)


@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str, current=Depends(get_current_user)):
    result = await db.expenses.delete_one({"_id": to_object_id(expense_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Utlägg hittades inte")
    return {"success": True}


# ---- Payroll settings ----
@api_router.get("/settings/payroll", response_model=PayrollSettings)
async def get_payroll_settings(current=Depends(get_current_user)):
    return await get_payroll_settings_obj()


@api_router.put("/settings/payroll", response_model=PayrollSettings)
async def set_payroll_settings(payload: PayrollSettings, current=Depends(get_current_user)):
    doc = payload.model_dump()
    doc["_key"] = "payroll"
    await db.settings.update_one({"_key": "payroll"}, {"$set": doc}, upsert=True)
    return payload


# ---- Payroll summary + export ----
@api_router.get("/payroll/summary")
async def payroll_summary(start: str, end: str, current=Depends(get_current_user)):
    summary, settings = await build_payroll_summary(start, end)
    return {
        "settings": settings.model_dump(),
        "rows": [{"employee_id": eid, **row} for eid, row in summary.items()],
    }


@api_router.get("/payroll/export")
async def payroll_export(start: str, end: str, format: str = "xlsx", current=Depends(get_current_user)):
    summary, settings = await build_payroll_summary(start, end)
    if format == "pdf":
        inv_settings = await get_invoice_settings_obj()
        company = inv_settings.company_name or "PureNorth Städ"
        buf = BytesIO()
        from reportlab.lib.pagesizes import landscape
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
            topMargin=12*mm, bottomMargin=12*mm, leftMargin=15*mm, rightMargin=15*mm,
            title=f"Lönesammanställning {start} - {end}",
            author="PureNorth Städ")
        styles = getSampleStyleSheet()
        elements = []
        def ps(name, **kw): return ParagraphStyle(name, parent=styles["Normal"], **kw)
        # Header
        hdr = Table([[
            ParagraphStyle and Paragraph(company, ps("h", fontSize=16, fontName="Helvetica-Bold", textColor=colors.white)),
            Paragraph(f"LÖNESAMMANSTÄLLNING<br/>{start} – {end}", ps("s", fontSize=10, fontName="Helvetica-Bold", textColor=colors.white, alignment=2))
        ]], colWidths=[100*mm, None])
        hdr.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#141414")),
            ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),
            ("TOPPADDING",(0,0),(-1,-1),10),("BOTTOMPADDING",(0,0),(-1,-1),10),
        ]))
        # Header block
        hdr_data = [[
            Paragraph(f"<b>{company}</b>", ps("ch", fontSize=16, fontName="Helvetica-Bold", textColor=colors.white)),
            Paragraph(f"<b>LÖNESAMMANSTÄLLNING</b><br/>{start} – {end}", ps("ct", fontSize=10, fontName="Helvetica-Bold", textColor=colors.white, alignment=2))
        ]]
        hdr_tbl = Table(hdr_data, colWidths=[100*mm, None])
        hdr_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#141414")),
            ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),
            ("TOPPADDING",(0,0),(-1,-1),12),("BOTTOMPADDING",(0,0),(-1,-1),12),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ]))
        elements.append(hdr_tbl)
        elements.append(Paragraph(
            f"Skapad: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC",
            ps("d", fontSize=8, textColor=colors.HexColor("#94a3b8"), spaceBefore=4, spaceAfter=8)
        ))
        PRELSKATT = 0.30
        # Table
        PRELSKATT = 0.30
        ARBETSGIVARAVGIFT = 0.3142
        SEMESTER = 0.12
        data = [["Namn", "Typ", "Lön/h", "Normal(h)", "OB(kr)", "Brutto", "Semester", "AG-avgift", "Prelskatt", "Nettolön", "Utlägg", "Frånvaro", "Förlorad lön", "Sjuklön"]]
        total_brutto = total_netto = total_utlagg = total_ag = total_sem = total_prel = 0
        for r in summary.values():
            brutto = r["normal_h"]*r["hourly_rate"] + r["ob1_h"]*settings.ob1_extra + r["ob2_h"]*settings.ob2_extra + r.get("sjuklon_net",0)
            ob_kr = r["ob1_h"]*settings.ob1_extra + r["ob2_h"]*settings.ob2_extra
            prelskatt = round(brutto * PRELSKATT, 2)
            netto = round(brutto - prelskatt, 2)
            ag = round(brutto * ARBETSGIVARAVGIFT, 2)
            sem = round(brutto * SEMESTER, 2)
            utlagg = r.get("expense_total", 0)
            franvaro = r.get("absence_days", 0)
            forlorad = round(r.get("absence_lost_amount", 0), 2)
            sjuklon = round(r.get("sjuklon_net", 0), 2)
            typ = "Vikarie" if r.get("employment_type") == "vikarie" else "Fast"
            total_brutto += brutto; total_netto += netto; total_utlagg += utlagg
            total_ag += ag; total_sem += sem; total_prel += prelskatt
            data.append([
                r["name"], typ, f'{r["hourly_rate"]:.0f} kr', f'{r["normal_h"]:.1f}',
                f'{ob_kr:.2f}', f'{brutto:.2f}', f'{sem:.2f}',
                f'{ag:.2f}', f'-{prelskatt:.2f}', f'{netto:.2f}', f'{utlagg:.2f}',
                str(franvaro), f'-{forlorad:.2f}' if forlorad > 0 else "-",
                f'{sjuklon:.2f}' if sjuklon > 0 else "-"
            ])
        data.append(["TOTALT", "", "", "", "", f'{total_brutto:.2f}', f'{total_sem:.2f}',
                     f'{total_ag:.2f}', f'-{total_prel:.2f}', f'{total_netto:.2f}', f'{total_utlagg:.2f}', "", "", ""])
        tbl = Table(data, colWidths=[32*mm, 14*mm, 14*mm, 16*mm, 15*mm, 19*mm, 18*mm, 19*mm, 17*mm, 19*mm, 15*mm, 14*mm, 18*mm, 15*mm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#141414")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTSIZE",(0,0),(-1,-1),7),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
            ("LEFTPADDING",(0,0),(-1,-1),2),("RIGHTPADDING",(0,0),(-1,-1),2),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("ALIGN",(2,0),(-1,-1),"RIGHT"),
            ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#F1F5F9")),
            ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
        ]))
        elements.append(tbl)
        doc.build(elements)
        return Response(content=buf.getvalue(), media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="lon_{start}_{end}.pdf"'})
    if format == "paxml":
        xml_str = build_paxml_str(summary, settings, start, end)
        return Response(
            content=xml_str,
            media_type="application/xml",
            headers={"Content-Disposition": f'attachment; filename="lon_{start}_{end}.paxml.xml"'},
        )
    xlsx_bytes = build_xlsx_bytes(summary)
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="lon_{start}_{end}.xlsx"'},
    )


# ---- Invoices (Fakturering) ----
@api_router.get("/settings/invoice", response_model=InvoiceSettings)
async def get_invoice_settings(current=Depends(get_current_user)):
    return await get_invoice_settings_obj()


@api_router.put("/settings/invoice", response_model=InvoiceSettings)
async def set_invoice_settings(payload: InvoiceSettings, current=Depends(get_current_user)):
    doc = payload.model_dump()
    doc["_key"] = "invoice"
    # preserve next_invoice_number and logo if not provided
    existing = await db.settings.find_one({"_key": "invoice"})
    if existing:
        if doc.get("company_logo") is None and existing.get("company_logo"):
            doc["company_logo"] = existing["company_logo"]
        if doc.get("next_invoice_number", 1001) == 1001 and existing.get("next_invoice_number", 1001) > 1001:
            doc["next_invoice_number"] = existing["next_invoice_number"]
    await db.settings.update_one({"_key": "invoice"}, {"$set": doc}, upsert=True)
    result = await db.settings.find_one({"_key": "invoice"})
    result.pop("_id", None)
    result.pop("_key", None)
    return InvoiceSettings(**result)


@api_router.put("/settings/invoice/logo")
async def set_invoice_logo(request: Request, current=Depends(get_current_user)):
    body = await request.json()
    logo = body.get("company_logo")
    await db.settings.update_one({"_key": "invoice"}, {"$set": {"company_logo": logo}}, upsert=True)
    return {"success": True}


@api_router.post("/invoices", response_model=Invoice, response_model_by_alias=False)
async def create_invoice(payload: InvoiceCreate, current=Depends(get_current_user)):
    inv_settings = await get_invoice_settings_obj()
    items = [i.model_dump() for i in payload.items]
    if not payload.subtotal:
        amounts = calc_invoice_amounts(items, payload.rut_eligible, payload.customer_type, inv_settings.vat_rate)
    else:
        amounts = {}
    number = await get_next_invoice_number()
    due_date = payload.due_date or (DateClass.today() + timedelta(days=inv_settings.payment_terms_days)).isoformat()

    doc = payload.model_dump()
    doc["items"] = items
    doc["invoice_number"] = number
    doc["due_date"] = due_date
    doc["status"] = "draft"
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["paid_at"] = None
    # Use frontend amounts if provided, otherwise recalculate
    if not payload.subtotal:
        doc.update(amounts)

    result = await db.invoices.insert_one(doc)
    doc["_id"] = str(result.inserted_id)
    return Invoice(**doc)


@api_router.get("/invoices", response_model=List[Invoice], response_model_by_alias=False)
async def list_invoices(current=Depends(get_current_user)):
    docs = await db.invoices.find().sort("invoice_number", -1).to_list(2000)
    return [Invoice(**{**d, "_id": str(d["_id"])}) for d in docs]



# ── Auto Reminder Cron ────────────────────────────────────────────────────────
@api_router.post("/invoices/auto-remind")
async def auto_remind_overdue(request: Request):
    """
    Cron endpoint - call daily to auto-send reminders.
    Protected by CRON_SECRET env var.
    Logic:
      - 1 dag efter due_date        → Påminnelse #1 (ingen avgift)
      - payment_terms_days efter #1 → Påminnelse #2 (60 kr avgift)
      - payment_terms_days efter #2 → Påminnelse #3 (inkasso-varning)
    """
    # Verify secret
    secret = os.environ.get("CRON_SECRET", "")
    auth = request.headers.get("x-cron-secret", "")
    if secret and auth != secret:
        raise HTTPException(status_code=401, detail="Unauthorized")

    inv_settings = await get_invoice_settings_obj()
    payment_days = inv_settings.payment_terms_days or 30
    today = datetime.now(timezone.utc).date()
    today_str = today.isoformat()

    # Find all unpaid invoices (sent or overdue) with email
    invoices = await db.invoices.find({
        "status": {"$in": ["sent", "overdue"]},
        "customer_email": {"$exists": True, "$ne": ""}
    }).to_list(1000)

    sent_count = 0
    skipped_count = 0
    results = []

    for inv in invoices:
        due_date_str = inv.get("due_date", "")
        if not due_date_str:
            continue

        try:
            from datetime import date as DateObj
            due = DateObj.fromisoformat(due_date_str)
        except:
            continue

        days_overdue = (today - due).days
        if days_overdue < 1:
            # Not yet overdue
            skipped_count += 1
            continue

        reminder_count = inv.get("reminder_count", 0)
        last_reminder = inv.get("last_reminder_at", "")

        # Calculate when next reminder should be sent
        if reminder_count == 0:
            # First reminder: 1 day after due_date
            should_remind = days_overdue >= 1
        elif reminder_count == 1:
            # Second reminder: payment_days after first reminder
            if last_reminder:
                try:
                    last_dt = DateObj.fromisoformat(last_reminder[:10])
                    should_remind = (today - last_dt).days >= payment_days
                except:
                    should_remind = False
            else:
                should_remind = days_overdue >= payment_days
        elif reminder_count == 2:
            # Third reminder: payment_days after second reminder
            if last_reminder:
                try:
                    last_dt = DateObj.fromisoformat(last_reminder[:10])
                    should_remind = (today - last_dt).days >= payment_days
                except:
                    should_remind = False
            else:
                should_remind = days_overdue >= payment_days * 2
        else:
            # Already sent 3+ reminders, skip
            skipped_count += 1
            continue

        if not should_remind:
            skipped_count += 1
            continue

        # Send reminder using existing logic
        invoice_id = str(inv["_id"])
        company = inv_settings.company_name or "PureNorth Stad"
        inv_num = inv.get("invoice_number", "")
        customer_name = inv.get("customer_name", "")
        customer_email = inv.get("customer_email", "")
        due_date = inv.get("due_date", "")
        amount = inv.get("customer_pays", 0)
        new_reminder_count = reminder_count + 1
        reminder_fee = 60 if new_reminder_count >= 2 else 0
        total_with_fee = amount + reminder_fee

        resend.api_key = os.environ.get("RESEND_API_KEY", "")
        if not resend.api_key:
            results.append({"invoice": inv_num, "status": "skipped", "reason": "no resend key"})
            continue

        admin_email = os.environ.get("ADMIN_EMAIL", "")
        subject = f"Paminnelse #{new_reminder_count}: Faktura #{inv_num} – {company}"
        if new_reminder_count >= 3:
            subject = f"Sista betalningspaminnelse: Faktura #{inv_num} – {company}"

        html = f"""
        <div style="font-family:-apple-system,BlinkMacSystemFont,sans-serif;max-width:520px;margin:0 auto;background:#fff;">
          <div style="background:#dc2626;padding:28px 36px;">
            <h1 style="color:#fff;font-size:20px;margin:0;font-weight:700;">{company}</h1>
            <p style="color:rgba(255,255,255,0.8);margin:6px 0 0;font-size:13px;text-transform:uppercase;">
              {"Paminnelse #" + str(new_reminder_count) if new_reminder_count < 3 else "Sista betalningspaminnelse"}
            </p>
          </div>
          <div style="padding:36px;border:1px solid #e2e8f0;border-top:none;">
            <p style="font-size:17px;color:#141414;font-weight:600;">Hej {customer_name},</p>
            <p style="color:#64748b;font-size:15px;line-height:1.7;">
              Vi har inte mottagit betalning for faktura <strong>#{inv_num}</strong>
              som forfoll for <strong style="color:#dc2626;">{days_overdue} dagar sedan</strong>.
            </p>
            <div style="background:#fef2f2;border-radius:8px;padding:20px;margin:24px 0;border-left:3px solid #dc2626;">
              <p style="margin:0 0 8px;font-size:14px;"><span style="color:#64748b;">Fakturanummer</span> &nbsp; <strong>#{inv_num}</strong></p>
              <p style="margin:0 0 8px;font-size:14px;"><span style="color:#64748b;">Forfallodatum</span> &nbsp; <strong>{due_date}</strong></p>
              <p style="margin:0 0 8px;font-size:14px;"><span style="color:#64748b;">Ursprungligt belopp</span> &nbsp; <strong>{amount:.2f} kr</strong></p>
              {"<p style=margin:0 0 8px;font-size:14px;><span style=color:#64748b;>Paminnelseavgift</span> &nbsp; <strong>" + str(reminder_fee) + " kr</strong></p>" if reminder_fee > 0 else ""}
              <p style="margin:0;font-size:16px;font-weight:700;color:#dc2626;">Att betala: {total_with_fee:.2f} kr</p>
            </div>
            {"<div style=background:#fef2f2;border:1px solid #fecaca;border-radius:6px;padding:12px 16px;margin:0 0 16px;><p style=color:#dc2626;font-size:13px;font-weight:600;margin:0;>Om betalning inte inkommer inom 10 dagar lamnas arendet till inkassobolag.</p></div>" if new_reminder_count >= 2 else ""}
            <hr style="border:none;border-top:1px solid #e2e8f0;margin:24px 0;"/>
            <p style="color:#94a3b8;font-size:12px;text-align:center;margin:0;">{company} · Automatisk paminnelse</p>
          </div>
        </div>
        """

        try:
            pdf_bytes = build_invoice_pdf(inv, inv_settings)
            import base64
            pdf_b64 = base64.b64encode(pdf_bytes).decode()

            resend.Emails.send({
                "from": f"{company} <onboarding@resend.dev>",
                "to": admin_email,
                "subject": subject,
                "html": html,
                "attachments": [{
                    "filename": f"faktura_{inv_num}.pdf",
                    "content": pdf_b64,
                }]
            })

            # Update invoice
            update_fields = {
                "status": "overdue",
                "reminder_count": new_reminder_count,
                "last_reminder_at": datetime.now(timezone.utc).isoformat()
            }

            if reminder_fee > 0:
                items = inv.get("items", [])
                items = [i for i in items if i.get("service") != "Paminnelseavgift"]
                items.append({
                    "service": "Paminnelseavgift",
                    "description": "Paminnelseavgift enligt inkassolagen (ingen moms)",
                    "quantity": 1,
                    "unit_price": reminder_fee,
                    "is_material": False,
                })
                work_items = [i for i in items if i.get("service") != "Paminnelseavgift"]
                subtotal = sum(i["quantity"] * i["unit_price"] for i in work_items)
                vat_amount = round(subtotal * 0.25, 2)
                rut_deduction = inv.get("rut_deduction", 0)
                customer_pays = round(subtotal + vat_amount - rut_deduction + reminder_fee, 2)
                update_fields["items"] = items
                update_fields["subtotal"] = round(subtotal, 2)
                update_fields["vat_amount"] = vat_amount
                update_fields["customer_pays"] = customer_pays

            await db.invoices.update_one(
                {"_id": to_object_id(invoice_id)},
                {"$set": update_fields}
            )

            sent_count += 1
            results.append({
                "invoice": inv_num,
                "customer": customer_name,
                "reminder_count": new_reminder_count,
                "reminder_fee": reminder_fee,
                "status": "sent"
            })

        except Exception as e:
            results.append({"invoice": inv_num, "status": "error", "reason": str(e)})

    return {
        "date": today_str,
        "sent": sent_count,
        "skipped": skipped_count,
        "payment_terms_days": payment_days,
        "results": results
    }

@api_router.patch("/invoices/{invoice_id}/status")
async def update_invoice_status(invoice_id: str, payload: InvoiceStatusUpdate, current=Depends(get_current_user)):
    updates = {"status": payload.status}
    if payload.status == "paid":
        updates["paid_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.invoices.update_one({"_id": to_object_id(invoice_id)}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Faktura hittades inte")
    return {"success": True}


@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str, current=Depends(get_current_user)):
    result = await db.invoices.delete_one({"_id": to_object_id(invoice_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Faktura hittades inte")
    return {"success": True}


@api_router.put("/invoices/{invoice_id}", response_model=Invoice, response_model_by_alias=False)
async def update_invoice(invoice_id: str, payload: InvoiceCreate, current=Depends(get_current_user)):
    existing = await db.invoices.find_one({"_id": to_object_id(invoice_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Faktura hittades inte")
    inv_settings = await get_invoice_settings_obj()
    items = [i.model_dump() for i in payload.items]
    amounts = calc_invoice_amounts(items, payload.rut_eligible, payload.customer_type, inv_settings.vat_rate)
    due_date = payload.due_date or existing.get("due_date") or (DateClass.today() + timedelta(days=inv_settings.payment_terms_days)).isoformat()

    updates = payload.model_dump()
    updates["items"] = items
    updates["due_date"] = due_date
    updates.update(amounts)

    await db.invoices.update_one({"_id": to_object_id(invoice_id)}, {"$set": updates})
    doc = await db.invoices.find_one({"_id": to_object_id(invoice_id)})
    doc["_id"] = str(doc["_id"])
    return Invoice(**doc)


@api_router.get("/invoices/{invoice_id}/pdf")
async def get_invoice_pdf(invoice_id: str, current=Depends(get_current_user)):
    doc = await db.invoices.find_one({"_id": to_object_id(invoice_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Faktura hittades inte")
    inv_settings = await get_invoice_settings_obj()
    doc["_id"] = str(doc["_id"])
    pdf_bytes = build_invoice_pdf(doc, inv_settings)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="faktura_{doc["invoice_number"]}.pdf"'},
    )



@api_router.get("/payroll/slip")
async def payroll_slip(start: str, end: str, employee_id: str, current=Depends(get_current_user)):
    summary, settings = await build_payroll_summary(start, end)
    if employee_id not in summary:
        raise HTTPException(status_code=404, detail="Anställd hittades inte")
    row = summary[employee_id]
    inv_settings = await get_invoice_settings_obj()

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        topMargin=20*mm, bottomMargin=20*mm, leftMargin=20*mm, rightMargin=20*mm,
        title="Lönebesked - " + row.get("name","") + " " + start + " - " + end,
            author="PureNorth Städ")
    styles = getSampleStyleSheet()
    elements = []

    # Header: logo + company name
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=18)
    if inv_settings.company_logo:
        try:
            from reportlab.platypus import Image as RLImage
            from reportlab.lib.utils import ImageReader
            logo_data = base64.b64decode(inv_settings.company_logo.split(",")[-1])
            logo_buf = BytesIO(logo_data)
            ir = ImageReader(logo_buf)
            iw, ih = ir.getSize()
            ratio = (15 * mm) / ih
            logo_buf.seek(0)
            logo_img = RLImage(logo_buf, width=iw * ratio, height=15 * mm)
            hdr = Table([[logo_img, Paragraph(f'<b>{inv_settings.company_name or "Lönebesked"}</b>', title_style)]],
                colWidths=[iw * ratio + 5 * mm, None])
            hdr.setStyle(TableStyle([("VALIGN", (0,0),(-1,-1),"MIDDLE"),
                ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),4),
                ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0)]))
            elements.append(hdr)
        except Exception:
            elements.append(Paragraph(inv_settings.company_name or "Lönebesked", title_style))
    else:
        elements.append(Paragraph(inv_settings.company_name or "Lönebesked", title_style))

    elements.append(Paragraph("Lönebesked", styles["Heading2"]))
    elements.append(Spacer(1, 4*mm))

    # Employee info
    emp_type = "Fast anställd" if row.get("employment_type") == "fastanstalld" else "Vikarie"
    info_data = [
        ["Anställd:", row["name"]],
        ["Personnummer:", row["personnummer"] or "-"],
        ["Anställningstyp:", emp_type],
        ["Period:", f"{start} – {end}"],
    ]
    info_table = Table(info_data, colWidths=[50*mm, None])
    info_table.setStyle(TableStyle([
        ("FONTSIZE", (0,0),(-1,-1), 10),
        ("FONTNAME", (0,0),(0,-1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 8*mm))

    # ── Pay details table ──────────────────────────────────────────────
    rate = row["hourly_rate"]
    pay_data = [["Post", "Timmar/Antal", "À-pris (kr)", "Belopp (kr)"]]
    pay_data.append(["Normaltid", f'{row["normal_h"]:.2f} tim', f'{rate:.2f}', f'{row["base_pay"]:.2f}'])
    if row["ob1_h"] > 0:
        pay_data.append([f'OB1 ({settings.ob1_label})', f'{row["ob1_h"]:.2f} tim', f'{settings.ob1_extra:.2f}', f'{row["ob1_pay"]:.2f}'])
    if row["ob2_h"] > 0:
        pay_data.append([f'OB2 ({settings.ob2_label})', f'{row["ob2_h"]:.2f} tim', f'{settings.ob2_extra:.2f}', f'{row["ob2_pay"]:.2f}'])
    if row["expense_total"] > 0:
        pay_data.append(["Utlägg", "-", "-", f'{row["expense_total"]:.2f}'])

    # Sjuklön section
    if row.get("sjuklon_gross", 0) > 0:
        pay_data.append(["── Sjukfrånvaro ──", "", "", ""])
        pay_data.append([f'Sjuklön brutto (80% av sjuklön)', f'{row.get("sick_shifts_h",0):.2f} tim', f'{rate*0.8:.2f}', f'{row["sjuklon_gross"]:.2f}'])
        pay_data.append(["Karensavdrag (20% av veckosjuklön)", "", "", f'-{row["karensavdrag"]:.2f}'])
        pay_data.append(["Sjuklön netto (utbetalas)", "", "", f'{row["sjuklon_net"]:.2f}'])

    # Absence/loss section (fast anställd only)
    if row.get("absence_lost_amount", 0) > 0 and row.get("employment_type") == "fastanstalld" and row.get("sjuklon_gross", 0) == 0:
        pay_data.append(["Förlorad lön (ej sjukfrånvaro)", "", "", f'-{row["absence_lost_amount"]:.2f}'])

    pay_table = Table(pay_data, colWidths=[90*mm, 30*mm, 28*mm, 28*mm])
    pay_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#141414")),
        ("TEXTCOLOR", (0,0),(-1,0), colors.white),
        ("FONTSIZE", (0,0),(-1,-1), 9),
        ("ALIGN", (1,0),(-1,-1), "RIGHT"),
        ("GRID", (0,0),(-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ("BOTTOMPADDING", (0,0),(-1,0), 8),
        ("TOPPADDING", (0,0),(-1,0), 8),
        ("BACKGROUND", (0,4),(-1,4), colors.HexColor("#F8FAFC")) if row.get("sjuklon_gross",0) > 0 else ("FONTSIZE",(0,0),(0,0),9),
        ("FONTNAME", (0,4),(0,4), "Helvetica-Bold") if row.get("sjuklon_gross",0) > 0 else ("FONTSIZE",(0,0),(0,0),9),
        ("TEXTCOLOR", (3,5),(3,7), colors.HexColor("#16a34a")) if row.get("sjuklon_gross",0) > 0 else ("FONTSIZE",(0,0),(0,0),9),
    ]))
    elements.append(pay_table)
    elements.append(Spacer(1, 6*mm))

    # ── Summary box ───────────────────────────────────────────────────
    elements.append(Spacer(1, 4*mm))
    summary_data = []
    PRELSKATT = 0.30
    ARBETSGIVARAVGIFT = 0.3142
    SEMESTER = 0.12
    brutto = row["normal_h"]*row["hourly_rate"] + row["ob1_pay"] + row["ob2_pay"] + row.get("sjuklon_net",0)
    semester_kr = round(brutto * SEMESTER, 2)
    ag_avg = round(brutto * ARBETSGIVARAVGIFT, 2)
    prelskatt = round(brutto * PRELSKATT, 2)
    netto = round(brutto - prelskatt, 2)
    total_cost_employer = round(brutto + ag_avg + semester_kr, 2)

    summary_data.append(["Grundlön (normaltid):", f'{row["base_pay"]:.2f} kr'])
    if row["ob1_pay"] > 0:
        summary_data.append([f'OB1-tillägg:', f'{row["ob1_pay"]:.2f} kr'])
    if row["ob2_pay"] > 0:
        summary_data.append([f'OB2-tillägg:', f'{row["ob2_pay"]:.2f} kr'])
    if row["expense_total"] > 0:
        summary_data.append(["Utlägg:", f'{row["expense_total"]:.2f} kr'])
    if row.get("sjuklon_net", 0) > 0:
        summary_data.append(["Sjuklön netto:", f'{row["sjuklon_net"]:.2f} kr'])
        summary_data.append(["  (varav karensavdrag):", f'-{row["karensavdrag"]:.2f} kr'])
    if row.get("absence_days", 0) > 0:
        summary_data.append(["Frånvarodagar:", f'{row["absence_days"]} dagar'])
    summary_data.append(["", ""])
    summary_data.append(["Bruttolön:", f'{brutto:.2f} kr'])
    summary_data.append(["Preliminärskatt (30%):", f'-{prelskatt:.2f} kr'])
    summary_data.append(["", ""])
    summary_data.append(["NETTOLÖN:", f'{netto:.2f} kr'])
    summary_data.append(["", ""])
    summary_data.append(["", ""])
    summary_data.append(["Semesterersättning (12%):", f'{semester_kr:.2f} kr'])
    summary_data.append(["Arbetsgivaravgift (31.42%):", f'{ag_avg:.2f} kr'])
    summary_data.append(["Total kostnad för arbetsgivare:", f'{total_cost_employer:.2f} kr'])

    # Split into two sections: employee view and employer view
    # Find nettolön index
    netto_idx = next((i for i,r in enumerate(summary_data) if "NETTOLÖN" in str(r[0])), len(summary_data)-1)
    employer_idx = next((i for i,r in enumerate(summary_data) if "Semesterersättning" in str(r[0])), None)

    total_table = Table(summary_data, colWidths=[110*mm, 45*mm])
    ts = [
        ("FONTSIZE", (0,0),(-1,-1), 10),
        ("ALIGN", (1,0),(-1,-1), "RIGHT"),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
        ("TOPPADDING", (0,0),(-1,-1), 3),
        # Brutto line
        ("LINEABOVE", (0,netto_idx-2),(1,netto_idx-2), 0.5, colors.HexColor("#E2E8F0")),
        # Netto line - highlight
        ("BACKGROUND", (0,netto_idx),(-1,netto_idx), colors.HexColor("#141414")),
        ("TEXTCOLOR", (0,netto_idx),(-1,netto_idx), colors.white),
        ("FONTNAME", (0,netto_idx),(-1,netto_idx), "Helvetica-Bold"),
        ("FONTSIZE", (0,netto_idx),(-1,netto_idx), 12),
        ("TOPPADDING", (0,netto_idx),(-1,netto_idx), 8),
        ("BOTTOMPADDING", (0,netto_idx),(-1,netto_idx), 8),
    ]
    if employer_idx:
        ts += [
            ("LINEABOVE", (0,employer_idx),(1,employer_idx), 0.5, colors.HexColor("#E2E8F0")),
            ("FONTSIZE", (0,employer_idx),(-1,-1), 9),
            ("TEXTCOLOR", (0,employer_idx),(-1,-1), colors.HexColor("#64748b")),
            ("FONTNAME", (0,-1),(-1,-1), "Helvetica-Bold"),
            ("TEXTCOLOR", (0,-1),(-1,-1), colors.HexColor("#141414")),
            ("FONTSIZE", (0,-1),(-1,-1), 10),
        ]
    total_table.setStyle(TableStyle(ts))
    elements.append(total_table)

    # ── Legal note ────────────────────────────────────────────────────
    elements.append(Spacer(1, 8*mm))
    note_style = ParagraphStyle("note", parent=styles["Normal"], fontSize=7, textColor=colors.HexColor("#94a3b8"))
    elements.append(Paragraph(
        "Sjuklön beräknad enligt sjuklönelagen (1991:1047): 80% av ordinarie lön dag 1–14. "
        "Karensavdrag = 20% av genomsnittlig veckosjuklön. Vid sjukdom >14 dagar ansöker den anställde om sjukpenning från Försäkringskassan.",
        note_style
    ))

    doc.build(elements)
    fname = f"lonebesked_{row['name'].replace(' ','_')}_{start}_{end}.pdf"
    return Response(content=buf.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{fname}"'})


# ---- Price List endpoints ----
DEFAULT_PRICES = [
    {"id": "hem", "service": "Hemstädning", "description": "Löpande hemstädning per timme", "unit": "tim", "price": 478, "is_rut_eligible": True, "is_active": True},
    {"id": "stor", "service": "Storstädning", "description": "Grundlig storstädning per timme", "unit": "tim", "price": 520, "is_rut_eligible": True, "is_active": True},
    {"id": "flytt", "service": "Flyttstädning", "description": "Komplett flyttstädning per kvm", "unit": "kvm", "price": 45, "is_rut_eligible": True, "is_active": True},
    {"id": "flytt_fast_lgh", "service": "Flyttstädning lägenhet (fast)", "description": "Fast pris 1-2 rok", "unit": "fast", "price": 2500, "is_rut_eligible": True, "is_active": True},
    {"id": "kontor", "service": "Kontorsstädning", "description": "Kontorsstädning per timme (ex. moms)", "unit": "tim", "price": 350, "is_rut_eligible": False, "is_active": True},
    {"id": "fonster", "service": "Fönsterputs", "description": "Per fönster (in- och utsida)", "unit": "st", "price": 80, "is_rut_eligible": True, "is_active": True},
    {"id": "ugn", "service": "Ugnstvätt", "description": "Djuprengöring av ugn", "unit": "st", "price": 400, "is_rut_eligible": True, "is_active": True},
    {"id": "kyl", "service": "Kyl/frys rengöring", "description": "Rengöring av kyl och frys", "unit": "st", "price": 300, "is_rut_eligible": True, "is_active": True},
    {"id": "bygg", "service": "Byggstädning", "description": "Byggstädning per timme (ex. moms)", "unit": "tim", "price": 450, "is_rut_eligible": False, "is_active": True},
    {"id": "trappa", "service": "Trappstädning", "description": "Per trapphus och tillfälle", "unit": "st", "price": 350, "is_rut_eligible": False, "is_active": True},
]


@api_router.get("/settings/pricelist")
async def get_pricelist(current=Depends(get_current_user)):
    doc = await db.settings.find_one({"_key": "pricelist"})
    if not doc or not doc.get("items"):
        return {"items": DEFAULT_PRICES}
    return {"items": doc["items"]}


@api_router.put("/settings/pricelist")
async def set_pricelist(payload: PriceListSettings, current=Depends(get_current_user)):
    doc = payload.model_dump()
    doc["_key"] = "pricelist"
    await db.settings.update_one({"_key": "pricelist"}, {"$set": doc}, upsert=True)
    return payload



# ── Economy / Ekonomiöversikt ────────────────────────────────────────────────
def recalc_invoice(inv):
    """Return invoice with stored values - DB is already updated when reminder fee added"""
    return inv

@api_router.get("/economy/overview")
async def economy_overview(start: str, end: str, current=Depends(get_current_user)):
    ARBETSGIVARAVGIFT = 0.3142
    PRELSKATT = 0.30
    SEMESTERERSATTNING = 0.12

    # ── Invoices ──────────────────────────────────────────────────
    invoices = await db.invoices.find({
        "created_at": {"$gte": start, "$lte": end + "T23:59:59"}
    }).to_list(2000)

    # Core revenue calculations
    forsaljning_excl_moms = sum(i.get("subtotal", 0) for i in invoices)
    utgaende_moms = sum(i.get("vat_amount", 0) for i in invoices)
    rut_avdrag = sum(i.get("rut_deduction", 0) for i in invoices)
    kund_betalar = sum(i.get("customer_pays", 0) for i in invoices)
    betalda = sum(i.get("customer_pays", 0) for i in invoices if i.get("status") == "paid")
    obetalda = sum(i.get("customer_pays", 0) for i in invoices if i.get("status") not in ["paid", "cancelled"])

    # Påminnelseavgifter (no moms)
    paminnelse_avgifter = sum(
        sum(item.get("quantity",1) * item.get("unit_price",0)
            for item in inv.get("items",[])
            if "Påminnelseavgift" in item.get("service",""))
        for inv in invoices
    )

    invoice_count = len(invoices)
    paid_count = sum(1 for i in invoices if i.get("status") == "paid")

    # ── Payroll ───────────────────────────────────────────────────
    payroll_summary, payroll_settings = await build_payroll_summary(start, end)
    gross_salary = sum(
        r["normal_h"] * r["hourly_rate"] +
        r["ob1_h"] * payroll_settings.ob1_extra +
        r["ob2_h"] * payroll_settings.ob2_extra +
        r.get("sjuklon_net", 0)
        for r in payroll_summary.values()
    )
    arbetsgivaravgifter = round(gross_salary * ARBETSGIVARAVGIFT, 2)
    semesterersattning = round(gross_salary * SEMESTERERSATTNING, 2)
    prelskatt_estimate = round(gross_salary * PRELSKATT, 2)
    total_payroll_cost = round(gross_salary + arbetsgivaravgifter + semesterersattning, 2)
    utlagg_total = sum(r.get("expense_total", 0) for r in payroll_summary.values())

    # Per employee
    employee_costs = []
    for eid, row in payroll_summary.items():
        gross = (row["normal_h"]*row["hourly_rate"] + row["ob1_h"]*payroll_settings.ob1_extra +
                 row["ob2_h"]*payroll_settings.ob2_extra + row.get("sjuklon_net",0))
        employee_costs.append({
            "name": row["name"],
            "employment_type": row.get("employment_type","fastanstalld"),
            "normal_h": row["normal_h"],
            "hourly_rate": row["hourly_rate"],
            "gross_salary": round(gross, 2),
            "ag_avgift": round(gross * ARBETSGIVARAVGIFT, 2),
            "semester": round(gross * SEMESTERERSATTNING, 2),
            "prelskatt": round(gross * PRELSKATT, 2),
            "netto_lon": round(gross * (1 - PRELSKATT), 2),
            "total_cost": round(gross * (1 + ARBETSGIVARAVGIFT + SEMESTERERSATTNING), 2),
            "expense_total": row.get("expense_total", 0),
        })

    # ── Material costs ────────────────────────────────────────────
    real_costs = await db.costs.find({
        "date": {"$gte": start[:10], "$lte": end[:10]}
    }).to_list(1000)
    material_total_incl = sum(c.get("amount", 0) for c in real_costs)
    ingaende_moms_mat = sum(
        c.get("amount",0) - c.get("amount",0)/(1+c.get("moms_rate",25)/100)
        for c in real_costs
    )
    material_excl_moms = material_total_incl - ingaende_moms_mat

    expenses = await db.expenses.find({
        "date": {"$gte": start[:10], "$lte": end[:10]}
    }).to_list(1000)
    ingaende_moms_exp = sum(
        e.get("amount",0) - e.get("amount",0)/(1+e.get("moms_rate",0)/100)
        for e in expenses if e.get("moms_rate",0) > 0
    )
    ingaende_moms_total = round(ingaende_moms_mat + ingaende_moms_exp, 2)

    # ── Moms ─────────────────────────────────────────────────────
    moms_att_betala = round(utgaende_moms - ingaende_moms_total, 2)

    # ── AGI ──────────────────────────────────────────────────────
    agi_to_pay = round(arbetsgivaravgifter + prelskatt_estimate, 2)

    # ── Profit ───────────────────────────────────────────────────
    total_intakter = forsaljning_excl_moms + paminnelse_avgifter
    total_kostnader = total_payroll_cost + utlagg_total + material_excl_moms
    rorelseresultat = round(total_intakter - total_kostnader, 2)
    rorselsemarginal = round((rorelseresultat / total_intakter * 100) if total_intakter > 0 else 0, 1)

    # ── Costs by category ─────────────────────────────────────────
    costs_by_category = {}
    for c in real_costs:
        cat = c.get("category","material")
        costs_by_category[cat] = costs_by_category.get(cat,0) + c.get("amount",0)

    return {
        "period": {"start": start, "end": end},
        "revenue": {
            "forsaljning_excl_moms": round(forsaljning_excl_moms, 2),
            "paminnelse_avgifter": round(paminnelse_avgifter, 2),
            "total_intakter": round(total_intakter, 2),
            "utgaende_moms": round(utgaende_moms, 2),
            "rut_avdrag": round(rut_avdrag, 2),
            "kund_betalar": round(kund_betalar, 2),
            "betalda": round(betalda, 2),
            "obetalda": round(obetalda, 2),
            "invoice_count": invoice_count,
            "paid_count": paid_count,
            # Legacy fields for compatibility
            "excl_vat": round(forsaljning_excl_moms, 2),
            "vat_collected": round(utgaende_moms, 2),
            "rut_deductions": round(rut_avdrag, 2),
            "total_invoiced": round(kund_betalar, 2),
            "paid": round(betalda, 2),
            "unpaid": round(obetalda, 2),
            "reminder_fees": round(paminnelse_avgifter, 2),
        },
        "payroll": {
            "gross_salary": round(gross_salary, 2),
            "arbetsgivaravgifter": arbetsgivaravgifter,
            "arbetsgivaravgift_pct": ARBETSGIVARAVGIFT * 100,
            "semesterersattning": semesterersattning,
            "prelskatt_estimate": prelskatt_estimate,
            "total_payroll_cost": total_payroll_cost,
            "utlagg": round(utlagg_total, 2),
            "employees": employee_costs,
        },
        "material_costs": {
            "total_incl_moms": round(material_total_incl, 2),
            "total_excl_moms": round(material_excl_moms, 2),
            "ingaende_moms": round(ingaende_moms_mat, 2),
            "by_category": {k: round(v,2) for k,v in costs_by_category.items()},
            "count": len(real_costs),
        },
        "vat": {
            "collected": round(utgaende_moms, 2),
            "ingoing_estimate": round(ingaende_moms_total, 2),
            "to_pay": round(moms_att_betala, 2),
            # Legacy
            "ingoing_estimate": round(ingaende_moms_total, 2),
        },
        "obligations": {
            "salaries_to_pay": round(gross_salary * (1-PRELSKATT), 2),
            "agi_to_pay": agi_to_pay,
            "arbetsgivaravgifter": arbetsgivaravgifter,
            "prelskatt_estimate": prelskatt_estimate,
            "vat_to_pay": round(moms_att_betala, 2),
            "total_to_pay": round(gross_salary*(1-PRELSKATT) + agi_to_pay + moms_att_betala, 2),
        },
        "result": {
            "revenue_excl_vat": round(total_intakter, 2),
            "total_costs": round(total_kostnader, 2),
            "operating_profit": rorelseresultat,
            "profit_margin": rorselsemarginal,
        },
        "constants": {
            "arbetsgivaravgift_pct": ARBETSGIVARAVGIFT * 100,
            "prelskatt_pct": PRELSKATT * 100,
            "semesterersattning_pct": SEMESTERERSATTNING * 100,
        }
    }


# ── CRM - Kundregister ────────────────────────────────────────────────────────
class CustomerCreate(BaseModel):
    name: str = Field(..., min_length=1)
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    personnummer: Optional[str] = None
    customer_type: str = "private"  # private / company
    notes: Optional[str] = None

class CustomerUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    personnummer: Optional[str] = None
    customer_type: Optional[str] = None
    notes: Optional[str] = None

class Customer(BaseModel):
    model_config = ConfigDict(populate_by_name=True)
    id: Optional[PyObjectId] = Field(default=None, alias="_id")
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    personnummer: Optional[str] = None
    customer_type: str = "private"
    notes: Optional[str] = None
    booking_count: int = 0
    invoice_count: int = 0
    total_invoiced: float = 0.0
    last_booking: Optional[str] = None
    created_at: str


@api_router.post("/customers", response_model=Customer, response_model_by_alias=False)
async def create_customer(payload: CustomerCreate, current=Depends(get_current_user)):
    # Check if customer already exists by email or phone
    query = {}
    if payload.email:
        query["email"] = payload.email.lower()
    elif payload.phone:
        query["phone"] = payload.phone
    if query:
        existing = await db.customers.find_one(query)
        if existing:
            existing["_id"] = str(existing["_id"])
            return Customer(**existing)
    doc = payload.model_dump()
    if doc.get("email"):
        doc["email"] = doc["email"].lower()
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["booking_count"] = 0
    doc["invoice_count"] = 0
    doc["total_invoiced"] = 0.0
    doc["last_booking"] = None
    result = await db.customers.insert_one(doc)
    doc["_id"] = str(result.inserted_id)
    return Customer(**doc)


async def enrich_customer(c: dict) -> dict:
    cid = str(c["_id"])
    c["_id"] = cid
    email = (c.get("email") or "").strip().lower()
    phone = (c.get("phone") or "").strip()

    # Only query if we have actual identifying info
    bookings = []
    invoices = []

    if email and phone:
        bookings = await db.bookings.find({"$or": [{"email": email}, {"phone": phone}]}).sort("created_at", -1).to_list(1000)
    elif email:
        bookings = await db.bookings.find({"email": email}).sort("created_at", -1).to_list(1000)
    elif phone:
        bookings = await db.bookings.find({"phone": phone}).sort("created_at", -1).to_list(1000)

    if email:
        invoices = await db.invoices.find({"customer_email": email}).to_list(1000)

    c["booking_count"] = len(bookings)
    c["last_booking"] = bookings[0].get("preferred_date") or bookings[0]["created_at"][:10] if bookings else None
    c["invoice_count"] = len(invoices)
    c["total_invoiced"] = sum(i.get("customer_pays", 0) for i in invoices)
    return c


@api_router.get("/customers", response_model=List[Customer], response_model_by_alias=False)
async def list_customers(current=Depends(get_current_user)):
    docs = await db.customers.find().sort("name", 1).to_list(2000)
    result = []
    for d in docs:
        enriched = await enrich_customer(d)
        result.append(Customer(**enriched))
    return result


# ── Customers Excel Export ────────────────────────────────────────────────────
@api_router.get("/customers/export-xlsx")
async def export_customers_xlsx(current=Depends(get_current_user)):
    """Export customers to Excel"""
    customers = await db.customers.find().sort("name", 1).to_list(5000)
    invoices = await db.invoices.find().to_list(5000)

    # Calculate stats per customer
    inv_by_customer = {}
    for inv in invoices:
        name = inv.get("customer_name","")
        if name not in inv_by_customer:
            inv_by_customer[name] = {"count":0,"total":0,"paid":0,"last":""}
        inv_by_customer[name]["count"] += 1
        inv_by_customer[name]["total"] += inv.get("customer_pays",0)
        if inv.get("status") == "paid":
            inv_by_customer[name]["paid"] += inv.get("customer_pays",0)
        date = inv.get("created_at","")[:10]
        if date > inv_by_customer[name]["last"]:
            inv_by_customer[name]["last"] = date

    wb = Workbook()
    ws = wb.active
    ws.title = "Kunder"

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    headers = [
        "Namn", "E-post", "Telefon", "Adress", "Personnummer",
        "Typ", "Antal fakturor", "Totalt fakturerat (kr)",
        "Totalt betalt (kr)", "Senaste faktura",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="141414", end_color="141414", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for c in customers:
        name = c.get("name","")
        stats = inv_by_customer.get(name, {"count":0,"total":0,"paid":0,"last":"-"})
        ws.append([
            name,
            c.get("email",""),
            c.get("phone",""),
            c.get("address",""),
            c.get("personnummer",""),
            "Företag" if c.get("customer_type") == "company" else "Privat",
            stats["count"],
            round(stats["total"],2),
            round(stats["paid"],2),
            stats["last"] or "-",
        ])
        for cell in ws[ws.max_row]:
            cell.font = Font(size=9)

    col_widths = [25, 25, 15, 30, 15, 10, 16, 22, 20, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="kunder.xlsx"'}
    )




# ── Customers PDF Export ──────────────────────────────────────────────────────
@api_router.get("/customers/export-pdf")
async def export_customers_pdf(current=Depends(get_current_user)):
    """Export customers list to PDF"""
    customers = await db.customers.find().sort("name", 1).to_list(5000)
    invoices = await db.invoices.find().to_list(5000)
    inv_settings = await get_invoice_settings_obj()
    company = inv_settings.company_name or "PureNorth Städ"

    # Stats per customer
    inv_by_customer = {}
    for inv in invoices:
        name = inv.get("customer_name","")
        if name not in inv_by_customer:
            inv_by_customer[name] = {"count":0,"total":0,"paid":0,"last":""}
        inv_by_customer[name]["count"] += 1
        inv_by_customer[name]["total"] += inv.get("customer_pays",0)
        if inv.get("status") == "paid":
            inv_by_customer[name]["paid"] += inv.get("customer_pays",0)
        date = inv.get("created_at","")[:10]
        if date > inv_by_customer[name]["last"]:
            inv_by_customer[name]["last"] = date

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        topMargin=15*mm, bottomMargin=15*mm, leftMargin=20*mm, rightMargin=20*mm,
        title="Kundregister", author=company)
    styles = getSampleStyleSheet()
    def ps(name, **kw): return ParagraphStyle(name, parent=styles["Normal"], **kw)
    page_w = A4[0]
    elements = []

    # Header
    hdr = Table([[
        Paragraph(f"<b>{company}</b>", ps("h", fontSize=14, fontName="Helvetica-Bold", textColor=colors.white)),
        Paragraph(f"<b>KUNDREGISTER</b><br/>Skapad: {datetime.now(timezone.utc).strftime('%Y-%m-%d')}", ps("s", fontSize=10, fontName="Helvetica-Bold", textColor=colors.white, alignment=2))
    ]], colWidths=[100*mm, None])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#141414")),
        ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),
        ("TOPPADDING",(0,0),(-1,-1),10),("BOTTOMPADDING",(0,0),(-1,-1),10),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    elements.append(hdr)
    elements.append(Spacer(1,3*mm))

    # Summary
    total_invoiced = sum(s["total"] for s in inv_by_customer.values())
    active = sum(1 for s in inv_by_customer.values() if s["count"] > 0)
    sum_data = [[
        Paragraph(f"<b>Antal kunder</b><br/>{len(customers)} st", ps("sb", fontSize=9)),
        Paragraph(f"<b>Aktiva kunder</b><br/>{active} st", ps("sb", fontSize=9, textColor=colors.HexColor("#15803d"))),
        Paragraph(f"<b>Totalt fakturerat</b><br/>{total_invoiced:.2f} kr", ps("sb", fontSize=9, textColor=colors.HexColor("#1e40af"))),
    ]]
    sum_tbl = Table(sum_data, colWidths=[(page_w-40*mm)/3]*3)
    sum_tbl.setStyle(TableStyle([
        ("BOX",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),
        ("INNERGRID",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),
        ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),
        ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8),
    ]))
    elements.append(sum_tbl)
    elements.append(Spacer(1,5*mm))

    # Customer table
    data = [["Namn", "E-post", "Telefon", "Typ", "Fakturor", "Totalt (kr)", "Senaste faktura"]]
    for c in customers:
        name = c.get("name","")
        stats = inv_by_customer.get(name, {"count":0,"total":0,"paid":0,"last":"-"})
        data.append([
            name,
            c.get("email","") or "-",
            c.get("phone","") or "-",
            "Företag" if c.get("customer_type") == "company" else "Privat",
            str(stats["count"]),
            f'{stats["total"]:.2f}' if stats["total"] > 0 else "-",
            stats["last"] or "-",
        ])

    tbl = Table(data, colWidths=[40*mm, 48*mm, 25*mm, 14*mm, 14*mm, 20*mm, 25*mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#141414")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),8.5),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
        ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
        ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("ALIGN",(4,0),(5,-1),"RIGHT"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#F8FAFC")]),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1,4*mm))
    elements.append(Paragraph(
        f"{company}  ·  Kundregister  ·  {datetime.now(timezone.utc).strftime('%Y-%m-%d')}",
        ps("ft", fontSize=7, textColor=colors.HexColor("#94a3b8"))
    ))

    doc.build(elements)
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": 'inline; filename="kundregister.pdf"'}
    )


@api_router.get("/customers/{customer_id}", response_model_by_alias=False)
async def get_customer(customer_id: str, current=Depends(get_current_user)):
    doc = await db.customers.find_one({"_id": to_object_id(customer_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Kund hittades inte")
    enriched = await enrich_customer(doc)
    email = enriched.get("email", "")
    phone = enriched.get("phone", "")
    query = {}
    if email:
        query = {"$or": [{"email": email}, {"phone": phone}]} if phone else {"email": email}
    elif phone:
        query = {"phone": phone}
    bookings = await db.bookings.find(query).sort("created_at", -1).to_list(100) if query else []
    invoices = await db.invoices.find({"customer_email": email}).sort("invoice_number", -1).to_list(100) if email else []
    for b in bookings:
        b["_id"] = str(b["_id"])
    for i in invoices:
        i["_id"] = str(i["_id"])
    return {
        "customer": Customer(**enriched).model_dump(),
        "bookings": bookings,
        "invoices": invoices,
    }


@api_router.patch("/customers/{customer_id}", response_model=Customer, response_model_by_alias=False)
async def update_customer(customer_id: str, payload: CustomerUpdate, current=Depends(get_current_user)):
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    if updates:
        await db.customers.update_one({"_id": to_object_id(customer_id)}, {"$set": updates})
    doc = await db.customers.find_one({"_id": to_object_id(customer_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Kund hittades inte")
    enriched = await enrich_customer(doc)
    return Customer(**enriched)


@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, current=Depends(get_current_user)):
    result = await db.customers.delete_one({"_id": to_object_id(customer_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kund hittades inte")
    return {"success": True}


@api_router.post("/customers/import-from-bookings")
async def import_customers_from_bookings(current=Depends(get_current_user)):
    bookings = await db.bookings.find().to_list(5000)
    created = 0
    updated = 0
    for b in bookings:
        email = b.get("email", "").lower()
        phone = b.get("phone", "")
        name = b.get("name", "")
        if not name:
            continue
        existing = None
        if email:
            existing = await db.customers.find_one({"email": email})
        if not existing and phone:
            existing = await db.customers.find_one({"phone": phone})
        if existing:
            # Update if missing info
            updates = {}
            if not existing.get("email") and email:
                updates["email"] = email
            if not existing.get("phone") and phone:
                updates["phone"] = phone
            if updates:
                await db.customers.update_one({"_id": existing["_id"]}, {"$set": updates})
                updated += 1
        else:
            doc = {
                "name": name,
                "email": email or None,
                "phone": phone or None,
                "address": None,
                "personnummer": None,
                "customer_type": "private",
                "notes": None,
                "booking_count": 0,
                "invoice_count": 0,
                "total_invoiced": 0.0,
                "last_booking": None,
                "created_at": b.get("created_at", datetime.now(timezone.utc).isoformat()),
            }
            await db.customers.insert_one(doc)
            created += 1
    return {"created": created, "updated": updated}


# ── Dashboard ─────────────────────────────────────────────────────────────────
@api_router.get("/dashboard")
async def dashboard(current=Depends(get_current_user)):
    today = DateClass.today().isoformat()
    week_start = (DateClass.today() - timedelta(days=DateClass.today().weekday())).isoformat()
    week_end = (DateClass.today() + timedelta(days=6 - DateClass.today().weekday())).isoformat()
    month_start = DateClass.today().replace(day=1).isoformat()
    month_end = DateClass.today().isoformat()

    # ── Today's shifts ──────────────────────────────────────────────
    todays_shifts = await db.shifts.find({"date": today}).sort("start_time", 1).to_list(50)
    employees = await db.employees.find().to_list(100)
    emp_map = {str(e["_id"]): e for e in employees}
    for s in todays_shifts:
        s["_id"] = str(s["_id"])
        emp = emp_map.get(s["employee_id"], {})
        s["employee_name"] = emp.get("name", "Okänd")
        s["employee_color"] = emp.get("color", "#141414")

    # ── New bookings ─────────────────────────────────────────────────
    new_bookings = await db.bookings.find({"status": "new"}).sort("created_at", -1).to_list(20)
    for b in new_bookings:
        b["_id"] = str(b["_id"])

    # ── Upcoming bookings (next 7 days) ──────────────────────────────
    upcoming = await db.bookings.find({
        "preferred_date": {"$gte": today, "$lte": week_end},
        "status": {"$ne": "done"}
    }).sort("preferred_date", 1).to_list(10)
    for b in upcoming:
        b["_id"] = str(b["_id"])

    # ── Unpaid invoices ───────────────────────────────────────────────
    unpaid = await db.invoices.find({"status": {"$in": ["draft","sent","overdue"]}}).sort("due_date", 1).to_list(20)
    for i in unpaid:
        i["_id"] = str(i["_id"])
    unpaid_total = sum(i.get("customer_pays", 0) for i in unpaid)
    overdue = [i for i in unpaid if i.get("due_date", "9999") < today]

    # ── This week stats ───────────────────────────────────────────────
    week_shifts = await db.shifts.find({"date": {"$gte": week_start, "$lte": week_end}}).to_list(500)
    week_bookings = await db.bookings.find({"created_at": {"$gte": week_start}}).to_list(500)

    # ── This month stats ──────────────────────────────────────────────
    month_invoices = await db.invoices.find({
        "created_at": {"$gte": month_start}
    }).to_list(500)
    month_revenue = sum(i.get("subtotal", 0) for i in month_invoices)
    month_paid = sum(i.get("customer_pays", 0) for i in month_invoices if i.get("status") == "paid")

    # ── Sick employees today ──────────────────────────────────────────
    sick_today = await db.absences.find({
        "start_date": {"$lte": today},
        "end_date": {"$gte": today},
        "type": "Sjuk"
    }).to_list(20)
    sick_names = []
    for a in sick_today:
        emp = emp_map.get(a["employee_id"], {})
        sick_names.append(emp.get("name", "Okänd"))

    # ── Absent employees today (all types) ────────────────────────────
    absent_today = await db.absences.find({
        "start_date": {"$lte": today},
        "end_date": {"$gte": today},
    }).to_list(20)

    # ── Material costs this month ─────────────────────────────────────
    month_costs = await db.costs.find({
        "date": {"$gte": month_start, "$lte": month_end}
    }).to_list(500)
    month_material_cost = sum(c.get("amount", 0) for c in month_costs)

    return {
        "today": today,
        "todays_shifts": todays_shifts,
        "todays_shift_count": len(todays_shifts),
        "sick_today": sick_names,
        "absent_today_count": len(absent_today),
        "new_bookings": new_bookings[:5],
        "new_bookings_count": len(new_bookings),
        "upcoming_bookings": upcoming,
        "unpaid_invoices": unpaid[:5],
        "unpaid_invoices_count": len(unpaid),
        "unpaid_total": round(unpaid_total, 2),
        "overdue_count": len(overdue),
        "week": {
            "shifts": len(week_shifts),
            "new_bookings": len(week_bookings),
            "start": week_start,
            "end": week_end,
        },
        "month": {
            "revenue": round(month_revenue, 2),
            "paid": round(month_paid, 2),
            "invoice_count": len(month_invoices),
        },
        "employee_count": len(employees),
        "month_material_cost": round(month_material_cost, 2),
        "month_costs_count": len(month_costs),
    }


# ── Password Reset ────────────────────────────────────────────────────────────
class ForgotPasswordRequest(BaseModel):
    email: EmailStr

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str = Field(..., min_length=6)


@api_router.post("/auth/forgot-password")
async def forgot_password(payload: ForgotPasswordRequest):
    email = payload.email.lower()
    user = await db.users.find_one({"email": email})
    # Always return success to avoid email enumeration
    if not user:
        return {"success": True, "message": "Om e-postadressen finns skickas ett återställningsmail."}

    import secrets
    token = secrets.token_urlsafe(32)
    expires = (datetime.now(timezone.utc) + timedelta(hours=2)).isoformat()

    await db.password_resets.insert_one({
        "email": email,
        "token": token,
        "expires": expires,
        "used": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    })

    reset_url = f"https://purenorth-stad.vercel.app/admin?reset_token={token}"

    resend.api_key = os.environ.get("RESEND_API_KEY", "")
    try:
        resend.Emails.send({
            "from": "PureNorth Städ <onboarding@resend.dev>",
            "to": email,
            "subject": "Återställ ditt lösenord — PureNorth Städ",
            "html": f"""
            <div style="font-family:sans-serif;max-width:480px;margin:0 auto;padding:32px;">
              <h2 style="color:#141414;margin-bottom:8px;">Återställ lösenord</h2>
              <p style="color:#64748b;margin-bottom:24px;">Klicka på knappen nedan för att skapa ett nytt lösenord. Länken är giltig i 2 timmar.</p>
              <a href="{reset_url}" style="display:inline-block;background:#141414;color:white;text-decoration:none;padding:12px 28px;border-radius:9999px;font-weight:600;font-size:15px;">
                Återställ lösenord
              </a>
              <p style="color:#94a3b8;font-size:13px;margin-top:24px;">Om du inte begärt detta kan du ignorera detta mail.</p>
              <hr style="border:none;border-top:1px solid #e2e8f0;margin:24px 0;"/>
              <p style="color:#94a3b8;font-size:12px;">PureNorth Städ — Miljövänlig städning i Umeå</p>
            </div>
            """
        })
    except Exception as e:
        logger.error("Failed to send reset email: %s", e)

    return {"success": True, "message": "Om e-postadressen finns skickas ett återställningsmail."}


@api_router.post("/auth/reset-password")
async def reset_password(payload: ResetPasswordRequest):
    now = datetime.now(timezone.utc).isoformat()
    reset_doc = await db.password_resets.find_one({
        "token": payload.token,
        "used": False,
        "expires": {"$gt": now},
    })
    if not reset_doc:
        raise HTTPException(status_code=400, detail="Ogiltig eller utgången återställningslänk.")

    email = reset_doc["email"]
    new_hash = hash_password(payload.new_password)
    await db.users.update_one({"email": email}, {"$set": {"password_hash": new_hash}})
    await db.password_resets.update_one({"_id": reset_doc["_id"]}, {"$set": {"used": True}})

    return {"success": True, "message": "Lösenordet har uppdaterats. Du kan nu logga in."}


# ── Recurring Bookings ────────────────────────────────────────────────────────
class RecurringBookingCreate(BaseModel):
    name: str = Field(..., min_length=1)
    email: str = "admin@purenorth.se"
    phone: str = Field(..., min_length=1)
    kvm: Optional[str] = None
    services: List[str] = Field(default_factory=list)
    preferred_date: str  # first occurrence YYYY-MM-DD
    other_description: Optional[str] = None
    invoice_id: Optional[str] = None
    recurrence: str = "weekly"  # weekly, biweekly, monthly
    occurrences: int = Field(default=6, ge=2, le=52)


@api_router.post("/bookings/recurring")
async def create_recurring_bookings(payload: RecurringBookingCreate, current=Depends(get_current_user)):
    from datetime import date as DateClass2
    y, m, d = map(int, payload.preferred_date.split("-"))
    start_date = DateClass2(y, m, d)

    if payload.recurrence == "weekly":
        delta = timedelta(weeks=1)
    elif payload.recurrence == "biweekly":
        delta = timedelta(weeks=2)
    else:  # monthly
        delta = None  # handled separately

    created = []
    current_date = start_date

    for i in range(payload.occurrences):
        if payload.recurrence == "monthly":
            # Add months
            month = start_date.month + i
            year = start_date.year + (month - 1) // 12
            month = ((month - 1) % 12) + 1
            import calendar
            day = min(start_date.day, calendar.monthrange(year, month)[1])
            current_date = DateClass2(year, month, day)
        
        doc = {
            "name": payload.name,
            "email": payload.email,
            "phone": payload.phone,
            "kvm": payload.kvm,
            "services": payload.services,
            "preferred_date": current_date.isoformat(),
            "other_description": payload.other_description,
            "status": "new",
            "is_recurring": True,
            "recurrence": payload.recurrence,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        result = await db.bookings.insert_one(doc)
        doc["_id"] = str(result.inserted_id)
        created.append(doc)

        if payload.recurrence != "monthly":
            current_date += delta

    return {"created": len(created), "bookings": [{"id": str(b["_id"]), "date": b["preferred_date"]} for b in created]}


# ── Statistics / Analytics ───────────────────────────────────────────────────
@api_router.get("/stats/overview")
async def stats_overview(months: int = 6, current=Depends(get_current_user)):
    from datetime import date as D
    today = D.today()

    # Build last N months
    monthly = []
    for i in range(months - 1, -1, -1):
        m = today.month - i
        y = today.year
        while m <= 0:
            m += 12
            y -= 1
        month_start = f"{y}-{m:02d}-01"
        import calendar
        last_day = calendar.monthrange(y, m)[1]
        month_end = f"{y}-{m:02d}-{last_day:02d}"
        label = f"{['Jan','Feb','Mar','Apr','Maj','Jun','Jul','Aug','Sep','Okt','Nov','Dec'][m-1]} {y}"

        # Revenue
        invoices = await db.invoices.find({
            "created_at": {"$gte": month_start, "$lte": month_end + "T23:59:59"}
        }).to_list(500)
        revenue = sum(i.get("subtotal", 0) for i in invoices)
        paid = sum(i.get("customer_pays", 0) for i in invoices if i.get("status") == "paid")

        # Bookings
        bookings = await db.bookings.find({
            "created_at": {"$gte": month_start, "$lte": month_end + "T23:59:59"}
        }).to_list(500)
        new_b = sum(1 for b in bookings if b.get("status") == "new")
        done_b = sum(1 for b in bookings if b.get("status") == "done")

        # Shifts / hours
        shifts = await db.shifts.find({"date": {"$gte": month_start, "$lte": month_end}}).to_list(500)
        total_hours = 0
        for s in shifts:
            try:
                sh, sm = map(int, s["start_time"].split(":"))
                eh, em = map(int, s["end_time"].split(":"))
                h = (eh * 60 + em - sh * 60 - sm) / 60
                if h < 0: h += 24
                total_hours += h
            except:
                pass

        monthly.append({
            "label": label,
            "revenue": round(revenue, 2),
            "paid": round(paid, 2),
            "bookings": len(bookings),
            "bookings_new": new_b,
            "bookings_done": done_b,
            "shifts": len(shifts),
            "hours": round(total_hours, 1),
        })

    # Top services
    all_bookings = await db.bookings.find().to_list(5000)
    service_count = {}
    for b in all_bookings:
        for s in b.get("services", []):
            service_count[s] = service_count.get(s, 0) + 1
    top_services = sorted(service_count.items(), key=lambda x: -x[1])[:6]

    # Booking status breakdown
    status_count = {"new": 0, "contacted": 0, "done": 0}
    for b in all_bookings:
        st = b.get("status", "new")
        if st in status_count:
            status_count[st] += 1

    # Employee hours this month
    month_start_curr = f"{today.year}-{today.month:02d}-01"
    import calendar as cal
    last_day_curr = cal.monthrange(today.year, today.month)[1]
    month_end_curr = f"{today.year}-{today.month:02d}-{last_day_curr:02d}"
    curr_shifts = await db.shifts.find({"date": {"$gte": month_start_curr, "$lte": month_end_curr}}).to_list(500)
    employees = await db.employees.find().to_list(100)
    emp_hours = {}
    emp_map = {str(e["_id"]): e["name"] for e in employees}
    for s in curr_shifts:
        eid = s["employee_id"]
        try:
            sh, sm = map(int, s["start_time"].split(":"))
            eh, em = map(int, s["end_time"].split(":"))
            h = (eh * 60 + em - sh * 60 - sm) / 60
            if h < 0: h += 24
            emp_hours[eid] = emp_hours.get(eid, 0) + h
        except:
            pass
    employee_hours = [{"name": emp_map.get(eid, "Okänd"), "hours": round(h, 1)} for eid, h in sorted(emp_hours.items(), key=lambda x: -x[1])]

    return {
        "monthly": monthly,
        "top_services": [{"service": s, "count": c} for s, c in top_services],
        "status_breakdown": status_count,
        "employee_hours_this_month": employee_hours,
        "totals": {
            "revenue": round(sum(m["revenue"] for m in monthly), 2),
            "bookings": sum(m["bookings"] for m in monthly),
            "hours": round(sum(m["hours"] for m in monthly), 1),
        }
    }


@api_router.get("/economy/report-pdf")
async def economy_report_pdf(start: str, end: str, current=Depends(get_current_user)):
    # Reuse economy overview data
    ARBETSGIVARAVGIFT = 0.3142
    PRELSKATT = 0.30
    SEMESTERERSATTNING = 0.12

    invoices = await db.invoices.find({
        "created_at": {"$gte": start, "$lte": end + "T23:59:59"}
    }).to_list(2000)

    revenue_excl_vat = sum(i.get("subtotal", 0) for i in invoices)
    vat_collected = sum(i.get("vat_amount", 0) for i in invoices)
    rut_deductions = sum(i.get("rut_deduction", 0) for i in invoices)
    total_invoiced = sum(i.get("total_amount", 0) for i in invoices)
    paid_invoices = sum(i.get("customer_pays", 0) for i in invoices if i.get("status") == "paid")
    unpaid_invoices = sum(i.get("customer_pays", 0) for i in invoices if i.get("status") != "paid")

    payroll_summary, payroll_settings = await build_payroll_summary(start, end)
    gross_salary = sum(
        r["normal_h"] * r["hourly_rate"] +
        r["ob1_h"] * payroll_settings.ob1_extra +
        r["ob2_h"] * payroll_settings.ob2_extra +
        r.get("sjuklon_net", 0)
        for r in payroll_summary.values()
    )
    arbetsgivaravgifter = round(gross_salary * ARBETSGIVARAVGIFT, 2)
    semesterersattning = round(gross_salary * SEMESTERERSATTNING, 2)
    prelskatt = round(gross_salary * PRELSKATT, 2)
    total_payroll_cost = round(gross_salary + arbetsgivaravgifter + semesterersattning, 2)
    utlagg_total = sum(r.get("expense_total", 0) for r in payroll_summary.values())
    paminnelse_fees = sum(
        sum(item.get("quantity",1)*item.get("unit_price",0)
            for item in inv.get("items",[])
            if "Påminnelseavgift" in item.get("service",""))
        for inv in invoices
    )
    total_intakter_pdf = revenue_excl_vat + paminnelse_fees
    operating_profit = round(total_intakter_pdf - total_payroll_cost - utlagg_total, 2)
    profit_margin = round((operating_profit / total_intakter_pdf * 100) if total_intakter_pdf > 0 else 0, 1)
    vat_to_pay = round(vat_collected, 2)
    agi_to_pay = round(arbetsgivaravgifter + prelskatt, 2)

    inv_settings = await get_invoice_settings_obj()

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        topMargin=20*mm, bottomMargin=20*mm, leftMargin=20*mm, rightMargin=20*mm,
        title=f"Ekonomirapport {start} - {end}",
            author="PureNorth Städ")
    styles = getSampleStyleSheet()
    elements = []

    # Header
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=20)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#64748b"))
    elements.append(Paragraph(inv_settings.company_name or "Ekonomirapport", title_style))
    elements.append(Paragraph(f"Ekonomirapport · Period: {start} – {end}", sub_style))
    elements.append(Paragraph(f"Skapad: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC", sub_style))
    elements.append(Spacer(1, 8*mm))

    def section(title):
        elements.append(Spacer(1, 4*mm))
        elements.append(Paragraph(title, styles["Heading2"]))
        elements.append(Spacer(1, 2*mm))

    def table2(rows, highlight_last=True):
        t = Table(rows, colWidths=[110*mm, 50*mm])
        style = [
            ("FONTSIZE", (0,0),(-1,-1), 10),
            ("ALIGN", (1,0),(-1,-1), "RIGHT"),
            ("GRID", (0,0),(-1,-1), 0.5, colors.HexColor("#E2E8F0")),
            ("BOTTOMPADDING", (0,0),(-1,-1), 6),
            ("TOPPADDING", (0,0),(-1,-1), 6),
            ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#F8FAFC")),
            ("FONTNAME", (0,0),(-1,0), "Helvetica-Bold"),
        ]
        if highlight_last:
            style += [
                ("FONTNAME", (0,-1),(-1,-1), "Helvetica-Bold"),
                ("LINEABOVE", (0,-1),(-1,-1), 1, colors.HexColor("#141414")),
            ]
        t.setStyle(TableStyle(style))
        elements.append(t)

    # Calculate reminder fees and correct totals
    kund_betalar = sum(i.get("customer_pays", 0) for i in invoices)
    betalda = sum(i.get("customer_pays", 0) for i in invoices if i.get("status") == "paid")
    obetalda = sum(i.get("customer_pays", 0) for i in invoices if i.get("status") not in ["paid","cancelled"])

    # Revenue
    section("Intäkter")
    rev_rows = [
        ["Post", "Belopp"],
        ["Försäljning tjänster (exkl. moms)", f"{revenue_excl_vat:,.2f} kr"],
    ]
    if paminnelse_fees > 0:
        rev_rows.append(["Påminnelseavgifter (ingen moms)", f"{paminnelse_fees:,.2f} kr"])
    rev_rows += [
        ["Utgående moms (25%)", f"{vat_collected:,.2f} kr"],
        ["RUT-avdrag (betalas av Skatteverket)", f"{rut_deductions:,.2f} kr"],
        ["Kund betalar totalt", f"{kund_betalar:,.2f} kr"],
        ["Betalda fakturor", f"{betalda:,.2f} kr"],
        ["Obetalda fakturor", f"{obetalda:,.2f} kr"],
        ["TOTALA INTÄKTER (exkl. moms)", f"{revenue_excl_vat + paminnelse_fees:,.2f} kr"],
    ]
    table2(rev_rows)

    # Payroll
    section("Personalkostnader")
    table2([
        ["Post", "Belopp"],
        ["Bruttolöner", f"{gross_salary:,.2f} kr"],
        [f"Arbetsgivaravgifter ({ARBETSGIVARAVGIFT*100:.2f}%)", f"{arbetsgivaravgifter:,.2f} kr"],
        [f"Semesterersättning ({SEMESTERERSATTNING*100:.0f}%)", f"{semesterersattning:,.2f} kr"],
        ["Utlägg", f"{utlagg_total:,.2f} kr"],
        ["TOTAL PERSONALKOSTNAD", f"{total_payroll_cost + utlagg_total:,.2f} kr"],
    ])

    # Per employee
    if payroll_summary:
        section("Kostnad per anställd")
        emp_rows = [["Anställd", "Typ", "Timmar", "Bruttolön", "Arb.avg", "Total"]]
        for row in payroll_summary.values():
            gross = row["normal_h"]*row["hourly_rate"] + row["ob1_h"]*payroll_settings.ob1_extra + row["ob2_h"]*payroll_settings.ob2_extra
            ag = round(gross * ARBETSGIVARAVGIFT, 2)
            sem = round(gross * SEMESTERERSATTNING, 2)
            total_c = round(gross + ag + sem, 2)
            emp_type = "Vikarie" if row.get("employment_type") == "vikarie" else "Fast"
            hours = round(row["normal_h"] + row["ob1_h"] + row["ob2_h"], 1)
            emp_rows.append([row["name"], emp_type, f"{hours} h", f"{gross:,.0f} kr", f"{ag:,.0f} kr", f"{total_c:,.0f} kr"])
        t = Table(emp_rows, colWidths=[45*mm, 18*mm, 18*mm, 26*mm, 22*mm, 26*mm])
        t.setStyle(TableStyle([
            ("FONTSIZE", (0,0),(-1,-1), 9),
            ("ALIGN", (2,0),(-1,-1), "RIGHT"),
            ("GRID", (0,0),(-1,-1), 0.5, colors.HexColor("#E2E8F0")),
            ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#141414")),
            ("TEXTCOLOR", (0,0),(-1,0), colors.white),
            ("BOTTOMPADDING", (0,0),(-1,-1), 5),
            ("TOPPADDING", (0,0),(-1,-1), 5),
        ]))
        elements.append(t)

    # Moms
    section("🧾 Momsredovisning")
    table2([
        ["Post", "Belopp"],
        ["Utgående moms", f"{vat_collected:,.2f} kr"],
        ["MOMS ATT BETALA", f"{vat_to_pay:,.2f} kr"],
    ])

    # Obligations
    section("📋 Skyldigheter till Skatteverket")
    table2([
        ["Post", "Belopp"],
        ["Arbetsgivaravgifter", f"{arbetsgivaravgifter:,.2f} kr"],
        [f"Preliminärskatt (ca {PRELSKATT*100:.0f}%)", f"{prelskatt:,.2f} kr"],
        ["AGI totalt (senast 12:e varje månad)", f"{agi_to_pay:,.2f} kr"],
        ["Moms (kvartalsvis)", f"{vat_to_pay:,.2f} kr"],
        ["TOTALT TILL SKATTEVERKET", f"{agi_to_pay + vat_to_pay:,.2f} kr"],
    ])

    # Result
    section("📊 Rörelseresultat")
    profit_color = colors.HexColor("#166534") if operating_profit >= 0 else colors.HexColor("#be123c")
    table2([
        ["Post", "Belopp"],
        ["Intäkter (exkl. moms)", f"{total_intakter_pdf:,.2f} kr"],
        ["Personalkostnader", f"-{total_payroll_cost + utlagg_total:,.2f} kr"],
        ["RÖRELSERESULTAT", f"{operating_profit:,.2f} kr ({profit_margin}%)"],
    ])

    # Footer note
    elements.append(Spacer(1, 10*mm))
    note_style = ParagraphStyle("note", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#94a3b8"))
    elements.append(Paragraph(
        "⚠️ Denna rapport är ett uppskattningsverktyg. Bokföringsskatt (20,6%) ingår inte. "
        "Kontrollera alltid med din redovisningskonsult eller revisor. "
        f"Arbetsgivaravgifter: {ARBETSGIVARAVGIFT*100:.2f}% (Skatteverket 2026). "
        f"Semesterersättning: {SEMESTERERSATTNING*100:.0f}% (semesterlagen).",
        note_style
    ))

    doc.build(elements)
    fname = f"ekonomirapport_{start}_{end}.pdf"
    return Response(content=buf.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{fname}"'})


# ── User Management ───────────────────────────────────────────────────────────
class UserCreate(BaseModel):
    email: EmailStr
    password: str = Field(..., min_length=6)
    name: str = Field(..., min_length=1)
    role: str = "staff"  # "admin", "staff", or "sales"
    employee_id: Optional[str] = None


class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    password: Optional[str] = None


class UserOut(BaseModel):
    model_config = ConfigDict(populate_by_name=True)
    id: Optional[PyObjectId] = Field(default=None, alias="_id")
    email: str
    name: str
    role: str
    employee_id: Optional[str] = None
    created_at: str


async def require_admin(current=Depends(get_current_user)):
    if current.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Åtkomst nekad. Kräver admin-behörighet.")
    return current


@api_router.get("/users", response_model=List[UserOut], response_model_by_alias=False)
async def list_users(current=Depends(require_admin)):
    docs = await db.users.find().sort("created_at", 1).to_list(100)
    return [UserOut(**{**d, "_id": str(d["_id"])}) for d in docs]


@api_router.post("/users", response_model=UserOut, response_model_by_alias=False)
async def create_user(payload: UserCreate, current=Depends(require_admin)):
    email = payload.email.lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="E-postadressen används redan.")
    doc = {
        "email": email,
        "password_hash": hash_password(payload.password),
        "name": payload.name,
        "role": payload.role if payload.role in ("admin", "staff", "sales") else "staff",
        "employee_id": payload.employee_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    result = await db.users.insert_one(doc)
    doc["_id"] = str(result.inserted_id)
    return UserOut(**doc)


@api_router.patch("/users/{user_id}", response_model=UserOut, response_model_by_alias=False)
async def update_user(user_id: str, payload: UserUpdate, current=Depends(require_admin)):
    updates = {}
    if payload.name:
        updates["name"] = payload.name
    if payload.role and payload.role in ("admin", "staff", "sales"):
        updates["role"] = payload.role
    if payload.password:
        updates["password_hash"] = hash_password(payload.password)
    if updates:
        await db.users.update_one({"_id": to_object_id(user_id)}, {"$set": updates})
    doc = await db.users.find_one({"_id": to_object_id(user_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Användare hittades inte")
    doc["_id"] = str(doc["_id"])
    return UserOut(**doc)


@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current=Depends(require_admin)):
    # Prevent deleting yourself
    if str(current.get("_id", "")) == user_id or current.get("id") == user_id:
        raise HTTPException(status_code=400, detail="Du kan inte ta bort ditt eget konto.")
    result = await db.users.delete_one({"_id": to_object_id(user_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Användare hittades inte")
    return {"success": True}


# ── Global Search ─────────────────────────────────────────────────────────────
@api_router.get("/search")
async def global_search(q: str, current=Depends(get_current_user)):
    if not q or len(q.strip()) < 2:
        return {"results": []}
    
    query = q.strip()
    results = []
    
    # Search bookings
    bookings = await db.bookings.find({
        "$or": [
            {"name": {"$regex": query, "$options": "i"}},
            {"phone": {"$regex": query, "$options": "i"}},
            {"email": {"$regex": query, "$options": "i"}},
        ]
    }).limit(5).to_list(5)
    for b in bookings:
        results.append({
            "type": "booking",
            "id": str(b["_id"]),
            "title": b.get("name", ""),
            "sub": f'{", ".join(b.get("services", []))} · {b.get("preferred_date", "")}',
            "status": b.get("status", ""),
        })

    # Search customers
    customers = await db.customers.find({
        "$or": [
            {"name": {"$regex": query, "$options": "i"}},
            {"phone": {"$regex": query, "$options": "i"}},
            {"email": {"$regex": query, "$options": "i"}},
        ]
    }).limit(5).to_list(5)
    for c in customers:
        results.append({
            "type": "customer",
            "id": str(c["_id"]),
            "title": c.get("name", ""),
            "sub": f'{c.get("phone", "")} · {c.get("email", "")}',
            "status": "",
        })

    # Search invoices
    invoices = await db.invoices.find({
        "$or": [
            {"customer_name": {"$regex": query, "$options": "i"}},
            {"customer_email": {"$regex": query, "$options": "i"}},
            {"invoice_number": {"$regex": query, "$options": "i"}},
        ]
    }).limit(5).to_list(5)
    for inv in invoices:
        results.append({
            "type": "invoice",
            "id": str(inv["_id"]),
            "title": f'Faktura #{inv.get("invoice_number", "")}',
            "sub": f'{inv.get("customer_name", "")} · {inv.get("customer_pays", 0):.0f} kr',
            "status": inv.get("status", ""),
        })

    # Search employees
    employees = await db.employees.find({
        "name": {"$regex": query, "$options": "i"}
    }).limit(3).to_list(3)
    for e in employees:
        results.append({
            "type": "employee",
            "id": str(e["_id"]),
            "title": e.get("name", ""),
            "sub": f'{e.get("employment_type", "")} · {e.get("hourly_rate", 0):.0f} kr/tim',
            "status": "",
        })

    return {"results": results, "query": query, "total": len(results)}


# ── Email Notifications ───────────────────────────────────────────────────────
async def send_booking_confirmation(booking: dict, inv_settings=None):
    """Send booking confirmation email to customer"""
    email = booking.get("email", "")
    if not email or email == "admin@purenorth.se":
        return False

    resend.api_key = os.environ.get("RESEND_API_KEY", "")
    if not resend.api_key:
        return False

    name = booking.get("name", "")
    services = ", ".join(booking.get("services", []))
    date = booking.get("preferred_date", "")
    company = inv_settings.company_name if inv_settings else "PureNorth Städ"
    phone = inv_settings.company_phone if hasattr(inv_settings, "company_phone") and inv_settings.company_phone else "070-624 04 03"

    date_str = ""
    if date:
        try:
            from datetime import datetime as DT
            d = DT.strptime(date, "%Y-%m-%d")
            MONTHS_SV = ["januari","februari","mars","april","maj","juni","juli","augusti","september","oktober","november","december"]
            date_str = f"{d.day} {MONTHS_SV[d.month-1]} {d.year}"
        except:
            date_str = date

    try:
        admin_email = os.environ.get("ADMIN_EMAIL", "akhazzane.othmane@gmail.com")
        resend.Emails.send({
            "from": f"{company} <onboarding@resend.dev>",
            "to": admin_email,  # Temporary: send to admin until domain verified
            "subject": f"Bokningsbekräftelse till {name} – {company}",
            "html": f"""
            <div style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;max-width:520px;margin:0 auto;background:#fff;border-radius:16px;overflow:hidden;border:1px solid #e2e8f0;">
              <div style="background:#141414;padding:32px 36px;">
                <h1 style="color:white;font-size:24px;margin:0;font-weight:700;">{company}</h1>
                <p style="color:rgba(255,255,255,0.6);margin:4px 0 0;font-size:14px;">Bokningsbekräftelse</p>
              </div>
              <div style="padding:36px;">
                <p style="font-size:16px;color:#141414;margin:0 0 24px;">Hej <strong>{name}</strong>! </p>
                <p style="color:#475569;line-height:1.6;margin:0 0 24px;">
                  Tack för att du kontaktade oss! Vi har tagit emot din bokningsförfrågan och återkommer inom kort för att bekräfta din tid.
                </p>

                <div style="background:#f8fafc;border-radius:12px;padding:20px;margin:0 0 24px;border:1px solid #e2e8f0;">
                  <h3 style="font-size:13px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:0.05em;margin:0 0 12px;">Din bokning</h3>
                  {"<table style='width:100%;border-collapse:collapse;'><tr><td style='padding:8px 0;border-bottom:1px solid #e2e8f0;color:#64748b;font-size:13px;width:80px;'>Datum</td><td style='padding:8px 0;border-bottom:1px solid #e2e8f0;color:#141414;font-size:14px;font-weight:600;'>" + date_str + "</td></tr></table>" if date_str else ""}
                  {"<table style='width:100%;border-collapse:collapse;'><tr><td style='padding:8px 0;border-bottom:1px solid #e2e8f0;color:#64748b;font-size:13px;width:80px;'>Tjänst</td><td style='padding:8px 0;border-bottom:1px solid #e2e8f0;color:#141414;font-size:14px;font-weight:600;'>" + services + "</td></tr></table>" if services else ""}
                  {"<table style='width:100%;border-collapse:collapse;'><tr><td style='padding:8px 0;color:#64748b;font-size:13px;width:80px;'>Yta</td><td style='padding:8px 0;color:#141414;font-size:14px;font-weight:600;'>" + str(booking.get("kvm","")) + " kvm</td></tr></table>" if booking.get("kvm") else ""}
                </div>

                <p style="color:#475569;line-height:1.6;margin:0 0 24px;">
                  Har du frågor? Tveka inte att höra av dig!
                </p>

                <div style="text-align:center;margin:0 0 24px;">
                  <a href="tel:{phone.replace('-','').replace(' ','')}"
                    style="display:inline-block;background:#141414;color:white;text-decoration:none;padding:14px 32px;border-radius:9999px;font-weight:600;font-size:15px;">
                    Ring oss: {phone}
                  </a>
                </div>

                <hr style="border:none;border-top:1px solid #e2e8f0;margin:24px 0;"/>
                <p style="color:#94a3b8;font-size:12px;text-align:center;margin:0;">
                  {company} · Miljövänlig städning i Umeå · 50% RUT-avdrag
                </p>
              </div>
            </div>
            """
        })
        return True
    except Exception as e:
        logger.error(f"Failed to send booking confirmation: {e}")
        return False


# ── Material Costs ────────────────────────────────────────────────────────────
class MaterialCost(BaseModel):
    model_config = ConfigDict(populate_by_name=True)
    id: Optional[PyObjectId] = Field(default=None, alias="_id")
    name: str
    category: str = "material"
    amount: float
    antal: Optional[int] = 1
    unit_price: Optional[float] = None
    moms_rate: Optional[float] = 25.0
    date: str
    notes: Optional[str] = None
    created_at: str

class MaterialCostCreate(BaseModel):
    name: str
    category: str = "material"
    amount: float
    antal: Optional[int] = 1
    unit_price: Optional[float] = None
    moms_rate: float = 25.0
    date: str
    notes: Optional[str] = None

@api_router.get("/costs")
async def list_costs(current=Depends(get_current_user)):
    docs = await db.costs.find().sort("date", -1).to_list(1000)
    for d in docs:
        d["_id"] = str(d["_id"])
    return docs

@api_router.post("/costs")
async def create_cost(payload: MaterialCostCreate, current=Depends(get_current_user)):
    doc = payload.model_dump()
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.costs.insert_one(doc)
    doc["_id"] = str(result.inserted_id)
    return doc

@api_router.patch("/costs/{cost_id}")
async def update_cost(cost_id: str, payload: MaterialCostCreate, current=Depends(get_current_user)):
    updates = payload.model_dump()
    await db.costs.update_one({"_id": to_object_id(cost_id)}, {"$set": updates})
    doc = await db.costs.find_one({"_id": to_object_id(cost_id)})
    doc["_id"] = str(doc["_id"])
    return doc

@api_router.delete("/costs/{cost_id}")
async def delete_cost(cost_id: str, current=Depends(get_current_user)):
    await db.costs.delete_one({"_id": to_object_id(cost_id)})
    return {"success": True}

@api_router.get("/costs/overview")
async def costs_overview(start: str = None, end: str = None, current=Depends(get_current_user)):
    # Build date filter
    date_filter = {}
    if start: date_filter["$gte"] = start
    if end:   date_filter["$lte"] = end

    cost_query = {"date": date_filter} if date_filter else {}
    inv_query = {}
    payroll_query = {}

    if start or end:
        if start: inv_query["created_at"] = {"$gte": start}
        if end:   inv_query["created_at"] = {**inv_query.get("created_at", {}), "$lte": end + "T23:59:59"}

    # Revenue from paid invoices
    invoices = await db.invoices.find({**inv_query, "status": "paid"}).to_list(1000)
    revenue = sum(i.get("customer_pays", 0) for i in invoices)

    # Material costs
    costs = await db.costs.find(cost_query).to_list(1000)
    material_total = sum(c.get("amount", 0) for c in costs)
    material_excl_moms = sum(c.get("amount", 0) / (1 + c.get("moms_rate", 25)/100) for c in costs)
    ingaende_moms = material_total - material_excl_moms
    material_by_cat = {}
    for c in costs:
        cat = c.get("category", "material")
        material_by_cat[cat] = material_by_cat.get(cat, 0) + c.get("amount", 0)


    # Profit calculation
    net_profit = revenue - material_total
    margin = (net_profit / revenue * 100) if revenue > 0 else 0

    return {
        "revenue": round(revenue, 2),
        "material_costs": round(material_total, 2),
        "material_excl_moms": round(material_excl_moms, 2),
        "ingaende_moms": round(ingaende_moms, 2),
        "material_by_category": {k: round(v, 2) for k, v in material_by_cat.items()},
        "employee_costs": 0,
        "total_costs": round(material_total, 2),
        "net_profit": round(net_profit, 2),
        "margin": round(margin, 1),
        "costs_list": [{**{k: str(v) if k == "_id" else v for k, v in c.items()}} for c in costs],
    }


# ── Receipt Upload ────────────────────────────────────────────────────────────
@api_router.post("/expenses/submit")
async def submit_expense(
    employee_id: str = Form(...),
    date: str = Form(...),
    amount: float = Form(...),
    antal: int = Form(1),
    unit_price: float = Form(0.0),
    moms_rate: float = Form(25.0),
    category: str = Form("Material"),
    description: str = Form(""),
    receipt: UploadFile = File(None),
    current=Depends(get_current_user)
):
    """Staff submits expense with optional receipt photo"""
    doc = {
        "employee_id": employee_id,
        "date": date,
        "amount": amount,
        "antal": antal,
        "unit_price": unit_price,
        "moms_rate": moms_rate,
        "category": category,
        "description": description,
        "status": "pending",
        "submitted_by": current.get("email", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "receipt_image": None,
    }

    # Handle receipt image upload
    if receipt and receipt.filename:
        try:
            contents = await receipt.read()
            import base64
            ext = receipt.filename.split(".")[-1].lower()
            mime = "image/jpeg" if ext in ["jpg","jpeg"] else "image/png" if ext == "png" else "image/jpeg"
            doc["receipt_image"] = f"data:{mime};base64,{base64.b64encode(contents).decode()}"
        except Exception as e:
            logger.error(f"Receipt upload error: {e}")

    result = await db.expenses.insert_one(doc)
    doc["_id"] = str(result.inserted_id)
    return doc

@api_router.get("/expenses/pending")
async def get_pending_expenses(current=Depends(get_current_user)):
    """Get all pending expense submissions"""
    docs = await db.expenses.find({"status": "pending"}).sort("created_at", -1).to_list(100)
    for d in docs:
        d["_id"] = str(d["_id"])
        # Get employee name
        try:
            emp = await db.employees.find_one({"_id": to_object_id(d["employee_id"])})
            d["employee_name"] = emp.get("name", "Okänd") if emp else "Okänd"
        except:
            d["employee_name"] = "Okänd"
    return docs

@api_router.patch("/expenses/{expense_id}/approve")
async def approve_expense(expense_id: str, current=Depends(get_current_user)):
    """Admin approves expense - moves to approved status"""
    await db.expenses.update_one(
        {"_id": to_object_id(expense_id)},
        {"$set": {"status": "approved", "approved_by": current.get("email"), "approved_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"success": True}

@api_router.patch("/expenses/{expense_id}/reject")
async def reject_expense(expense_id: str, current=Depends(get_current_user)):
    """Admin rejects expense"""
    await db.expenses.update_one(
        {"_id": to_object_id(expense_id)},
        {"$set": {"status": "rejected"}}
    )
    return {"success": True}

@api_router.get("/expenses/my")
async def get_my_expenses(current=Depends(get_current_user)):
    """Get current user's submitted expenses"""
    email = current.get("email", "")
    docs = await db.expenses.find({"submitted_by": email}).sort("created_at", -1).to_list(100)
    for d in docs:
        d["_id"] = str(d["_id"])
    return docs


# ── Send Invoice by Email ─────────────────────────────────────────────────────
@api_router.post("/invoices/{invoice_id}/send")
async def send_invoice_email(invoice_id: str, current=Depends(get_current_user)):
    doc = await db.invoices.find_one({"_id": to_object_id(invoice_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Faktura hittades inte")

    customer_email = doc.get("customer_email", "")
    if not customer_email:
        raise HTTPException(status_code=400, detail="Kunden har ingen e-postadress")

    inv_settings = await get_invoice_settings_obj()
    company = inv_settings.company_name or "PureNorth Städ"
    inv_num = doc.get("invoice_number", "")
    customer_name = doc.get("customer_name", "")
    due_date = doc.get("due_date", "")
    amount = doc.get("customer_pays", 0)

    # Generate PDF
    pdf_bytes = build_invoice_pdf(doc, inv_settings)
    import base64
    pdf_b64 = base64.b64encode(pdf_bytes).decode()

    resend.api_key = os.environ.get("RESEND_API_KEY", "")
    if not resend.api_key:
        raise HTTPException(status_code=500, detail="E-posttjänst ej konfigurerad")

    admin_email = os.environ.get("ADMIN_EMAIL", "akhazzane.othmane@gmail.com")

    try:
        resend.Emails.send({
            "from": f"{company} <onboarding@resend.dev>",
            "to": admin_email,  # Temp until domain verified
            "subject": f"Faktura #{inv_num} från {company}",
            "html": f"""
            <div style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;max-width:520px;margin:0 auto;background:#fff;">
              <div style="background:#141414;padding:28px 36px;">
                <h1 style="color:#fff;font-size:22px;margin:0;font-weight:700;">{company}</h1>
                <p style="color:rgba(255,255,255,0.55);margin:6px 0 0;font-size:13px;text-transform:uppercase;letter-spacing:0.05em;">Faktura</p>
              </div>
              <div style="padding:36px;border:1px solid #e2e8f0;border-top:none;">
                <p style="font-size:17px;color:#141414;margin:0 0 8px;font-weight:600;">Hej {customer_name},</p>
                <p style="color:#64748b;font-size:15px;line-height:1.7;margin:0 0 24px;">
                  Tack för att du anlitade oss! Bifogat finner du din faktura.
                </p>
                <div style="background:#f8fafc;border-radius:8px;padding:20px;margin:0 0 24px;border-left:3px solid #141414;">
                  <p style="margin:0 0 8px;font-size:14px;color:#141414;"><span style="color:#64748b;">Fakturanummer</span> &nbsp; <strong>#{inv_num}</strong></p>
                  <p style="margin:0 0 8px;font-size:14px;color:#141414;"><span style="color:#64748b;">Förfallodatum</span> &nbsp; <strong>{due_date}</strong></p>
                  <p style="margin:0;font-size:14px;color:#141414;"><span style="color:#64748b;">Att betala</span> &nbsp; <strong>{amount:.2f} kr</strong></p>
                </div>
                <p style="color:#64748b;font-size:14px;margin:0 0 24px;">
                  Fakturan finns bifogad som PDF. Har du frågor är du välkommen att kontakta oss.
                </p>
                <hr style="border:none;border-top:1px solid #e2e8f0;margin:0 0 20px;"/>
                <p style="color:#94a3b8;font-size:12px;text-align:center;margin:0;">
                  {company} · Miljövänlig städning i Umeå · 50% RUT-avdrag
                </p>
              </div>
            </div>
            """,
            "attachments": [{
                "filename": f"faktura_{inv_num}.pdf",
                "content": pdf_b64,
            }]
        })

        # Update invoice status to sent
        await db.invoices.update_one(
            {"_id": to_object_id(invoice_id)},
            {"$set": {"status": "sent", "sent_at": datetime.now(timezone.utc).isoformat()}}
        )

        return {"success": True, "message": f"Faktura skickad till {customer_email}"}

    except Exception as e:
        logger.error(f"Failed to send invoice: {e}")
        raise HTTPException(status_code=500, detail=f"Kunde inte skicka faktura: {str(e)}")


# ── Monthly Report for Konsult ───────────────────────────────────────────────
@api_router.get("/reports/monthly")
async def monthly_report_pdf(month: str, current=Depends(get_current_user)):
    """Generate comprehensive monthly report for redovisningskonsult
    month format: YYYY-MM (e.g. 2025-01)
    """
    from reportlab.platypus import Image as RLImage
    year, mon = month.split("-")
    import calendar
    last_day = calendar.monthrange(int(year), int(mon))[1]
    start = f"{month}-01"
    end = f"{month}-{last_day:02d}"
    start_dt = f"{start}T00:00:00"
    end_dt = f"{end}T23:59:59"

    ARBETSGIVARAVGIFT = 0.3142
    PRELSKATT = 0.30
    SEMESTERERSATTNING = 0.12

    # ── Data collection ────────────────────────────────────────────
    inv_settings = await get_invoice_settings_obj()
    company = inv_settings.company_name or "PureNorth Städ"
    orgnr = inv_settings.company_orgnr or "—"

    invoices_raw = await db.invoices.find({"created_at": {"$gte": start_dt, "$lte": end_dt}}).to_list(2000)
    invoices = [recalc_invoice(i) for i in invoices_raw]
    payroll_summary, payroll_settings = await build_payroll_summary(start, end)
    costs = await db.costs.find({"date": {"$gte": start, "$lte": end}}).to_list(1000)
    expenses = await db.expenses.find({"date": {"$gte": start, "$lte": end}}).to_list(1000)
    bookings = await db.bookings.find({"created_at": {"$gte": start_dt, "$lte": end_dt}}).to_list(1000)

    # ── Calculations ───────────────────────────────────────────────
    # Revenue
    revenue_excl_vat = sum(i.get("subtotal", 0) for i in invoices)
    paminnelse_fees = sum(
        sum(item.get("quantity",1)*item.get("unit_price",0)
            for item in inv.get("items",[])
            if "Påminnelseavgift" in item.get("service",""))
        for inv in invoices
    )
    total_intakter = revenue_excl_vat + paminnelse_fees
    vat_collected = sum(i.get("vat_amount", 0) for i in invoices)
    rut_deductions = sum(i.get("rut_deduction", 0) for i in invoices)
    paid_amount = sum(i.get("customer_pays", 0) for i in invoices if i.get("status") == "paid")
    unpaid_amount = sum(i.get("customer_pays", 0) for i in invoices if i.get("status") != "paid")

    # Payroll
    gross_salary = sum(
        r["normal_h"] * r["hourly_rate"] +
        r["ob1_h"] * payroll_settings.ob1_extra +
        r["ob2_h"] * payroll_settings.ob2_extra +
        r.get("sjuklon_net", 0)
        for r in payroll_summary.values()
    )
    arbetsgivaravgifter = round(gross_salary * ARBETSGIVARAVGIFT, 2)
    semesterersattning = round(gross_salary * SEMESTERERSATTNING, 2)
    prelskatt = round(gross_salary * PRELSKATT, 2)
    total_payroll = round(gross_salary + arbetsgivaravgifter + semesterersattning, 2)
    utlagg_total = sum(e.get("amount", 0) for e in expenses)

    # Material costs
    material_total_incl = sum(c.get("amount", 0) for c in costs)
    ingaende_moms = sum(c.get("amount", 0) - c.get("amount", 0) / (1 + c.get("moms_rate", 25)/100) for c in costs)
    utlagg_moms = sum(e.get("amount", 0) - e.get("amount", 0) / (1 + e.get("moms_rate", 0)/100) for e in expenses if e.get("moms_rate", 0) > 0)
    total_ingaende_moms = round(ingaende_moms + utlagg_moms, 2)

    # VAT
    vat_to_pay = round(vat_collected - total_ingaende_moms, 2)
    agi_to_pay = round(arbetsgivaravgifter + prelskatt, 2)

    # Profit
    total_costs = total_payroll + utlagg_total + (material_total_incl - ingaende_moms)
    operating_profit = round(total_intakter - total_costs, 2)
    margin = round((operating_profit / total_intakter * 100) if total_intakter > 0 else 0, 1)

    # ── PDF Build ──────────────────────────────────────────────────
    buf = BytesIO()
    page_w, page_h = A4
    doc = SimpleDocTemplate(buf, pagesize=A4,
        topMargin=15*mm, bottomMargin=20*mm, leftMargin=20*mm, rightMargin=20*mm,
        title=f"Månadsrapport {month}",
            author="PureNorth Städ")

    styles = getSampleStyleSheet()
    def ps(name, **kw): return ParagraphStyle(name, parent=styles["Normal"], **kw)

    h1 = ps("h1", fontSize=20, fontName="Helvetica-Bold", textColor=colors.HexColor("#141414"), spaceAfter=2)
    h2 = ps("h2", fontSize=12, fontName="Helvetica-Bold", textColor=colors.HexColor("#141414"), spaceBefore=8, spaceAfter=4, borderPad=2)
    normal = ps("n", fontSize=9, textColor=colors.HexColor("#1a1a1a"), leading=13)
    small = ps("s", fontSize=8, textColor=colors.HexColor("#64748b"), leading=11)
    bold = ps("b", fontSize=9, fontName="Helvetica-Bold", textColor=colors.HexColor("#1a1a1a"), leading=13)

    elements = []

    # ── HEADER ─────────────────────────────────────────────────────
    import locale
    MONTHS_SV = ["","Januari","Februari","Mars","April","Maj","Juni","Juli","Augusti","September","Oktober","November","December"]
    month_name = MONTHS_SV[int(mon)]

    header_data = [[
        Paragraph(f"<b>{company}</b>", ps("ch", fontSize=16, fontName="Helvetica-Bold", textColor=colors.white)),
        Paragraph(f"<b>MÅNADSRAPPORT</b><br/>{month_name} {year}", ps("ct", fontSize=11, fontName="Helvetica-Bold", textColor=colors.white, alignment=2))
    ]]
    header_tbl = Table(header_data, colWidths=[100*mm, None])
    header_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#141414")),
        ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),
        ("TOPPADDING",(0,0),(-1,-1),12),("BOTTOMPADDING",(0,0),(-1,-1),12),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    elements.append(header_tbl)

    # Company info
    info_line = f"Org.nr: {orgnr}  ·  Period: {start} – {end}  ·  Skapad: {datetime.now(timezone.utc).strftime('%Y-%m-%d')}"
    elements.append(Paragraph(info_line, ps("info", fontSize=8, textColor=colors.HexColor("#94a3b8"), spaceBefore=3, spaceAfter=6)))

    # Divider
    div = Table([[""]], colWidths=[page_w-40*mm])
    div.setStyle(TableStyle([("LINEBELOW",(0,0),(-1,-1),1,colors.HexColor("#141414")),("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0)]))
    elements.append(div)
    elements.append(Spacer(1,4*mm))

    def section_header(title, color="#141414"):
        t = Table([[Paragraph(title, ps("sh", fontSize=10, fontName="Helvetica-Bold", textColor=colors.white))]], colWidths=[page_w-40*mm])
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),colors.HexColor(color)),
            ("LEFTPADDING",(0,0),(-1,-1),8),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
        ]))
        elements.append(t)

    def data_row(label, value, is_total=False, color=None):
        lbl_style = bold if is_total else normal
        val_style = bold if is_total else normal
        val_para = Paragraph(value, ps("vr", fontSize=9, fontName="Helvetica-Bold" if is_total else "Helvetica",
            textColor=colors.HexColor(color) if color else colors.HexColor("#1a1a1a"), alignment=2))
        row = [[Paragraph(label, lbl_style), val_para]]
        t = Table(row, colWidths=[120*mm, None])
        bg = colors.HexColor("#F8FAFC") if is_total else colors.white
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),bg),
            ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("LINEBELOW",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
        ]))
        elements.append(t)

    def kr(n): return f"{n:,.2f} kr".replace(",",".")
    def konto(k): return f"[Konto {k}]"

    # ── 1. INTÄKTER ────────────────────────────────────────────────
    elements.append(Spacer(1,3*mm))
    section_header("1. INTÄKTER (FÖRSÄLJNING)", "#1e40af")
    data_row(f"Antal fakturor", str(len(invoices)))
    data_row(f"Antal bokningar", str(len(bookings)))
    data_row(f"Försäljning tjänster  {konto('3000')}", kr(revenue_excl_vat))
    if paminnelse_fees > 0:
        data_row(f"Påminnelseavgifter  {konto('3590')}", kr(paminnelse_fees))
    data_row(f"Utgående moms 25%  {konto('2610')}", kr(vat_collected))
    data_row(f"RUT-avdrag  {konto('3001')}", kr(rut_deductions))
    data_row("Betalt av kunder", kr(paid_amount))
    data_row("Obetalt (kundfordringar)", kr(unpaid_amount))
    data_row("TOTALT INTÄKTER (exkl. moms)", kr(total_intakter), is_total=True, color="#1e40af")

    elements.append(Spacer(1,3*mm))

    # ── 2. PERSONALKOSTNADER ───────────────────────────────────────
    section_header("2. PERSONALKOSTNADER", "#7c3aed")
    data_row(f"Bruttolön  {konto('7210')}", kr(gross_salary))
    data_row(f"Arbetsgivaravgift 31.42%  {konto('7510')}", kr(arbetsgivaravgifter))
    data_row(f"Semesterersättning 12%  {konto('7290')}", kr(semesterersattning))
    data_row(f"Preliminärskatt (skuld)  {konto('2710')}", kr(prelskatt))
    data_row(f"Utlägg reseersättning  {konto('7331')}", kr(utlagg_total))
    data_row("TOTALT PERSONALKOSTNADER", kr(total_payroll + utlagg_total), is_total=True, color="#7c3aed")

    # Per employee
    if payroll_summary:
        elements.append(Spacer(1,2*mm))
        emp_header = [["Namn", "Tim", "Bruttolön", "AG-avgift", "Semester", "Summa"]]
        emp_data = emp_header.copy()
        for r in payroll_summary.values():
            gross = r["normal_h"]*r["hourly_rate"] + r["ob1_h"]*payroll_settings.ob1_extra + r["ob2_h"]*payroll_settings.ob2_extra + r.get("sjuklon_net",0)
            emp_data.append([
                r["name"], f"{r['normal_h']:.1f}",
                kr(gross), kr(gross*ARBETSGIVARAVGIFT),
                kr(gross*SEMESTERERSATTNING), kr(gross*(1+ARBETSGIVARAVGIFT+SEMESTERERSATTNING))
            ])
        emp_tbl = Table(emp_data, colWidths=[50*mm, 18*mm, 30*mm, 25*mm, 25*mm, 30*mm])
        emp_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#F1F5F9")),
            ("FONTSIZE",(0,0),(-1,-1),8),("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
            ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("ALIGN",(1,0),(-1,-1),"RIGHT"),
        ]))
        elements.append(emp_tbl)

    elements.append(Spacer(1,3*mm))

    # ── 3. MATERIALKOSTNADER ───────────────────────────────────────
    section_header("3. MATERIALKOSTNADER & UTLÄGG", "#0369a1")
    data_row(f"Städmaterial  {konto('5410')}", kr(material_total_incl))
    data_row(f"Bränsle  {konto('5612')}", "Se utlägg")
    data_row(f"Milersättning  {konto('7331')}", "Se utlägg")
    data_row(f"Parkering  {konto('5651')}", "Se utlägg")
    data_row(f"Ingående moms material  {konto('2640')}", f"-{kr(ingaende_moms)}", color="#15803d")
    data_row(f"Ingående moms utlägg  {konto('2640')}", f"-{kr(utlagg_moms)}", color="#15803d")
    data_row("NETTOKOSTNAD MATERIAL (exkl. moms)", kr(material_total_incl - ingaende_moms), is_total=True, color="#0369a1")

    elements.append(Spacer(1,3*mm))

    # ── 4. MOMSREDOVISNING ─────────────────────────────────────────
    section_header("4. MOMSREDOVISNING (ATT REDOVISA TILL SKATTEVERKET)", "#b45309")
    data_row(f"Utgående moms försäljning  {konto('2610')}", kr(vat_collected))
    data_row(f"Ingående moms inköp + utlägg  {konto('2640')}", f"-{kr(total_ingaende_moms)}", color="#15803d")
    data_row("MOMS ATT BETALA → Bankgiro 5050-1055", kr(vat_to_pay), is_total=True, color="#b45309")

    elements.append(Spacer(1,3*mm))

    # ── 5. ARBETSGIVARDEKLARATION ──────────────────────────────────
    section_header("5. ARBETSGIVARDEKLARATION (AGI) - BETALA SENAST DEN 12:E", "#dc2626")
    data_row(f"Arbetsgivaravgift  {konto('7510')}", kr(arbetsgivaravgifter))
    data_row(f"Preliminärskatt  {konto('2710')}", kr(prelskatt))
    data_row("TOTALT ATT BETALA SKATTEVERKET → Bankgiro 5050-1055", kr(agi_to_pay), is_total=True, color="#dc2626")

    elements.append(Spacer(1,3*mm))

    # ── 6. RESULTAT ────────────────────────────────────────────────
    section_header("6. RÖRELSERESULTAT", "#15803d")
    data_row(f"Intäkter exkl. moms  {konto('3000')}", kr(total_intakter))
    data_row(f"Personalkostnader  {konto('7210/7510')}", f"-{kr(total_payroll + utlagg_total)}")
    data_row(f"Materialkostnader exkl. moms  {konto('5410')}", f"-{kr(material_total_incl - ingaende_moms)}")
    profit_color = "#15803d" if operating_profit >= 0 else "#dc2626"
    data_row("RÖRELSERESULTAT", kr(operating_profit), is_total=True, color=profit_color)
    data_row("Rörelsemarginal", f"{margin}%")

    elements.append(Spacer(1,4*mm))

    # ── BAS KONTOPLAN NOTE ─────────────────────────────────────────
    note_text = "Kontonummer enligt BAS-kontoplanen 2025. Kontrollera med din redovisningskonsult att kontona stämmer med ert bokföringssystem."
    elements.append(Paragraph(note_text, ps("note", fontSize=7, textColor=colors.HexColor("#94a3b8"), spaceBefore=2, spaceAfter=4)))

    # ── FAKTURA LIST ───────────────────────────────────────────────
    if invoices:
        section_header("7. FAKTURALISTA", "#475569")
        inv_header = [["Faktura #", "Kund", "Datum", "Exkl. moms", "Påm.avg", "Moms", "Att betala", "Status"]]
        inv_rows = inv_header.copy()
        for i in sorted(invoices, key=lambda x: x.get("created_at","")):
            status_map = {"paid":"Betald","draft":"Utkast","sent":"Skickad","overdue":"Förfallen"}
            pam_fee = sum(
                item.get("quantity",1)*item.get("unit_price",0)
                for item in i.get("items",[])
                if "Påminnelseavgift" in item.get("service","")
            )
            inv_rows.append([
                i.get("invoice_number",""),
                i.get("customer_name","")[:20],
                i.get("created_at","")[:10],
                f'{i.get("subtotal",0):.2f}',
                f'{pam_fee:.2f}' if pam_fee > 0 else "-",
                f'{i.get("vat_amount",0):.2f}',
                f'{i.get("customer_pays",0):.2f}',
                status_map.get(i.get("status",""),""),
            ])
        inv_tbl = Table(inv_rows, colWidths=[18*mm, 38*mm, 18*mm, 20*mm, 16*mm, 18*mm, 22*mm, 18*mm])
        inv_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#F1F5F9")),
            ("FONTSIZE",(0,0),(-1,-1),7.5),("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
            ("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("ALIGN",(3,0),(-1,-1),"RIGHT"),
        ]))
        elements.append(inv_tbl)

    elements.append(Spacer(1,4*mm))

    # ── FOOTER ─────────────────────────────────────────────────────
    footer_line = f"{company}  ·  Org.nr {orgnr}  ·  Månadsrapport {month_name} {year}  ·  Konfidentiellt"
    ft = Table([[Paragraph(footer_line, ps("ft", fontSize=7, textColor=colors.HexColor("#94a3b8")))]], colWidths=[page_w-40*mm])
    ft.setStyle(TableStyle([
        ("LINEABOVE",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),
        ("TOPPADDING",(0,0),(-1,-1),4),("LEFTPADDING",(0,0),(-1,-1),0),
    ]))
    elements.append(ft)

    doc.build(elements)
    pdf_bytes = buf.getvalue()

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="manadsrapport_{month}.pdf"'}
    )


# ── SIE4 Export ───────────────────────────────────────────────────────────────
@api_router.get("/reports/sie")
async def export_sie(month: str, current=Depends(get_current_user)):
    """Export SIE4 file for redovisningskonsult
    month format: YYYY-MM
    """
    import calendar
    year, mon = month.split("-")
    last_day = calendar.monthrange(int(year), int(mon))[1]
    start = f"{month}-01"
    end = f"{month}-{last_day:02d}"
    start_dt = f"{start}T00:00:00"
    end_dt = f"{end}T23:59:59"

    ARBETSGIVARAVGIFT = 0.3142
    PRELSKATT = 0.30
    SEMESTERERSATTNING = 0.12

    inv_settings = await get_invoice_settings_obj()
    company = inv_settings.company_name or "PureNorth Städ"
    orgnr = (inv_settings.company_orgnr or "").replace("-","")

    # Collect data
    invoices_raw = await db.invoices.find({"created_at": {"$gte": start_dt, "$lte": end_dt}}).to_list(2000)
    invoices = [recalc_invoice(i) for i in invoices_raw]
    payroll_summary, payroll_settings = await build_payroll_summary(start, end)
    costs = await db.costs.find({"date": {"$gte": start, "$lte": end}}).to_list(1000)
    expenses = await db.expenses.find({"date": {"$gte": start, "$lte": end}}).to_list(1000)

    # Calculate totals
    revenue = sum(i.get("subtotal", 0) for i in invoices)
    paminnelse_fees = sum(
        sum(item.get("quantity",1)*item.get("unit_price",0)
            for item in inv.get("items",[])
            if "Påminnelseavgift" in item.get("service",""))
        for inv in invoices
    )
    vat_out = sum(i.get("vat_amount", 0) for i in invoices)
    rut = sum(i.get("rut_deduction", 0) for i in invoices)

    gross_salary = sum(
        r["normal_h"] * r["hourly_rate"] +
        r["ob1_h"] * payroll_settings.ob1_extra +
        r["ob2_h"] * payroll_settings.ob2_extra +
        r.get("sjuklon_net", 0)
        for r in payroll_summary.values()
    )
    ag_avg = round(gross_salary * ARBETSGIVARAVGIFT, 2)
    semester = round(gross_salary * SEMESTERERSATTNING, 2)
    prelskatt = round(gross_salary * PRELSKATT, 2)
    utlagg = sum(e.get("amount", 0) for e in expenses)

    material = sum(c.get("amount", 0) for c in costs)
    vat_in_mat = sum(c.get("amount",0) - c.get("amount",0)/(1+c.get("moms_rate",25)/100) for c in costs)
    vat_in_exp = sum(e.get("amount",0) - e.get("amount",0)/(1+e.get("moms_rate",0)/100) for e in expenses if e.get("moms_rate",0) > 0)
    vat_in = round(vat_in_mat + vat_in_exp, 2)

    # Build SIE4 content
    now = datetime.now(timezone.utc)
    gen_date = now.strftime("%Y%m%d")
    period = month.replace("-","")

    # Verification entries (verifikationer)
    ver_num = 1
    verifikationer = []

    # 1. Revenue entries per invoice
    for inv in invoices:
        date = (inv.get("created_at","")[:10] or start).replace("-","")
        sub = inv.get("subtotal", 0)
        vat = inv.get("vat_amount", 0)
        pays = inv.get("customer_pays", 0)
        rut_d = inv.get("rut_deduction", 0)
        inv_num = inv.get("invoice_number", "")
        customer = inv.get("customer_name", "")
        ver = f'#VER A {ver_num} {date} "Faktura #{inv_num} - {customer}"\n'
        ver += f'  #TRANS 3000 {{}} {-sub:.2f} {date} "Försäljning"\n'
        ver += f'  #TRANS 2610 {{}} {-vat:.2f} {date} "Utgående moms"\n'
        if rut_d > 0:
            ver += f'  #TRANS 3001 {{}} {rut_d:.2f} {date} "RUT-avdrag"\n'
        pam_fee = sum(item.get("quantity",1)*item.get("unit_price",0)
            for item in inv.get("items",[])
            if "Påminnelseavgift" in item.get("service",""))
        if pam_fee > 0:
            ver += f'  #TRANS 3590 {{}} {-pam_fee:.2f} {date} "Påminnelseavgift"\n'
        ver += f'  #TRANS 1510 {{}} {pays:.2f} {date} "Kundfordran"\n'
        verifikationer.append(ver)
        ver_num += 1

    # 2. Payroll entry
    if gross_salary > 0:
        date = end.replace("-","")
        ver = f'#VER L {ver_num} {date} "Löner {month}"\n'
        ver += f'  #TRANS 7210 {{}} {gross_salary:.2f} {date} "Bruttolön"\n'
        ver += f'  #TRANS 7510 {{}} {ag_avg:.2f} {date} "Arbetsgivaravgift"\n'
        ver += f'  #TRANS 7290 {{}} {semester:.2f} {date} "Semesterersättning"\n'
        ver += f'  #TRANS 2710 {{}} {-prelskatt:.2f} {date} "Prelskatt skuld"\n'
        ver += f'  #TRANS 2731 {{}} {-ag_avg:.2f} {date} "AG-avgift skuld"\n'
        ver += f'  #TRANS 1930 {{}} {-(gross_salary+semester-prelskatt):.2f} {date} "Utbetald lön"\n'
        verifikationer.append(ver)
        ver_num += 1

    # 3. Material costs
    for c in costs:
        date = c.get("date","").replace("-","")
        amount = c.get("amount", 0)
        vat_c = amount - amount/(1+c.get("moms_rate",25)/100)
        net = amount - vat_c
        cat = c.get("category","material")
        konto = {"material":"5410","equipment":"5410","transport":"5612","other":"6990"}.get(cat,"5410")
        name = c.get("name","Material")
        ver = f'#VER K {ver_num} {date} "{name}"\n'
        ver += f'  #TRANS {konto} {{}} {net:.2f} {date} "{name}"\n'
        ver += f'  #TRANS 2640 {{}} {vat_c:.2f} {date} "Ingående moms"\n'
        ver += f'  #TRANS 1930 {{}} {-amount:.2f} {date} "Betalning"\n'
        verifikationer.append(ver)
        ver_num += 1

    # Build SIE file
    MONTHS_SV = ["","Januari","Februari","Mars","April","Maj","Juni","Juli","Augusti","September","Oktober","November","December"]
    sie_lines = [
        f'#FLAGGA 0',
        f'#FORMAT UTF-8',
        f'#SIETYP 4',
        f'#PROGRAM "PureNorth Admin" "1.0"',
        f'#GEN {gen_date} "PureNorth Admin"',
        f'#FNAMN "{company}"',
        f'#ORGNR {orgnr}',
        f'#RAR 0 {period}01 {period}{last_day:02d}',
        f'#TAXAR {year}',
        f'#VALUTA SEK',
        f'',
        f'#KONTO 1510 "Kundfordringar"',
        f'#KONTO 1930 "Företagskonto"',
        f'#KONTO 2610 "Utgående moms 25%"',
        f'#KONTO 2640 "Ingående moms"',
        f'#KONTO 2710 "Personalskatt"',
        f'#KONTO 2731 "Arbetsgivaravgifter"',
        f'#KONTO 3000 "Försäljning tjänster"',
        f'#KONTO 3001 "RUT-avdrag"',
        f'#KONTO 3590 "Påminnelseavgifter"',
        f'#KONTO 5410 "Förbrukningsinventarier"',
        f'#KONTO 5612 "Bränsle och drivmedel"',
        f'#KONTO 6990 "Övriga kostnader"',
        f'#KONTO 7210 "Löner"',
        f'#KONTO 7290 "Semesterersättning"',
        f'#KONTO 7331 "Milersättning"',
        f'#KONTO 7510 "Arbetsgivaravgifter"',
        f'',
    ]

    # Add IB (ingående balans) - simplified
    sie_lines.append(f'#IB 0 1930 0.00')
    sie_lines.append(f'')

    # Add verifications
    for ver in verifikationer:
        sie_lines.append(ver)

    # Add UB (utgående balans)
    sie_lines.append(f'')
    sie_lines.append(f'#UB 0 1510 {sum(i.get("customer_pays",0) for i in invoices if i.get("status") != "paid"):.2f}')

    sie_content = "\n".join(sie_lines)

    # Encode to UTF-8
    sie_bytes = sie_content.encode("utf-8")

    return Response(
        content=sie_bytes,
        media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="sie_{month}.se"'}
    )


# ── AGI XML Export ────────────────────────────────────────────────────────────
@api_router.get("/reports/agi")
async def export_agi(month: str, current=Depends(get_current_user)):
    """Export AGI (Arbetsgivardeklaration) XML for Skatteverket
    month format: YYYY-MM
    """
    import calendar
    year, mon = month.split("-")
    last_day = calendar.monthrange(int(year), int(mon))[1]
    start = f"{month}-01"
    end = f"{month}-{last_day:02d}"

    ARBETSGIVARAVGIFT = 0.3142
    PRELSKATT = 0.30

    inv_settings = await get_invoice_settings_obj()
    orgnr = (inv_settings.company_orgnr or "").replace("-","")
    company = inv_settings.company_name or "PureNorth Städ"

    payroll_summary, payroll_settings = await build_payroll_summary(start, end)
    employees = await db.employees.find().to_list(100)
    emp_map = {str(e["_id"]): e for e in employees}

    now = datetime.now(timezone.utc)
    period = month.replace("-","")  # YYYYMM

    # Build XML
    lines = []
    lines.append('<?xml version="1.0" encoding="UTF-8"?>')
    lines.append('<Skatteverket xmlns="http://www.skatteverket.se/schema/instans/arbetsgivardeklaration" xmlns:agd="http://www.skatteverket.se/schema/instans/arbetsgivardeklaration">')
    lines.append(f'  <Avsandare>')
    lines.append(f'    <ProgramnameProduktnr>{company} Admin 1.0</ProgramnameProduktnr>')
    lines.append(f'    <Tidpunkt>{now.strftime("%Y-%m-%dT%H:%M:%S")}</Tidpunkt>')
    lines.append(f'  </Avsandare>')
    lines.append(f'  <Blankettgemensamt>')
    lines.append(f'    <Arbetsgivare>')
    lines.append(f'      <AgRegistreradId>{orgnr}</AgRegistreradId>')
    lines.append(f'    </Arbetsgivare>')
    lines.append(f'  </Blankettgemensamt>')

    total_ag_avg = 0
    total_prelskatt = 0

    for eid, row in payroll_summary.items():
        emp = emp_map.get(eid, {})
        personnr = emp.get("personnummer", "").replace("-","")
        name = row["name"]
        brutto = (
            row["normal_h"] * row["hourly_rate"] +
            row["ob1_h"] * payroll_settings.ob1_extra +
            row["ob2_h"] * payroll_settings.ob2_extra +
            row.get("sjuklon_net", 0)
        )
        prelskatt = round(brutto * PRELSKATT, 2)
        ag_avg = round(brutto * ARBETSGIVARAVGIFT, 2)
        total_ag_avg += ag_avg
        total_prelskatt += prelskatt

        if not personnr:
            continue

        lines.append(f'  <Blankett>')
        lines.append(f'    <Arendeinformation>')
        lines.append(f'      <Arendeagare>{orgnr}</Arendeagare>')
        lines.append(f'      <Period>{period}</Period>')
        lines.append(f'    </Arendeinformation>')
        lines.append(f'    <Blankettinnehall>')
        lines.append(f'      <HU>')
        lines.append(f'        <ArbetsgivareHUGROUP>')
        lines.append(f'          <AgRegistreradId faltkod="201">{orgnr}</AgRegistreradId>')
        lines.append(f'        </ArbetsgivareHUGROUP>')
        lines.append(f'        <RedovisningsPeriod faltkod="006">{period}</RedovisningsPeriod>')
        lines.append(f'        <SummaArbAvgSlf faltkod="487">{int(total_ag_avg)}</SummaArbAvgSlf>')
        lines.append(f'        <SummaSkatteavdr faltkod="497">{int(total_prelskatt)}</SummaSkatteavdr>')
        lines.append(f'      </HU>')
        lines.append(f'      <IU>')
        lines.append(f'        <ArbetsgivareIUGROUP>')
        lines.append(f'          <AgRegistreradId faltkod="201">{orgnr}</AgRegistreradId>')
        lines.append(f'        </ArbetsgivareIUGROUP>')
        lines.append(f'        <BetalningsmottagareIUGROUP>')
        lines.append(f'          <Fodelsetid faltkod="215">{personnr}</Fodelsetid>')
        lines.append(f'        </BetalningsmottagareIUGROUP>')
        lines.append(f'        <RedovisningsPeriod faltkod="006">{period}</RedovisningsPeriod>')
        lines.append(f'        <Specifikationsnummer faltkod="570">001</Specifikationsnummer>')
        lines.append(f'        <AvdrPrelSkatt faltkod="001">{int(prelskatt)}</AvdrPrelSkatt>')
        lines.append(f'        <KontantErsattningKU10 faltkod="011">{int(brutto)}</KontantErsattningKU10>')
        lines.append(f'      </IU>')
        lines.append(f'    </Blankettinnehall>')
        lines.append(f'  </Blankett>')

    lines.append('</Skatteverket>')

    xml_content = "\n".join(lines)

    return Response(
        content=xml_content.encode("utf-8"),
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="agi_{month}.xml"'}
    )


# ── Momsdeklaration PDF ───────────────────────────────────────────────────────
@api_router.get("/reports/moms")
async def export_moms_pdf(month: str, current=Depends(get_current_user)):
    """Export Momsdeklaration PDF in Skatteverket format
    month format: YYYY-MM
    """
    import calendar
    year, mon = month.split("-")
    last_day = calendar.monthrange(int(year), int(mon))[1]
    start = f"{month}-01"
    end = f"{month}-{last_day:02d}"
    start_dt = f"{start}T00:00:00"
    end_dt = f"{end}T23:59:59"

    inv_settings = await get_invoice_settings_obj()
    company = inv_settings.company_name or "PureNorth Städ"
    orgnr = inv_settings.company_orgnr or ""

    MONTHS_SV = ["","Januari","Februari","Mars","April","Maj","Juni","Juli","Augusti","September","Oktober","November","December"]
    month_name = MONTHS_SV[int(mon)]

    # Collect data
    invoices_raw = await db.invoices.find({"created_at": {"$gte": start_dt, "$lte": end_dt}}).to_list(2000)
    invoices = [recalc_invoice(i) for i in invoices_raw]
    costs = await db.costs.find({"date": {"$gte": start, "$lte": end}}).to_list(1000)
    expenses = await db.expenses.find({"date": {"$gte": start, "$lte": end}}).to_list(1000)

    # Calculate moms
    # Utgående moms (from invoices)
    sales_excl = sum(i.get("subtotal", 0) for i in invoices)
    paminnelse_fees = sum(
        sum(item.get("quantity",1)*item.get("unit_price",0)
            for item in inv.get("items",[])
            if "Påminnelseavgift" in item.get("service",""))
        for inv in invoices
    )
    total_sales = sales_excl + paminnelse_fees
    vat_out_25 = sum(i.get("vat_amount", 0) for i in invoices)

    # Ingående moms (from costs + expenses)
    vat_in_mat = sum(c.get("amount",0) - c.get("amount",0)/(1+c.get("moms_rate",25)/100) for c in costs if c.get("moms_rate",25) > 0)
    vat_in_exp = sum(e.get("amount",0) - e.get("amount",0)/(1+e.get("moms_rate",0)/100) for e in expenses if e.get("moms_rate",0) > 0)
    vat_in_total = round(vat_in_mat + vat_in_exp, 2)

    moms_to_pay = round(vat_out_25 - vat_in_total, 2)

    # Build PDF
    buf = BytesIO()
    page_w, page_h = A4
    doc = SimpleDocTemplate(buf, pagesize=A4,
        topMargin=15*mm, bottomMargin=15*mm, leftMargin=20*mm, rightMargin=20*mm,
        title=f"Momsdeklaration {month}",
            author="PureNorth Städ")

    styles = getSampleStyleSheet()
    def ps(name, **kw): return ParagraphStyle(name, parent=styles["Normal"], **kw)

    elements = []

    # ── HEADER ─────────────────────────────────────────────────────
    hdr = Table([[
        Paragraph(f"<b>MOMSDEKLARATION</b>", ps("h", fontSize=16, fontName="Helvetica-Bold", textColor=colors.white)),
        Paragraph(f"{month_name} {year}", ps("s", fontSize=11, textColor=colors.white, alignment=2))
    ]], colWidths=[120*mm, None])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#141414")),
        ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),
        ("TOPPADDING",(0,0),(-1,-1),12),("BOTTOMPADDING",(0,0),(-1,-1),12),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    elements.append(hdr)

    # Company info
    elements.append(Paragraph(
        f"{company}  ·  Org.nr: {orgnr}  ·  Redovisningsperiod: {start} – {end}",
        ps("info", fontSize=8, textColor=colors.HexColor("#64748b"), spaceBefore=4, spaceAfter=6)
    ))

    # Divider
    div = Table([[""]], colWidths=[page_w-40*mm])
    div.setStyle(TableStyle([("LINEBELOW",(0,0),(-1,-1),1,colors.HexColor("#141414")),
        ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0)]))
    elements.append(div)
    elements.append(Spacer(1,5*mm))

    def row(ruta, label, value, bold=False, color=None):
        val_color = colors.HexColor(color) if color else colors.HexColor("#141414")
        fn = "Helvetica-Bold" if bold else "Helvetica"
        data = [[
            Paragraph(f"<b>Ruta {ruta}</b>", ps("r", fontSize=8, textColor=colors.HexColor("#94a3b8"))),
            Paragraph(label, ps("l", fontSize=9, fontName=fn, textColor=colors.HexColor("#1a1a1a"))),
            Paragraph(f"{value:,.2f} kr".replace(",","."), ps("v", fontSize=9, fontName=fn, textColor=val_color, alignment=2))
        ]]
        t = Table(data, colWidths=[18*mm, 120*mm, None])
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#F8FAFC") if bold else colors.white),
            ("LINEBELOW",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
            ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
        ]))
        elements.append(t)

    def section(title, color="#141414"):
        t = Table([[Paragraph(title, ps("sh", fontSize=9, fontName="Helvetica-Bold", textColor=colors.white))]], colWidths=[page_w-40*mm])
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),colors.HexColor(color)),
            ("LEFTPADDING",(0,0),(-1,-1),8),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
        ]))
        elements.append(t)

    # ── A. MOMS ATT REDOVISA ────────────────────────────────────────
    section("A. MOMS ATT REDOVISA (UTGÅENDE MOMS)", "#1e40af")
    row("05", "Momspliktig försäljning (exkl. moms)", sales_excl)
    if paminnelse_fees > 0:
        row("06", "Påminnelseavgifter (momsfria intäkter)", paminnelse_fees)
    row("10", "Utgående moms 25%", vat_out_25)
    row("11", "Utgående moms 12%", 0)
    row("12", "Utgående moms 6%", 0)

    elements.append(Spacer(1,3*mm))

    # ── B. AVDRAGSGILL INGÅENDE MOMS ───────────────────────────────
    section("B. AVDRAGSGILL INGÅENDE MOMS", "#0369a1")
    row("48", "Ingående moms att dra av", vat_in_total, color="#15803d")

    elements.append(Spacer(1,3*mm))

    # ── SUMMARY ────────────────────────────────────────────────────
    section("SAMMANFATTNING - ATT BETALA/FÅ TILLBAKA", "#b45309")
    row("49", "Moms att betala (+) eller få tillbaka (−)",
        moms_to_pay, bold=True, color="#dc2626" if moms_to_pay > 0 else "#15803d")

    elements.append(Spacer(1,5*mm))

    # ── PAYMENT INFO ───────────────────────────────────────────────
    if moms_to_pay > 0:
        pay_info = f"Betala {moms_to_pay:,.2f} kr till Skatteverket · Bankgiro: 5050-1055 · Ref: {orgnr} · Senast 26:e".replace(",",".")
        pay_color = "#dc2626"
    else:
        pay_info = f"Du får tillbaka {abs(moms_to_pay):,.2f} kr från Skatteverket".replace(",",".")
        pay_color = "#15803d"

    pt = Table([[Paragraph(pay_info, ps("pi", fontSize=9, fontName="Helvetica-Bold", textColor=colors.white))]], colWidths=[page_w-40*mm])
    pt.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),colors.HexColor(pay_color)),
        ("LEFTPADDING",(0,0),(-1,-1),10),("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8),
    ]))
    elements.append(pt)

    elements.append(Spacer(1,5*mm))

    # ── INVOICE BREAKDOWN ──────────────────────────────────────────
    if invoices:
        section("FAKTUROR I PERIODEN", "#475569")
        inv_data = [["Faktura #", "Kund", "Datum", "Exkl. moms", "Påm.avg", "Moms 25%", "Att betala"]]
        for i in sorted(invoices, key=lambda x: x.get("created_at","")):
            pam = sum(item.get("quantity",1)*item.get("unit_price",0)
                for item in i.get("items",[])
                if "Påminnelseavgift" in item.get("service",""))
            inv_data.append([
                i.get("invoice_number",""),
                i.get("customer_name","")[:20],
                i.get("created_at","")[:10],
                f'{i.get("subtotal",0):.2f}',
                f'{pam:.2f}' if pam > 0 else "-",
                f'{i.get("vat_amount",0):.2f}',
                f'{i.get("customer_pays",0):.2f}',
            ])
        inv_tbl = Table(inv_data, colWidths=[18*mm, 48*mm, 20*mm, 24*mm, 16*mm, 20*mm, 24*mm])
        inv_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#F1F5F9")),
            ("FONTSIZE",(0,0),(-1,-1),8),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
            ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("ALIGN",(3,0),(-1,-1),"RIGHT"),
        ]))
        elements.append(inv_tbl)

    elements.append(Spacer(1,4*mm))

    # ── FOOTER ─────────────────────────────────────────────────────
    note = "Denna deklaration är baserad på fakturametoden. Kontrollera med din redovisningskonsult innan du skickar till Skatteverket."
    elements.append(Paragraph(note, ps("note", fontSize=7, textColor=colors.HexColor("#94a3b8"))))
    elements.append(Paragraph(
        f"{company}  ·  Org.nr {orgnr}  ·  Momsdeklaration {month_name} {year}  ·  Skapad {datetime.now(timezone.utc).strftime('%Y-%m-%d')}",
        ps("ft", fontSize=7, textColor=colors.HexColor("#94a3b8"), spaceBefore=3)
    ))

    doc.build(elements)
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="momsdeklaration_{month}.pdf"'}
    )


# ── Bokföringsunderlag PDF ────────────────────────────────────────────────────
@api_router.get("/reports/bokforing")
async def export_bokforing_pdf(month: str, current=Depends(get_current_user)):
    """Export Bokföringsunderlag with BAS account codes per transaction"""
    import calendar
    year, mon = month.split("-")
    last_day = calendar.monthrange(int(year), int(mon))[1]
    start = f"{month}-01"
    end = f"{month}-{last_day:02d}"
    start_dt = f"{start}T00:00:00"
    end_dt = f"{end}T23:59:59"

    inv_settings = await get_invoice_settings_obj()
    company = inv_settings.company_name or "PureNorth Städ"
    orgnr = inv_settings.company_orgnr or ""

    MONTHS_SV = ["","Januari","Februari","Mars","April","Maj","Juni","Juli","Augusti","September","Oktober","November","December"]
    month_name = MONTHS_SV[int(mon)]

    invoices = await db.invoices.find({"created_at": {"$gte": start_dt, "$lte": end_dt}}).to_list(2000)
    costs = await db.costs.find({"date": {"$gte": start, "$lte": end}}).to_list(1000)
    expenses = await db.expenses.find({"date": {"$gte": start, "$lte": end}}).to_list(1000)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        topMargin=15*mm, bottomMargin=15*mm, leftMargin=20*mm, rightMargin=20*mm,
        title=f"Bokföringsunderlag {month}",
            author="PureNorth Städ")
    styles = getSampleStyleSheet()
    def ps(name, **kw): return ParagraphStyle(name, parent=styles["Normal"], **kw)
    page_w = A4[0]
    elements = []

    # Header
    hdr = Table([[
        Paragraph("<b>BOKFÖRINGSUNDERLAG</b>", ps("h", fontSize=14, fontName="Helvetica-Bold", textColor=colors.white)),
        Paragraph(f"{month_name} {year}", ps("s", fontSize=10, textColor=colors.white, alignment=2))
    ]], colWidths=[120*mm, None])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#141414")),
        ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),
        ("TOPPADDING",(0,0),(-1,-1),10),("BOTTOMPADDING",(0,0),(-1,-1),10),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    elements.append(hdr)
    elements.append(Paragraph(
        f"{company}  ·  Org.nr: {orgnr}  ·  Period: {start} – {end}  ·  Skapad: {datetime.now(timezone.utc).strftime('%Y-%m-%d')}",
        ps("info", fontSize=8, textColor=colors.HexColor("#64748b"), spaceBefore=3, spaceAfter=6)
    ))

    def section_hdr(title, color="#1e40af"):
        t = Table([[Paragraph(title, ps("sh", fontSize=9, fontName="Helvetica-Bold", textColor=colors.white))]], colWidths=[page_w-40*mm])
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),colors.HexColor(color)),
            ("LEFTPADDING",(0,0),(-1,-1),8),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
        ]))
        elements.append(t)

    def trans_table(rows):
        """rows: list of [konto, kontonamn, debet, kredit, beskrivning]"""
        hrow = [["Konto", "Kontonamn", "Debet (kr)", "Kredit (kr)", "Beskrivning"]]
        data = hrow + rows
        t = Table(data, colWidths=[18*mm, 55*mm, 28*mm, 28*mm, None])
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#F1F5F9")),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),8),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
            ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("ALIGN",(2,0),(3,-1),"RIGHT"),
            ("TEXTCOLOR",(2,1),(2,-1),colors.HexColor("#1e40af")),
            ("TEXTCOLOR",(3,1),(3,-1),colors.HexColor("#dc2626")),
        ]))
        elements.append(t)

    # ── 1. FAKTUROR ────────────────────────────────────────────────
    elements.append(Spacer(1,3*mm))
    section_hdr("1. FAKTUROR (KUNDFORDRINGAR)", "#1e40af")

    for inv in sorted(invoices, key=lambda x: x.get("created_at","")):
        inv_num = inv.get("invoice_number","")
        customer = inv.get("customer_name","")
        date = inv.get("created_at","")[:10]
        status = inv.get("status","")
        items = inv.get("items", [])
        rut = inv.get("rut_deduction", 0)
        reminder_count = inv.get("reminder_count", 0)

        # Calculate excl. Påminnelseavgift (which has no moms)
        work_items = [i for i in items if i.get("service") != "Påminnelseavgift"]
        sub = sum(i.get("quantity",1) * i.get("unit_price",0) for i in work_items)
        vat_rate = 25
        vat = round(sub * vat_rate / 100, 2)
        pam_fee_total = sum(i.get("quantity",1)*i.get("unit_price",0) for i in items if i.get("service") == "Påminnelseavgift")
        pays = round(sub + vat - rut + pam_fee_total, 2)

        # Check for reminder fee
        reminder_fee_item = next((i for i in items if i.get("service") == "Påminnelseavgift"), None)

        elements.append(Spacer(1,2*mm))
        elements.append(Paragraph(
            f"<b>Faktura #{inv_num}</b>  ·  {customer}  ·  {date}  ·  Status: {status}" +
            (f"  ·  Påminnelse #{reminder_count}" if reminder_count > 0 else ""),
            ps("fh", fontSize=8, fontName="Helvetica-Bold", textColor=colors.HexColor("#1e40af"), spaceBefore=2, spaceAfter=1)
        ))

        rows = [
            ["1510", "Kundfordringar", f"{pays:.2f}", "", f"Faktura #{inv_num} - {customer}"],
        ]
        # Regular items
        for item in items:
            if item.get("service") == "Påminnelseavgift":
                continue
            item_total = item.get("quantity",1) * item.get("unit_price",0)
            rows.append(["3000", item.get("service","Försäljning"), "", f"{item_total:.2f}", "Intäkt exkl. moms"])

        rows.append(["2610", "Utgående moms 25%", "", f"{vat:.2f}", "Moms på försäljning"])

        if rut > 0:
            rows.append(["3001", "RUT-avdrag", f"{rut:.2f}", "", "RUT-reduktion"])

        # Reminder fee as separate entry
        if reminder_fee_item:
            fee = reminder_fee_item.get("quantity",1) * reminder_fee_item.get("unit_price",0)
            rows.append(["3590", "Påminnelseavgifter", "", f"{fee:.2f}", f"Påminnelseavgift #{reminder_count}"])

        trans_table(rows)

    # ── 2. MATERIALKOSTNADER ───────────────────────────────────────
    if costs:
        elements.append(Spacer(1,4*mm))
        section_hdr("2. MATERIALKOSTNADER", "#0369a1")
        cat_konto = {"material":"5410","equipment":"5420","transport":"5612","other":"6990","parking":"5651"}
        cat_name = {"material":"Förbrukningsinventarier","equipment":"Maskiner","transport":"Bränsle","other":"Övriga kostnader","parking":"Parkering"}

        for c in sorted(costs, key=lambda x: x.get("date","")):
            amount = c.get("amount", 0)
            mrate = c.get("moms_rate", 25)
            vat_c = round(amount - amount/(1+mrate/100), 2) if mrate > 0 else 0
            net = round(amount - vat_c, 2)
            cat = c.get("category","material")
            konto = cat_konto.get(cat,"5410")
            kname = cat_name.get(cat,"Kostnad")
            name = c.get("name","")
            date = c.get("date","")

            elements.append(Spacer(1,2*mm))
            elements.append(Paragraph(
                f"<b>{name}</b>  ·  {date}  ·  {c.get('category','')}",
                ps("ch", fontSize=8, fontName="Helvetica-Bold", textColor=colors.HexColor("#0369a1"), spaceBefore=2, spaceAfter=1)
            ))
            rows = [
                [konto, kname, f"{net:.2f}", "", f"Kostnad exkl. moms"],
                ["2640", "Ingående moms", f"{vat_c:.2f}", "", f"Moms {mrate}%"],
                ["1930", "Företagskonto", "", f"{amount:.2f}", "Betalning"],
            ]
            trans_table(rows)

    # ── 3. UTLÄGG ──────────────────────────────────────────────────
    if expenses:
        elements.append(Spacer(1,4*mm))
        section_hdr("3. UTLÄGG (ANSTÄLLDA)", "#7c3aed")
        exp_konto = {"Material":"5410","Bränsle":"5612","Milersättning":"7331","Parkering":"5651","Övrigt":"6990"}

        for e in sorted(expenses, key=lambda x: x.get("date","")):
            amount = e.get("amount", 0)
            mrate = e.get("moms_rate", 0)
            vat_e = round(amount - amount/(1+mrate/100), 2) if mrate > 0 else 0
            net = round(amount - vat_e, 2)
            cat = e.get("category","Övrigt")
            konto = exp_konto.get(cat,"6990")
            name = e.get("description","") or cat
            date = e.get("date","")

            elements.append(Spacer(1,2*mm))
            elements.append(Paragraph(
                f"<b>{name}</b>  ·  {date}  ·  {cat}",
                ps("eh", fontSize=8, fontName="Helvetica-Bold", textColor=colors.HexColor("#7c3aed"), spaceBefore=2, spaceAfter=1)
            ))
            rows = [
                [konto, cat, f"{net:.2f}", "", "Utlägg exkl. moms"],
            ]
            if vat_e > 0:
                rows.append(["2640", "Ingående moms", f"{vat_e:.2f}", "", f"Moms {mrate}%"])
            rows.append(["2890", "Skuld anställd", "", f"{amount:.2f}", "Att ersätta"])
            trans_table(rows)

    # Footer
    elements.append(Spacer(1,4*mm))
    elements.append(Paragraph(
        "Bokföringsunderlag enligt BAS-kontoplanen 2025. Kontrollera med din redovisningskonsult.",
        ps("ft", fontSize=7, textColor=colors.HexColor("#94a3b8"))
    ))

    doc.build(elements)
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="bokforingsunderlag_{month}.pdf"'}
    )


# ── Invoice Reminder ──────────────────────────────────────────────────────────
@api_router.post("/invoices/{invoice_id}/remind")
async def send_invoice_reminder(invoice_id: str, current=Depends(get_current_user)):
    """Send payment reminder to customer for overdue invoice"""
    doc = await db.invoices.find_one({"_id": to_object_id(invoice_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Faktura hittades inte")

    customer_email = doc.get("customer_email", "")
    if not customer_email:
        raise HTTPException(status_code=400, detail="Kunden har ingen e-postadress")

    inv_settings = await get_invoice_settings_obj()
    company = inv_settings.company_name or "PureNorth Städ"
    inv_num = doc.get("invoice_number", "")
    customer_name = doc.get("customer_name", "")
    due_date = doc.get("due_date", "")
    amount = doc.get("customer_pays", 0)
    reminder_count = doc.get("reminder_count", 0) + 1

    # Calculate days overdue
    try:
        from datetime import date as DateObj
        due = DateObj.fromisoformat(due_date)
        days_overdue = (DateObj.today() - due).days
    except:
        days_overdue = 0

    # Reminder fee (Swedish law: 60 kr after first reminder)
    reminder_fee = 60 if reminder_count >= 2 else 0
    total_with_fee = amount + reminder_fee

    resend.api_key = os.environ.get("RESEND_API_KEY", "")
    if not resend.api_key:
        raise HTTPException(status_code=500, detail="E-posttjänst ej konfigurerad")

    admin_email = os.environ.get("ADMIN_EMAIL", "akhazzane.othmane@gmail.com")

    subject = f"Påminnelse #{reminder_count}: Faktura #{inv_num} – {company}"
    if reminder_count >= 3:
        subject = f"Sista betalningspåminnelse: Faktura #{inv_num} – {company}"

    html = f"""
    <div style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;max-width:520px;margin:0 auto;background:#fff;">
      <div style="background:#dc2626;padding:28px 36px;">
        <h1 style="color:#fff;font-size:20px;margin:0;font-weight:700;">{company}</h1>
        <p style="color:rgba(255,255,255,0.8);margin:6px 0 0;font-size:13px;letter-spacing:0.05em;text-transform:uppercase;">
          {'Påminnelse #' + str(reminder_count) if reminder_count < 3 else 'Sista betalningspåminnelse'}
        </p>
      </div>
      <div style="padding:36px;border:1px solid #e2e8f0;border-top:none;">
        <p style="font-size:17px;color:#141414;margin:0 0 8px;font-weight:600;">Hej {customer_name},</p>
        <p style="color:#64748b;font-size:15px;line-height:1.7;margin:0 0 24px;">
          Vi har inte mottagit betalning för nedanstående faktura som förföll för
          <strong style="color:#dc2626;">{days_overdue} dagar sedan</strong>.
          Vänligen betala snarast möjligt för att undvika inkasso.
        </p>
        <div style="background:#fef2f2;border-radius:8px;padding:20px;margin:0 0 24px;border-left:3px solid #dc2626;">
          <p style="margin:0 0 8px;font-size:14px;"><span style="color:#64748b;">Fakturanummer</span> &nbsp; <strong>#{inv_num}</strong></p>
          <p style="margin:0 0 8px;font-size:14px;"><span style="color:#64748b;">Förfallodatum</span> &nbsp; <strong>{due_date}</strong></p>
          <p style="margin:0 0 8px;font-size:14px;"><span style="color:#64748b;">Ursprungligt belopp</span> &nbsp; <strong>{amount:.2f} kr</strong></p>
          {'<p style="margin:0 0 8px;font-size:14px;"><span style="color:#64748b;">Påminnelseavgift</span> &nbsp; <strong>' + str(reminder_fee) + ' kr</strong></p>' if reminder_fee > 0 else ''}
          <p style="margin:0;font-size:16px;font-weight:700;color:#dc2626;">Att betala: {total_with_fee:.2f} kr</p>
        </div>
        {'<div style="background:#fef2f2;border:1px solid #fecaca;border-radius:6px;padding:12px 16px;margin:0 0 16px;"><p style="color:#dc2626;font-size:13px;font-weight:600;margin:0;">Om betalning inte inkommer inom 10 dagar lämnas ärendet till inkassobolag för vidare åtgärd.</p></div>' if reminder_count >= 2 else ''}
        <p style="color:#64748b;font-size:14px;margin:0 0 16px;">
          Vid frågor, kontakta oss gärna.
        </p>
        <hr style="border:none;border-top:1px solid #e2e8f0;margin:0 0 20px;"/>
        <p style="color:#94a3b8;font-size:12px;text-align:center;margin:0;">
          {company} · Miljövänlig städning i Umeå
        </p>
      </div>
    </div>
    """

    try:
        # Generate PDF
        pdf_bytes = build_invoice_pdf(doc, inv_settings)
        import base64
        pdf_b64 = base64.b64encode(pdf_bytes).decode()

        resend.Emails.send({
            "from": f"{company} <onboarding@resend.dev>",
            "to": admin_email,
            "subject": subject,
            "html": html,
            "attachments": [{
                "filename": f"faktura_{inv_num}.pdf",
                "content": pdf_b64,
            }]
        })

        # Update invoice - add reminder fee if reminder_count >= 2
        update_fields = {
            "status": "overdue",
            "reminder_count": reminder_count,
            "last_reminder_at": datetime.now(timezone.utc).isoformat()
        }

        if reminder_fee > 0:
            # Add påminnelseavgift to invoice items (NO moms on reminder fee)
            items = doc.get("items", [])
            # Remove old reminder fee if exists
            items = [i for i in items if i.get("service") != "Påminnelseavgift"]
            # Add new reminder fee
            items.append({
                "service": "Påminnelseavgift",
                "description": "Påminnelseavgift enligt inkassolagen (ingen moms)",
                "quantity": 1,
                "unit_price": reminder_fee,
                "is_material": False,
            })

            # Recalculate: reminder fee is AFTER moms and RUT
            work_items = [i for i in items if i.get("service") != "Påminnelseavgift"]
            subtotal = sum(i["quantity"] * i["unit_price"] for i in work_items)
            vat_rate = doc.get("vat_amount", 0) / doc.get("subtotal", 1) * 100 if doc.get("subtotal", 0) > 0 else 25
            vat_amount = round(subtotal * vat_rate / 100, 2)
            rut_deduction = doc.get("rut_deduction", 0)
            total_excl_reminder = round(subtotal + vat_amount - rut_deduction, 2)
            customer_pays = round(total_excl_reminder + reminder_fee, 2)
            total_amount = round(subtotal + vat_amount + reminder_fee, 2)

            update_fields["items"] = items
            update_fields["subtotal"] = round(subtotal, 2)
            update_fields["vat_amount"] = vat_amount
            update_fields["total_amount"] = total_amount
            update_fields["customer_pays"] = customer_pays

        await db.invoices.update_one(
            {"_id": to_object_id(invoice_id)},
            {"$set": update_fields}
        )

        return {
            "success": True,
            "reminder_count": reminder_count,
            "reminder_fee": reminder_fee,
            "message": f"Påminnelse #{reminder_count} skickad!"
        }

    except Exception as e:
        logger.error(f"Failed to send reminder: {e}")
        raise HTTPException(status_code=500, detail=f"Kunde inte skicka påminnelse: {str(e)}")





# ── RUT Report ────────────────────────────────────────────────────────────────
@api_router.get("/reports/rut")
async def rut_report_pdf(month: str, current=Depends(get_current_user)):
    """Export RUT report for Skatteverket application"""
    import calendar
    year, mon = month.split("-")
    last_day = calendar.monthrange(int(year), int(mon))[1]
    start = f"{month}-01"
    end = f"{month}-{last_day:02d}"
    start_dt = f"{start}T00:00:00"
    end_dt = f"{end}T23:59:59"

    inv_settings = await get_invoice_settings_obj()
    company = inv_settings.company_name or "PureNorth Städ"
    orgnr = inv_settings.company_orgnr or ""

    MONTHS_SV = ["","Januari","Februari","Mars","April","Maj","Juni","Juli","Augusti","September","Oktober","November","December"]
    month_name = MONTHS_SV[int(mon)]

    # Get invoices with RUT
    invoices = await db.invoices.find({
        "created_at": {"$gte": start_dt, "$lte": end_dt},
        "rut_deduction": {"$gt": 0}
    }).to_list(2000)

    total_rut = sum(i.get("rut_deduction", 0) for i in invoices)
    total_paid = sum(i.get("rut_deduction", 0) for i in invoices if i.get("status") == "paid")
    total_pending = total_rut - total_paid

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        topMargin=15*mm, bottomMargin=15*mm, leftMargin=20*mm, rightMargin=20*mm,
        title=f"RUT-rapport {month}",
            author="PureNorth Städ")
    styles = getSampleStyleSheet()
    def ps(name, **kw): return ParagraphStyle(name, parent=styles["Normal"], **kw)
    page_w = A4[0]
    elements = []

    # Header
    hdr = Table([[
        Paragraph(f"<b>{company}</b>", ps("h", fontSize=14, fontName="Helvetica-Bold", textColor=colors.white)),
        Paragraph(f"<b>RUT-AVDRAG RAPPORT</b><br/>{month_name} {year}", ps("s", fontSize=10, fontName="Helvetica-Bold", textColor=colors.white, alignment=2))
    ]], colWidths=[100*mm, None])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#15803d")),
        ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),
        ("TOPPADDING",(0,0),(-1,-1),12),("BOTTOMPADDING",(0,0),(-1,-1),12),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    elements.append(hdr)
    elements.append(Paragraph(
        f"{company}  ·  Org.nr: {orgnr}  ·  Period: {start} – {end}  ·  Skapad: {datetime.now(timezone.utc).strftime('%Y-%m-%d')}",
        ps("info", fontSize=8, textColor=colors.HexColor("#64748b"), spaceBefore=4, spaceAfter=6)
    ))

    # Summary boxes
    summary_data = [[
        Paragraph(f"<b>Totalt RUT att ansöka</b><br/><br/>{total_rut:.2f} kr", ps("sb", fontSize=9, textColor=colors.HexColor("#15803d"))),
        Paragraph(f"<b>Antal fakturor</b><br/><br/>{len(invoices)} st", ps("sb", fontSize=9, textColor=colors.HexColor("#141414"))),
        Paragraph(f"<b>Ansökt (betalda)</b><br/><br/>{total_paid:.2f} kr", ps("sb", fontSize=9, textColor=colors.HexColor("#1e40af"))),
    ]]
    summary_tbl = Table(summary_data, colWidths=[(page_w-40*mm)/3]*3)
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(0,0),colors.HexColor("#f0fdf4")),
        ("BACKGROUND",(1,0),(1,0),colors.HexColor("#f8fafc")),
        ("BACKGROUND",(2,0),(2,0),colors.HexColor("#eff6ff")),
        ("BOX",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),
        ("INNERGRID",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),
        ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),
        ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8),
    ]))
    elements.append(summary_tbl)
    elements.append(Spacer(1,5*mm))

    # How to apply
    elements.append(Paragraph("HUR DU ANSÖKER OM RUT-AVDRAG", ps("sh", fontSize=9, fontName="Helvetica-Bold", spaceBefore=4, spaceAfter=2)))
    steps = [
        "1. Gå till skatteverket.se → Logga in med BankID",
        "2. Välj 'ROT och RUT-avdrag' → 'Begär utbetalning'",
        "3. Fyll i kundens personnummer och RUT-belopp per faktura",
        "4. Skatteverket betalar ut inom 2-4 veckor",
    ]
    for step in steps:
        elements.append(Paragraph(step, ps("s", fontSize=8, textColor=colors.HexColor("#475569"), spaceBefore=1)))
    elements.append(Spacer(1,5*mm))

    # Invoice list
    elements.append(Paragraph("FAKTUROR MED RUT-AVDRAG", ps("sh", fontSize=9, fontName="Helvetica-Bold", spaceBefore=2, spaceAfter=2)))

    if not invoices:
        elements.append(Paragraph("Inga fakturor med RUT-avdrag denna period.", ps("s", fontSize=9, textColor=colors.HexColor("#94a3b8"))))
    else:
        inv_data = [["Faktura #", "Kund", "Personnummer", "Tjänst", "Arbetskostnad", "RUT 50%", "Status"]]
        for inv in sorted(invoices, key=lambda x: x.get("created_at","")):
            work_items = [i for i in inv.get("items",[]) if "Påminnelseavgift" not in i.get("service","")]
            svc_names = set()
            for it in work_items:
                svc = it.get("service","") or it.get("description","")
                if svc: svc_names.add(svc.split("(")[0].strip())
            services = ", ".join(svc_names) or "Städtjänst"
            rut = inv.get("rut_deduction", 0)
            labor = rut * 2
            personnr = inv.get("customer_personnummer", "") or "—"
            status = "Betald" if inv.get("status") == "paid" else "Obetald"
            inv_data.append([
                str(inv.get("invoice_number","")),
                inv.get("customer_name","")[:18],
                personnr,
                services[:22],
                f"{labor:.2f}",
                f"{rut:.2f}",
                status,
            ])

        # Totals row
        inv_data.append(["", "TOTALT", "", "", f"{total_rut*2:,.2f}".replace(",","."), f"{total_rut:,.2f}".replace(",","."), ""])

        inv_tbl = Table(inv_data, colWidths=[15*mm, 35*mm, 28*mm, 32*mm, 24*mm, 20*mm, 18*mm])
        inv_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#15803d")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#f0fdf4")),
            ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),8),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
            ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("ALIGN",(4,0),(5,-1),"RIGHT"),
        ]))
        elements.append(inv_tbl)

    elements.append(Spacer(1,5*mm))
    elements.append(Paragraph(
        f"OBS: Kunden måste ha betalat sin del innan du kan ansöka om RUT. Kontrollera att personnummer är korrekt.",
        ps("note", fontSize=7, textColor=colors.HexColor("#94a3b8"))
    ))

    doc.build(elements)
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="rut_rapport_{month}.pdf"'}
    )


# ── Kostnader PDF Export ──────────────────────────────────────────────────────
@api_router.get("/costs/report-pdf")
async def costs_report_pdf(month: str, current=Depends(get_current_user)):
    """Export Kostnader PDF with monthly comparison"""
    import calendar
    year, mon = month.split("-")
    last_day = calendar.monthrange(int(year), int(mon))[1]
    start = f"{month}-01"
    end = f"{month}-{last_day:02d}"

    inv_settings = await get_invoice_settings_obj()
    company = inv_settings.company_name or "PureNorth Städ"
    MONTHS_SV = ["","Januari","Februari","Mars","April","Maj","Juni","Juli","Augusti","September","Oktober","November","December"]
    month_name = MONTHS_SV[int(mon)]

    # Current month costs
    costs = await db.costs.find({"date": {"$gte": start, "$lte": end}}).to_list(1000)
    total = sum(c.get("amount",0) for c in costs)
    ingaende_moms = sum(c.get("amount",0) - c.get("amount",0)/(1+c.get("moms_rate",25)/100) for c in costs)
    excl_moms = total - ingaende_moms
    by_cat = {}
    for c in costs:
        cat = c.get("category","material")
        by_cat[cat] = by_cat.get(cat,0) + c.get("amount",0)

    # Last 6 months comparison
    monthly_data = []
    for i in range(5, -1, -1):
        d = DateClass(int(year), int(mon), 1)
        for _ in range(i):
            d = (d.replace(day=1) - timedelta(days=1)).replace(day=1)
        m_start = d.strftime("%Y-%m-01")
        m_last = calendar.monthrange(d.year, d.month)[1]
        m_end = f"{d.strftime('%Y-%m')}-{m_last:02d}"
        m_costs = await db.costs.find({"date": {"$gte": m_start, "$lte": m_end}}).to_list(1000)
        m_total = sum(c.get("amount",0) for c in m_costs)
        monthly_data.append({
            "month": MONTHS_SV[d.month][:3] + f" {d.year}",
            "total": m_total,
            "count": len(m_costs),
        })

    # Build PDF
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        topMargin=15*mm, bottomMargin=15*mm, leftMargin=20*mm, rightMargin=20*mm,
        title=f"Kostnadsrapport {month}",
        author=company)
    styles = getSampleStyleSheet()
    def ps(name, **kw): return ParagraphStyle(name, parent=styles["Normal"], **kw)
    page_w = A4[0]
    elements = []

    # Header
    hdr = Table([[
        Paragraph(f"<b>{company}</b>", ps("h", fontSize=14, fontName="Helvetica-Bold", textColor=colors.white)),
        Paragraph(f"<b>KOSTNADSRAPPORT</b><br/>{month_name} {year}", ps("s", fontSize=10, fontName="Helvetica-Bold", textColor=colors.white, alignment=2))
    ]], colWidths=[100*mm, None])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#141414")),
        ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),
        ("TOPPADDING",(0,0),(-1,-1),10),("BOTTOMPADDING",(0,0),(-1,-1),10),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    elements.append(hdr)
    elements.append(Spacer(1,4*mm))

    # Summary boxes
    sum_data = [[
        Paragraph(f"<b>Totalt inkl. moms</b><br/>{total:.2f} kr", ps("sb", fontSize=9, textColor=colors.HexColor("#141414"))),
        Paragraph(f"<b>Ingående moms</b><br/>{ingaende_moms:.2f} kr", ps("sb", fontSize=9, textColor=colors.HexColor("#15803d"))),
        Paragraph(f"<b>Nettokostnad (exkl. moms)</b><br/>{excl_moms:.2f} kr", ps("sb", fontSize=9, textColor=colors.HexColor("#1e40af"))),
        Paragraph(f"<b>Antal poster</b><br/>{len(costs)} st", ps("sb", fontSize=9, textColor=colors.HexColor("#64748b"))),
    ]]
    sum_tbl = Table(sum_data, colWidths=[(page_w-40*mm)/4]*4)
    sum_tbl.setStyle(TableStyle([
        ("BOX",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),
        ("INNERGRID",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),
        ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),
        ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8),
        ("BACKGROUND",(0,0),(0,0),colors.HexColor("#F8FAFC")),
    ]))
    elements.append(sum_tbl)
    elements.append(Spacer(1,5*mm))

    # Category breakdown
    if by_cat:
        elements.append(Paragraph("KOSTNADER PER KATEGORI", ps("sh", fontSize=9, fontName="Helvetica-Bold", spaceBefore=2, spaceAfter=2)))
        cat_names = {"material":"Städmaterial","equipment":"Utrustning","transport":"Transport","parking":"Parkering","other":"Övrigt"}
        cat_data = [["Kategori", "Belopp inkl. moms", "Andel %"]]
        for cat, amt in sorted(by_cat.items(), key=lambda x: -x[1]):
            pct = (amt/total*100) if total > 0 else 0
            cat_data.append([cat_names.get(cat,cat), f"{amt:.2f} kr", f"{pct:.1f}%"])
        cat_data.append(["TOTALT", f"{total:.2f} kr", "100%"])
        cat_tbl = Table(cat_data, colWidths=[80*mm, 60*mm, 30*mm])
        cat_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#141414")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),9),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
            ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
            ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#F1F5F9")),
            ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
            ("ALIGN",(1,0),(-1,-1),"RIGHT"),
        ]))
        elements.append(cat_tbl)
        elements.append(Spacer(1,5*mm))

    # Detailed list
    if costs:
        elements.append(Paragraph("DETALJERAD KOSTNADSLISTA", ps("sh", fontSize=9, fontName="Helvetica-Bold", spaceBefore=2, spaceAfter=2)))
        det_data = [["Datum", "Namn", "Kategori", "Moms %", "Belopp"]]
        cat_names = {"material":"Material","equipment":"Utrustning","transport":"Transport","parking":"Parkering","other":"Övrigt"}
        for c in sorted(costs, key=lambda x: x.get("date","")):
            det_data.append([
                c.get("date",""),
                c.get("name","")[:30],
                cat_names.get(c.get("category",""),""),
                f"{c.get('moms_rate',25)}%",
                f"{c.get('amount',0):.2f} kr",
            ])
        det_tbl = Table(det_data, colWidths=[25*mm, 70*mm, 30*mm, 20*mm, 30*mm])
        det_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#475569")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),8),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
            ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("ALIGN",(4,0),(-1,-1),"RIGHT"),
        ]))
        elements.append(det_tbl)
        elements.append(Spacer(1,5*mm))

    # Monthly comparison
    elements.append(Paragraph("JÄMFÖRELSE SENASTE 6 MÅNADER", ps("sh", fontSize=9, fontName="Helvetica-Bold", spaceBefore=2, spaceAfter=2)))
    max_val = max(m["total"] for m in monthly_data) or 1
    comp_data = [["Månad", "Kostnader", "Poster", "Graf"]]
    for m in monthly_data:
        bar_len = int((m["total"] / max_val) * 30) if max_val > 0 else 0
        bar = "█" * bar_len
        is_current = m["month"].startswith(month_name[:3])
        comp_data.append([
            m["month"],
            f"{m['total']:.2f} kr",
            f"{m['count']} st",
            Paragraph(f'<font color="{"#141414" if is_current else "#94a3b8"}">{bar}</font>', ps("bar", fontSize=7)),
        ])
    comp_tbl = Table(comp_data, colWidths=[35*mm, 40*mm, 20*mm, None])
    comp_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#475569")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),8),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
        ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("ALIGN",(1,0),(2,-1),"RIGHT"),
    ]))
    elements.append(comp_tbl)

    # Lönsamhet section
    elements.append(Spacer(1,5*mm))
    elements.append(Paragraph("LÖNSAMHETSANALYS", ps("sh", fontSize=9, fontName="Helvetica-Bold", spaceBefore=2, spaceAfter=2)))

    # Get revenue for this month
    invoices = await db.invoices.find({
        "created_at": {"$gte": start + "T00:00:00", "$lte": end + "T23:59:59"},
        "status": "paid"
    }).to_list(1000)
    revenue = sum(i.get("customer_pays",0) for i in invoices)
    bruttovinst = round(revenue - total, 2)
    bruttomarginal = round((bruttovinst / revenue * 100) if revenue > 0 else 0, 1)

    lon_data = [
        ["Post", "Belopp"],
        ["Intäkter (betalda fakturor)", f"{revenue:.2f} kr"],
        ["Materialkostnader (inkl. moms)", f"-{total:.2f} kr"],
        ["Ingående moms (avdragsgill)", f"+{round(ingaende_moms,2):.2f} kr"],
        ["Nettokostnader (exkl. moms)", f"-{round(excl_moms,2):.2f} kr"],
        ["BRUTTOVINST", f"{bruttovinst:.2f} kr"],
        ["Bruttomarginal", f"{bruttomarginal}%"],
    ]
    lon_tbl = Table(lon_data, colWidths=[120*mm, None])
    lon_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#141414")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),9),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
        ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("ALIGN",(1,0),(-1,-1),"RIGHT"),
        ("BACKGROUND",(0,-2),(-1,-1),colors.HexColor("#F1F5F9")),
        ("FONTNAME",(0,-2),(-1,-1),"Helvetica-Bold"),
        ("TEXTCOLOR",(1,-2),(1,-2),colors.HexColor("#15803d") if bruttovinst >= 0 else colors.HexColor("#dc2626")),
        ("LINEABOVE",(0,-2),(-1,-2),1,colors.HexColor("#141414")),
    ]))
    elements.append(lon_tbl)

    elements.append(Spacer(1,4*mm))
    elements.append(Paragraph(
        f"{company}  ·  Kostnadsrapport {month_name} {year}  ·  Skapad {datetime.now(timezone.utc).strftime('%Y-%m-%d')}",
        ps("ft", fontSize=7, textColor=colors.HexColor("#94a3b8"))
    ))

    doc.build(elements)
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="kostnader_{month}.pdf"'}
    )


# ── Invoices Excel Export ─────────────────────────────────────────────────────
@api_router.get("/invoices/export-xlsx")
async def export_invoices_xlsx(start: str = None, end: str = None, current=Depends(get_current_user)):
    """Export invoices to Excel"""
    query = {}
    if start and end:
        query["created_at"] = {"$gte": start, "$lte": end + "T23:59:59"}
    elif start:
        query["created_at"] = {"$gte": start}
    elif end:
        query["created_at"] = {"$lte": end + "T23:59:59"}

    invoices = await db.invoices.find(query).sort("created_at", -1).to_list(5000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Fakturor"

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Headers
    headers = [
        "Faktura #", "Datum", "Förfallodatum", "Kund", "E-post",
        "Typ", "Tjänst", "Exkl. moms (kr)", "Moms 25% (kr)",
        "Totalt inkl. moms (kr)", "RUT-avdrag (kr)", "Påminnelseavgift (kr)",
        "Att betala (kr)", "Status", "Påminnelser",
    ]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color="141414", end_color="141414", fill_type="solid")
    for col, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    status_map = {"paid": "Betald", "draft": "Utkast", "sent": "Skickad", "overdue": "Förfallen", "cancelled": "Avbruten"}

    total_excl = total_moms = total_rut = total_pam = total_pays = 0
    paid_count = unpaid_count = 0

    for inv in invoices:
        items = inv.get("items", [])
        pam_fee = sum(item.get("quantity",1)*item.get("unit_price",0) for item in items if "Påminnelseavgift" in item.get("service",""))
        services = ", ".join(set(i.get("service","") for i in items if "Påminnelseavgift" not in i.get("service","")))
        status = inv.get("status","")
        excl = inv.get("subtotal", 0)
        moms = inv.get("vat_amount", 0)
        rut = inv.get("rut_deduction", 0)
        pays = inv.get("customer_pays", 0)

        total_excl += excl
        total_moms += moms
        total_rut += rut
        total_pam += pam_fee
        total_pays += pays
        if status == "paid": paid_count += 1
        else: unpaid_count += 1

        row = [
            inv.get("invoice_number",""),
            inv.get("created_at","")[:10],
            inv.get("due_date",""),
            inv.get("customer_name",""),
            inv.get("customer_email",""),
            "Företag" if inv.get("customer_type") == "company" else "Privat",
            services[:50],
            round(excl, 2),
            round(moms, 2),
            round(excl + moms, 2),
            round(rut, 2),
            round(pam_fee, 2),
            round(pays, 2),
            status_map.get(status, status),
            inv.get("reminder_count", 0),
        ]
        ws.append(row)

        # Color status
        status_colors = {"paid": "DCFCE7", "overdue": "FEE2E2", "sent": "EFF6FF", "draft": "F8FAFC"}
        fill_color = status_colors.get(status, "FFFFFF")
        for cell in ws[ws.max_row]:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="left")

    # Totals row
    ws.append(["TOTALT", "", "", "", "", "", "",
        round(total_excl,2), round(total_moms,2), round(total_excl+total_moms,2),
        round(total_rut,2), round(total_pam,2), round(total_pays,2), "", ""])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    # Summary sheet
    ws2 = wb.create_sheet("Sammanfattning")
    ws2.append(["Post", "Värde"])
    ws2.append(["Antal fakturor", len(invoices)])
    ws2.append(["Betalda", paid_count])
    ws2.append(["Obetalda", unpaid_count])
    ws2.append(["Total exkl. moms", round(total_excl,2)])
    ws2.append(["Total moms", round(total_moms,2)])
    ws2.append(["Total RUT-avdrag", round(total_rut,2)])
    ws2.append(["Total påminnelseavgifter", round(total_pam,2)])
    ws2.append(["Totalt att betala", round(total_pays,2)])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="141414", end_color="141414", fill_type="solid")

    # Column widths
    col_widths = [12, 12, 14, 25, 25, 10, 30, 16, 14, 18, 14, 18, 14, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 18

    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="fakturor.xlsx"'}
    )


# ── Invoices List PDF Export ──────────────────────────────────────────────────
@api_router.get("/invoices/export-pdf")
async def export_invoices_pdf(start: str = None, end: str = None, current=Depends(get_current_user)):
    query = {}
    if start and end:
        query["created_at"] = {"$gte": start, "$lte": end + "T23:59:59"}
    invoices = await db.invoices.find(query).sort("created_at", -1).to_list(5000)
    inv_settings = await get_invoice_settings_obj()
    company = inv_settings.company_name or "PureNorth Städ"

    total_excl = sum(i.get("subtotal",0) for i in invoices)
    total_moms = sum(i.get("vat_amount",0) for i in invoices)
    total_rut = sum(i.get("rut_deduction",0) for i in invoices)
    total_pam = sum(sum(item.get("quantity",1)*item.get("unit_price",0) for item in i.get("items",[]) if "Påminnelseavgift" in item.get("service","")) for i in invoices)
    total_pays = sum(i.get("customer_pays",0) for i in invoices)
    paid = sum(1 for i in invoices if i.get("status") == "paid")

    from reportlab.lib.pagesizes import landscape
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
        topMargin=12*mm, bottomMargin=12*mm, leftMargin=15*mm, rightMargin=15*mm,
        title="Fakturalista", author=company)
    styles = getSampleStyleSheet()
    def ps(name, **kw): return ParagraphStyle(name, parent=styles["Normal"], **kw)
    page_w = landscape(A4)[0]
    elements = []

    # Header
    hdr = Table([[
        Paragraph(f"<b>{company}</b>", ps("h", fontSize=14, fontName="Helvetica-Bold", textColor=colors.white)),
        Paragraph(f"<b>FAKTURALISTA</b><br/>Skapad: {datetime.now(timezone.utc).strftime('%Y-%m-%d')}", ps("s", fontSize=10, fontName="Helvetica-Bold", textColor=colors.white, alignment=2))
    ]], colWidths=[120*mm, None])
    hdr.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#141414")),("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),("TOPPADDING",(0,0),(-1,-1),10),("BOTTOMPADDING",(0,0),(-1,-1),10),("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
    elements.append(hdr)
    elements.append(Spacer(1,3*mm))

    # Summary
    sum_data = [[
        Paragraph(f"<b>Antal</b><br/>{len(invoices)} st", ps("sb", fontSize=9)),
        Paragraph(f"<b>Betalda</b><br/>{paid} st", ps("sb", fontSize=9, textColor=colors.HexColor("#15803d"))),
        Paragraph(f"<b>Obetalda</b><br/>{len(invoices)-paid} st", ps("sb", fontSize=9, textColor=colors.HexColor("#dc2626"))),
        Paragraph(f"<b>Exkl. moms</b><br/>{total_excl:.2f} kr", ps("sb", fontSize=9, textColor=colors.HexColor("#1e40af"))),
        Paragraph(f"<b>RUT-avdrag</b><br/>{total_rut:.2f} kr", ps("sb", fontSize=9, textColor=colors.HexColor("#7c3aed"))),
        Paragraph(f"<b>Att betala</b><br/>{total_pays:.2f} kr", ps("sb", fontSize=9)),
    ]]
    sum_tbl = Table(sum_data, colWidths=[(page_w-30*mm)/6]*6)
    sum_tbl.setStyle(TableStyle([("BOX",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),("INNERGRID",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6)]))
    elements.append(sum_tbl)
    elements.append(Spacer(1,4*mm))

    status_map = {"paid":"Betald","draft":"Utkast","sent":"Skickad","overdue":"Förfallen","cancelled":"Avbruten"}
    status_colors = {"paid":colors.HexColor("#15803d"),"overdue":colors.HexColor("#dc2626"),"sent":colors.HexColor("#1e40af"),"draft":colors.HexColor("#64748b")}

    data = [["#","Datum","Förfaller","Kund","Tjänst","Exkl. moms","Moms","RUT","Påm.avg","Att betala","Status","Påm."]]
    for inv in invoices:
        items = inv.get("items",[])
        pam = sum(item.get("quantity",1)*item.get("unit_price",0) for item in items if "Påminnelseavgift" in item.get("service",""))
        svc_set = set()
        for it in items:
            if "Påminnelseavgift" not in it.get("service",""):
                svc = it.get("service","") or it.get("description","")
                svc = svc.split("(")[0].strip()  # Remove "(100 kvm → 5 tim)" part
                if svc: svc_set.add(svc)
        services = ", ".join(svc_set)[:30] or "Städtjänst"
        data.append([str(inv.get("invoice_number","")),inv.get("created_at","")[:10],inv.get("due_date",""),inv.get("customer_name","")[:20],services,f'{inv.get("subtotal",0):.2f}',f'{inv.get("vat_amount",0):.2f}',f'{inv.get("rut_deduction",0):.2f}',f'{pam:.2f}' if pam>0 else "-",f'{inv.get("customer_pays",0):.2f}',status_map.get(inv.get("status",""),""),str(inv.get("reminder_count",0))])
    data.append(["TOTALT","","","","",f"{total_excl:.2f}",f"{total_moms:.2f}",f"{total_rut:.2f}",f"{total_pam:.2f}",f"{total_pays:.2f}","",""])

    tbl = Table(data, colWidths=[16*mm,22*mm,22*mm,38*mm,38*mm,24*mm,18*mm,18*mm,18*mm,24*mm,20*mm,12*mm])
    ts = [("BACKGROUND",(0,0),(-1,0),colors.HexColor("#141414")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),7.5),("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3),("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),("ALIGN",(5,0),(-1,-1),"RIGHT"),("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#F1F5F9")),("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold")]
    for row_idx, inv in enumerate(invoices, 1):
        col = status_colors.get(inv.get("status",""))
        if col: ts.append(("TEXTCOLOR",(10,row_idx),(10,row_idx),col))
    tbl.setStyle(TableStyle(ts))
    elements.append(tbl)
    elements.append(Spacer(1,3*mm))
    elements.append(Paragraph(f"{company}  ·  Fakturalista  ·  {datetime.now(timezone.utc).strftime('%Y-%m-%d')}", ps("ft", fontSize=7, textColor=colors.HexColor("#94a3b8"))))
    doc.build(elements)
    return Response(content=buf.getvalue(), media_type="application/pdf", headers={"Content-Disposition": 'inline; filename="fakturor.pdf"'})



# ── Costs Excel Export ────────────────────────────────────────────────────────
@api_router.get("/costs/export-xlsx")
async def export_costs_xlsx(start: str = None, end: str = None, current=Depends(get_current_user)):
    """Export costs to Excel with monthly comparison"""
    import calendar
    query = {}
    if start and end:
        query["date"] = {"$gte": start, "$lte": end}
    elif start:
        query["date"] = {"$gte": start}
    elif end:
        query["date"] = {"$lte": end}

    costs = await db.costs.find(query).sort("date", -1).to_list(5000)

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Kostnader ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Kostnader"

    headers = ["Datum", "Namn", "Kategori", "Moms %", "Belopp inkl. moms (kr)", "Ingående moms (kr)", "Nettokostnad exkl. moms (kr)"]
    ws.append(headers)

    header_fill = PatternFill(start_color="141414", end_color="141414", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    cat_names = {"material":"Städmaterial","equipment":"Utrustning","transport":"Transport","parking":"Parkering","other":"Övrigt"}

    total_incl = total_moms = total_excl = 0
    for c in costs:
        amt = c.get("amount",0)
        mrate = c.get("moms_rate",25)
        moms = round(amt - amt/(1+mrate/100), 2) if mrate > 0 else 0
        excl = round(amt - moms, 2)
        total_incl += amt; total_moms += moms; total_excl += excl
        ws.append([
            c.get("date",""),
            c.get("name",""),
            cat_names.get(c.get("category",""),""),
            f"{mrate}%",
            round(amt,2),
            moms,
            excl,
        ])
        for cell in ws[ws.max_row]:
            cell.font = Font(size=9)

    # Totals
    ws.append(["TOTALT","","","", round(total_incl,2), round(total_moms,2), round(total_excl,2)])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    col_widths = [14, 30, 16, 10, 24, 20, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Per kategori ──────────────────────────────────────
    ws2 = wb.create_sheet("Per kategori")
    ws2.append(["Kategori", "Antal poster", "Belopp inkl. moms (kr)", "Ingående moms (kr)", "Nettokostnad (kr)"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = header_fill

    by_cat = {}
    for c in costs:
        cat = cat_names.get(c.get("category","material"),"Övrigt")
        if cat not in by_cat:
            by_cat[cat] = {"count":0,"total":0,"moms":0,"excl":0}
        amt = c.get("amount",0)
        mrate = c.get("moms_rate",25)
        moms = round(amt - amt/(1+mrate/100), 2) if mrate > 0 else 0
        by_cat[cat]["count"] += 1
        by_cat[cat]["total"] += amt
        by_cat[cat]["moms"] += moms
        by_cat[cat]["excl"] += round(amt-moms, 2)

    for cat, vals in sorted(by_cat.items()):
        ws2.append([cat, vals["count"], round(vals["total"],2), round(vals["moms"],2), round(vals["excl"],2)])
        for cell in ws2[ws2.max_row]: cell.font = Font(size=9)

    for i, w in enumerate([20,14,24,20,20], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Månadsöversikt ────────────────────────────────────
    ws3 = wb.create_sheet("Månadsöversikt")
    ws3.append(["Månad", "Antal poster", "Totalt inkl. moms (kr)", "Ingående moms (kr)", "Nettokostnad (kr)"])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = header_fill

    by_month = {}
    for c in costs:
        month = c.get("date","")[:7]
        if month not in by_month:
            by_month[month] = {"count":0,"total":0,"moms":0,"excl":0}
        amt = c.get("amount",0)
        mrate = c.get("moms_rate",25)
        moms = round(amt - amt/(1+mrate/100), 2) if mrate > 0 else 0
        by_month[month]["count"] += 1
        by_month[month]["total"] += amt
        by_month[month]["moms"] += moms
        by_month[month]["excl"] += round(amt-moms, 2)

    for month in sorted(by_month.keys(), reverse=True):
        vals = by_month[month]
        ws3.append([month, vals["count"], round(vals["total"],2), round(vals["moms"],2), round(vals["excl"],2)])
        for cell in ws3[ws3.max_row]: cell.font = Font(size=9)

    for i, w in enumerate([14,14,24,20,20], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="kostnader.xlsx"'}
    )


# ── Bookings PDF Export ───────────────────────────────────────────────────────
@api_router.get("/bookings/export-pdf")
async def export_bookings_pdf(start: str = None, end: str = None, current=Depends(get_current_user)):
    """Export bookings list to PDF"""
    query = {}
    if start and end:
        query["created_at"] = {"$gte": start, "$lte": end + "T23:59:59"}

    bookings = await db.bookings.find(query).sort("created_at", -1).to_list(5000)
    inv_settings = await get_invoice_settings_obj()
    company = inv_settings.company_name or "PureNorth Städ"

    status_map = {"new":"Ny","contacted":"Kontaktad","quoted":"Offert","done":"Klar","cancelled":"Avbruten"}
    status_colors_map = {"new":colors.HexColor("#1e40af"),"contacted":colors.HexColor("#b45309"),"quoted":colors.HexColor("#7c3aed"),"done":colors.HexColor("#15803d"),"cancelled":colors.HexColor("#dc2626")}

    # Summary
    total = len(bookings)
    by_status = {}
    for b in bookings:
        s = b.get("status","new")
        by_status[s] = by_status.get(s,0) + 1

    from reportlab.lib.pagesizes import landscape
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
        topMargin=12*mm, bottomMargin=12*mm, leftMargin=15*mm, rightMargin=15*mm,
        title="Bokningslista", author=company)
    styles = getSampleStyleSheet()
    def ps(name, **kw): return ParagraphStyle(name, parent=styles["Normal"], **kw)
    page_w = landscape(A4)[0]
    elements = []

    # Header
    hdr = Table([[
        Paragraph(f"<b>{company}</b>", ps("h", fontSize=14, fontName="Helvetica-Bold", textColor=colors.white)),
        Paragraph(f"<b>BOKNINGSLISTA</b><br/>Skapad: {datetime.now(timezone.utc).strftime('%Y-%m-%d')}", ps("s", fontSize=10, fontName="Helvetica-Bold", textColor=colors.white, alignment=2))
    ]], colWidths=[120*mm, None])
    hdr.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#141414")),("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),("TOPPADDING",(0,0),(-1,-1),10),("BOTTOMPADDING",(0,0),(-1,-1),10),("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
    elements.append(hdr)
    elements.append(Spacer(1,3*mm))

    # Summary
    sum_data = [[
        Paragraph(f"<b>Totalt</b><br/>{total} st", ps("sb", fontSize=9)),
        Paragraph(f"<b>Nya</b><br/>{by_status.get('new',0)} st", ps("sb", fontSize=9, textColor=colors.HexColor("#1e40af"))),
        Paragraph(f"<b>Kontaktade</b><br/>{by_status.get('contacted',0)} st", ps("sb", fontSize=9, textColor=colors.HexColor("#b45309"))),
        Paragraph(f"<b>Offert</b><br/>{by_status.get('quoted',0)} st", ps("sb", fontSize=9, textColor=colors.HexColor("#7c3aed"))),
        Paragraph(f"<b>Klara</b><br/>{by_status.get('done',0)} st", ps("sb", fontSize=9, textColor=colors.HexColor("#15803d"))),
        Paragraph(f"<b>Avbrutna</b><br/>{by_status.get('cancelled',0)} st", ps("sb", fontSize=9, textColor=colors.HexColor("#dc2626"))),
    ]]
    sum_tbl = Table(sum_data, colWidths=[(page_w-30*mm)/6]*6)
    sum_tbl.setStyle(TableStyle([("BOX",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),("INNERGRID",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6)]))
    elements.append(sum_tbl)
    elements.append(Spacer(1,4*mm))

    # Bookings table
    data = [["Datum", "Kund", "E-post", "Telefon", "Tjänst", "Adress", "Meddelande", "Status"]]
    for b in bookings:
        data.append([
            b.get("created_at","")[:10],
            b.get("name","")[:20],
            b.get("email","")[:25] or "-",
            b.get("phone","") or "-",
            (", ".join(b.get("services",[]) or [b.get("service","")]))[:25] or "-",
            b.get("address","")[:25] or "-",
            (b.get("other_description","") or b.get("message","") or "-")[:30],
            status_map.get(b.get("status","new"),""),
        ])

    tbl = Table(data, colWidths=[22*mm, 35*mm, 45*mm, 25*mm, 30*mm, 40*mm, 45*mm, 25*mm])
    ts = [
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#141414")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),7.5),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
        ("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#F8FAFC")]),
    ]
    for row_idx, b in enumerate(bookings, 1):
        col = status_colors_map.get(b.get("status","new"))
        if col: ts.append(("TEXTCOLOR",(7,row_idx),(7,row_idx),col))
    tbl.setStyle(TableStyle(ts))
    elements.append(tbl)
    elements.append(Spacer(1,3*mm))
    elements.append(Paragraph(f"{company}  ·  Bokningslista  ·  {datetime.now(timezone.utc).strftime('%Y-%m-%d')}", ps("ft", fontSize=7, textColor=colors.HexColor("#94a3b8"))))

    doc.build(elements)
    return Response(content=buf.getvalue(), media_type="application/pdf", headers={"Content-Disposition": 'inline; filename="bokningar.pdf"'})


# ── Bookings Excel Export ─────────────────────────────────────────────────────
@api_router.get("/bookings/export-xlsx")
async def export_bookings_xlsx(start: str = None, end: str = None, current=Depends(get_current_user)):
    query = {}
    if start and end:
        query["created_at"] = {"$gte": start, "$lte": end + "T23:59:59"}
    bookings = await db.bookings.find(query).sort("created_at", -1).to_list(5000)

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Bokningar"

    headers = ["Datum", "Namn", "E-post", "Telefon", "Adress", "Tjänster", "Yta/Antal", "Önskat datum", "Anteckning", "Status"]
    ws.append(headers)

    header_fill = PatternFill(start_color="141414", end_color="141414", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    status_map = {"new":"Ny","contacted":"Kontaktad","quoted":"Offert","done":"Klar","cancelled":"Avbruten"}
    status_colors = {"new":"DBEAFE","contacted":"FEF3C7","quoted":"EDE9FE","done":"DCFCE7","cancelled":"FEE2E2"}

    by_status = {}
    for b in bookings:
        s = b.get("status","new")
        by_status[s] = by_status.get(s,0) + 1
        services = ", ".join(b.get("services",[]) or [b.get("service","")])
        status = status_map.get(s, s)
        ws.append([
            b.get("created_at","")[:10],
            b.get("name",""),
            b.get("email",""),
            b.get("phone",""),
            b.get("address","") or "-",
            services,
            b.get("kvm","") or "-",
            b.get("preferred_date","") or "-",
            b.get("other_description","") or "-",
            status,
        ])
        fill_color = status_colors.get(b.get("status","new"), "FFFFFF")
        for cell in ws[ws.max_row]:
            cell.font = Font(size=9)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Summary sheet
    ws2 = wb.create_sheet("Sammanfattning")
    ws2.append(["Status", "Antal"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = header_fill
    for s, label in status_map.items():
        ws2.append([label, by_status.get(s,0)])
        for cell in ws2[ws2.max_row]: cell.font = Font(size=9)
    ws2.append(["TOTALT", len(bookings)])
    for cell in ws2[ws2.max_row]: cell.font = Font(bold=True, size=9)

    col_widths = [14, 22, 28, 16, 25, 30, 12, 16, 30, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 10

    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="bokningar.xlsx"'}
    )


# ── Schema PDF Export ─────────────────────────────────────────────────────────
@api_router.get("/shifts/export-pdf")
async def export_shifts_pdf(start: str, end: str, current=Depends(get_current_user)):
    """Export schema/shifts to PDF"""
    shifts = await db.shifts.find({"date": {"$gte": start, "$lte": end}}).sort("date", 1).to_list(5000)
    employees = await db.employees.find().to_list(100)
    emp_map = {str(e["_id"]): e for e in employees}
    # Get absences for this period
    absences = await db.absences.find({"start_date": {"$lte": end}, "end_date": {"$gte": start}}).to_list(1000)
    absence_map = {}
    for ab in absences:
        eid = ab.get("employee_id","")
        if eid not in absence_map: absence_map[eid] = []
        absence_map[eid].append({"type": ab.get("absence_type","sjuk"), "start": ab.get("start_date",""), "end": ab.get("end_date","")})
    inv_settings = await get_invoice_settings_obj()
    company = inv_settings.company_name or "PureNorth Städ"

    # Group by employee
    by_emp = {}
    for s in shifts:
        eid = s.get("employee_id","")
        if eid not in by_emp:
            emp = emp_map.get(eid, {})
            by_emp[eid] = {"name": emp.get("name","Okänd"), "shifts": [], "total_h": 0}
        # Calculate hours
        try:
            from datetime import datetime as DT
            st = DT.strptime(s["start_time"], "%H:%M")
            et = DT.strptime(s["end_time"], "%H:%M")
            h = round((et - st).seconds / 3600, 2)
        except:
            h = 0
        by_emp[eid]["shifts"].append({**s, "hours": h})
        by_emp[eid]["total_h"] += h

    from reportlab.lib.pagesizes import landscape
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
        topMargin=12*mm, bottomMargin=12*mm, leftMargin=15*mm, rightMargin=15*mm,
        title=f"Schema {start} - {end}", author=company)
    styles = getSampleStyleSheet()
    def ps(name, **kw): return ParagraphStyle(name, parent=styles["Normal"], **kw)
    page_w = landscape(A4)[0]
    elements = []

    # Header
    hdr = Table([[
        Paragraph(f"<b>{company}</b>", ps("h", fontSize=14, fontName="Helvetica-Bold", textColor=colors.white)),
        Paragraph(f"<b>SCHEMA</b><br/>{start} – {end}", ps("s", fontSize=10, fontName="Helvetica-Bold", textColor=colors.white, alignment=2))
    ]], colWidths=[120*mm, None])
    hdr.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#141414")),("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),("TOPPADDING",(0,0),(-1,-1),10),("BOTTOMPADDING",(0,0),(-1,-1),10),("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
    elements.append(hdr)
    elements.append(Spacer(1,3*mm))

    # Summary
    total_shifts = len(shifts)
    total_hours = sum(e["total_h"] for e in by_emp.values())
    sum_data = [[
        Paragraph(f"<b>Antal pass</b><br/>{total_shifts} st", ps("sb", fontSize=9)),
        Paragraph(f"<b>Antal anställda</b><br/>{len(by_emp)} st", ps("sb", fontSize=9, textColor=colors.HexColor("#1e40af"))),
        Paragraph(f"<b>Totala timmar</b><br/>{total_hours:.1f} tim", ps("sb", fontSize=9, textColor=colors.HexColor("#15803d"))),
    ]]
    sum_tbl = Table(sum_data, colWidths=[(page_w-30*mm)/3]*3)
    sum_tbl.setStyle(TableStyle([("BOX",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),("INNERGRID",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8)]))
    elements.append(sum_tbl)
    elements.append(Spacer(1,5*mm))

    # Shifts table
    data = [["Datum", "Anställd", "Start", "Slut", "Timmar", "OB", "Notering"]]
    for eid, emp_data in sorted(by_emp.items(), key=lambda x: x[1]["name"]):
        for s in emp_data["shifts"]:
            ob = "OB1" if s.get("is_ob1") else ("OB2" if s.get("is_ob2") else "-")
            # Check if employee is absent this day
            emp_absences = absence_map.get(eid, [])
            sick_note = ""
            for ab in emp_absences:
                if ab["start"] <= s.get("date","") <= ab["end"]:
                    atype_map = {"sjuk":"🤒 Sjuk","semester":"🏖 Semester","vab":"👶 VAB","permission":"📋 Permission"}
                    sick_note = atype_map.get(ab["type"], ab["type"])
                    break
            data.append([
                s.get("date",""),
                emp_data["name"],
                s.get("start_time",""),
                s.get("end_time",""),
                f'{s["hours"]:.1f} tim',
                ob,
                sick_note or s.get("note","") or "-",
            ])
        # Employee subtotal
        data.append(["", f'SUMMA {emp_data["name"]}', "", "", f'{emp_data["total_h"]:.1f} tim', "", ""])

    # Grand total
    data.append(["TOTALT", "", "", "", f'{total_hours:.1f} tim', "", ""])

    tbl = Table(data, colWidths=[25*mm, 55*mm, 20*mm, 20*mm, 22*mm, 16*mm, None])
    ts = [
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#141414")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),8.5),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
        ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#F8FAFC")]),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#F1F5F9")),
        ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
    ]
    # Style subtotal rows
    for i, row in enumerate(data[1:], 1):
        if row[1].startswith("SUMMA"):
            ts.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor("#EFF6FF")))
            ts.append(("FONTNAME",(0,i),(-1,i),"Helvetica-Bold"))
            ts.append(("TEXTCOLOR",(0,i),(-1,i),colors.HexColor("#1e40af")))

    tbl.setStyle(TableStyle(ts))
    elements.append(tbl)
    elements.append(Spacer(1,3*mm))
    elements.append(Paragraph(f"{company}  ·  Schema {start} – {end}  ·  {datetime.now(timezone.utc).strftime('%Y-%m-%d')}", ps("ft", fontSize=7, textColor=colors.HexColor("#94a3b8"))))

    doc.build(elements)
    return Response(content=buf.getvalue(), media_type="application/pdf", headers={"Content-Disposition": f'inline; filename="schema_{start}_{end}.pdf"'})


# ── Schema Excel Export ───────────────────────────────────────────────────────
@api_router.get("/shifts/export-xlsx")
async def export_shifts_xlsx(start: str, end: str, current=Depends(get_current_user)):
    """Export schema/shifts to Excel"""
    shifts = await db.shifts.find({"date": {"$gte": start, "$lte": end}}).sort("date", 1).to_list(5000)
    employees = await db.employees.find().to_list(100)
    emp_map = {str(e["_id"]): e for e in employees}

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Alla pass ────────────────────────────────────────
    ws = wb.active
    ws.title = "Alla pass"

    headers = ["Datum", "Veckodag", "Anställd", "Start", "Slut", "Timmar", "OB1", "OB2", "Notering"]
    ws.append(headers)

    header_fill = PatternFill(start_color="141414", end_color="141414", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    DAYS_SV = ["Måndag","Tisdag","Onsdag","Torsdag","Fredag","Lördag","Söndag"]
    total_hours = 0

    # Load absences
    absences_list = await db.absences.find({"start_date": {"$lte": end}, "end_date": {"$gte": start}}).to_list(1000)
    absence_map_xl = {}
    for ab in absences_list:
        eid = ab.get("employee_id","")
        if eid not in absence_map_xl: absence_map_xl[eid] = []
        absence_map_xl[eid].append({"type": ab.get("absence_type","sjuk"), "start": ab.get("start_date",""), "end": ab.get("end_date","")})

    for s in shifts:
        eid = s.get("employee_id","")
        emp = emp_map.get(eid, {})
        try:
            from datetime import datetime as DT, date as DateObj
            d = DateObj.fromisoformat(s["date"])
            weekday = DAYS_SV[d.weekday()]
            st = DT.strptime(s["start_time"], "%H:%M")
            et = DT.strptime(s["end_time"], "%H:%M")
            h = round((et - st).seconds / 3600, 2)
        except:
            weekday = ""
            h = 0
        total_hours += h
        # Check absence
        sick_note = ""
        for ab in absence_map_xl.get(eid, []):
            if ab["start"] <= s.get("date","") <= ab["end"]:
                atype_map = {"sjuk":"Sjuk","semester":"Semester","vab":"VAB","permission":"Permission"}
                sick_note = atype_map.get(ab["type"], ab["type"])
                break
        ws.append([
            s.get("date",""),
            weekday,
            emp.get("name","Okänd"),
            s.get("start_time",""),
            s.get("end_time",""),
            h,
            "✓" if s.get("is_ob1") else "",
            "✓" if s.get("is_ob2") else "",
            sick_note or s.get("note","") or "",
        ])
        # Highlight sick rows
        if sick_note:
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        for cell in ws[ws.max_row]:
            cell.font = Font(size=9)

    # Totals
    ws.append(["TOTALT","","","","", round(total_hours,2),"","",""])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    col_widths = [14, 12, 25, 10, 10, 10, 8, 8, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Per anställd ─────────────────────────────────────
    ws2 = wb.create_sheet("Per anställd")
    ws2.append(["Anställd", "Antal pass", "Totala timmar", "OB1 timmar", "OB2 timmar"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = header_fill

    by_emp = {}
    for s in shifts:
        eid = s.get("employee_id","")
        emp = emp_map.get(eid,{})
        name = emp.get("name","Okänd")
        if name not in by_emp:
            by_emp[name] = {"count":0,"hours":0,"ob1":0,"ob2":0}
        try:
            st = DT.strptime(s["start_time"],"%H:%M")
            et = DT.strptime(s["end_time"],"%H:%M")
            h = round((et-st).seconds/3600,2)
        except: h = 0
        by_emp[name]["count"] += 1
        by_emp[name]["hours"] += h
        if s.get("is_ob1"): by_emp[name]["ob1"] += h
        if s.get("is_ob2"): by_emp[name]["ob2"] += h

    for name, vals in sorted(by_emp.items()):
        ws2.append([name, vals["count"], round(vals["hours"],2), round(vals["ob1"],2), round(vals["ob2"],2)])
        for cell in ws2[ws2.max_row]: cell.font = Font(size=9)

    ws2.append(["TOTALT", sum(v["count"] for v in by_emp.values()), round(total_hours,2),"",""])
    for cell in ws2[ws2.max_row]:
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    for i, w in enumerate([25,14,16,14,14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="schema_{start}_{end}.xlsx"'}
    )


# ── Utlägg PDF Export ─────────────────────────────────────────────────────────
@api_router.get("/expenses/export-pdf")
async def export_expenses_pdf(start: str = None, end: str = None, current=Depends(get_current_user)):
    """Export expenses/utlägg to PDF"""
    query = {}
    if start and end:
        query["date"] = {"$gte": start, "$lte": end}
    expenses = await db.expenses.find(query).sort("date", -1).to_list(5000)
    employees = await db.employees.find().to_list(100)
    emp_map = {str(e["_id"]): e["name"] for e in employees}
    inv_settings = await get_invoice_settings_obj()
    company = inv_settings.company_name or "PureNorth Städ"

    total = sum(e.get("amount",0) for e in expenses)
    total_moms = sum(e.get("amount",0) - e.get("amount",0)/(1+e.get("moms_rate",0)/100) for e in expenses if e.get("moms_rate",0) > 0)
    by_status = {"approved":0,"pending":0,"rejected":0}
    for e in expenses:
        s = e.get("status","pending")
        by_status[s] = by_status.get(s,0) + e.get("amount",0)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        topMargin=15*mm, bottomMargin=15*mm, leftMargin=20*mm, rightMargin=20*mm,
        title="Utlägg", author=company)
    styles = getSampleStyleSheet()
    def ps(name, **kw): return ParagraphStyle(name, parent=styles["Normal"], **kw)
    page_w = A4[0]
    elements = []

    # Header
    hdr = Table([[
        Paragraph(f"<b>{company}</b>", ps("h", fontSize=14, fontName="Helvetica-Bold", textColor=colors.white)),
        Paragraph(f"<b>UTLÄGG</b><br/>Skapad: {datetime.now(timezone.utc).strftime('%Y-%m-%d')}", ps("s", fontSize=10, fontName="Helvetica-Bold", textColor=colors.white, alignment=2))
    ]], colWidths=[100*mm, None])
    hdr.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#141414")),("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),("TOPPADDING",(0,0),(-1,-1),10),("BOTTOMPADDING",(0,0),(-1,-1),10),("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
    elements.append(hdr)
    elements.append(Spacer(1,3*mm))

    # Summary
    sum_data = [[
        Paragraph(f"<b>Totalt belopp</b><br/>{total:.2f} kr", ps("sb", fontSize=9)),
        Paragraph(f"<b>Godkänt</b><br/>{by_status['approved']:.2f} kr", ps("sb", fontSize=9, textColor=colors.HexColor("#15803d"))),
        Paragraph(f"<b>Väntar</b><br/>{by_status['pending']:.2f} kr", ps("sb", fontSize=9, textColor=colors.HexColor("#b45309"))),
        Paragraph(f"<b>Ingående moms</b><br/>{total_moms:.2f} kr", ps("sb", fontSize=9, textColor=colors.HexColor("#1e40af"))),
    ]]
    sum_tbl = Table(sum_data, colWidths=[(page_w-40*mm)/4]*4)
    sum_tbl.setStyle(TableStyle([("BOX",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),("INNERGRID",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8)]))
    elements.append(sum_tbl)
    elements.append(Spacer(1,5*mm))

    # Expenses table
    status_map = {"approved":"Godkänd","pending":"Väntar","rejected":"Avvisad","paid":"Utbetald"}
    status_colors_map = {"approved":colors.HexColor("#15803d"),"pending":colors.HexColor("#b45309"),"rejected":colors.HexColor("#dc2626")}

    data = [["Datum","Anställd","Kategori","Beskrivning","Moms %","Belopp","Status"]]
    for e in expenses:
        eid = e.get("employee_id","")
        emp_name = emp_map.get(eid) or e.get("employee_name","") or "Okänd"
        data.append([
            e.get("date",""),
            emp_name[:20],
            e.get("category","") or "-",
            (e.get("description","") or "-")[:30],
            f'{e.get("moms_rate",0)}%',
            f'{e.get("amount",0):.2f} kr',
            status_map.get(e.get("status","pending"),""),
        ])

    data.append(["TOTALT","","","","",f'{total:.2f} kr',""])

    tbl = Table(data, colWidths=[22*mm, 38*mm, 25*mm, 45*mm, 16*mm, 25*mm, 22*mm])
    ts = [
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#141414")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),8.5),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
        ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("ROWBACKGROUNDS",(0,1),(-1,-2),[colors.white, colors.HexColor("#F8FAFC")]),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#F1F5F9")),
        ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
        ("ALIGN",(5,0),(5,-1),"RIGHT"),
    ]
    for row_idx, e in enumerate(expenses, 1):
        col = status_colors_map.get(e.get("status","pending"))
        if col: ts.append(("TEXTCOLOR",(6,row_idx),(6,row_idx),col))
    tbl.setStyle(TableStyle(ts))
    elements.append(tbl)

    elements.append(Spacer(1,4*mm))
    elements.append(Paragraph(f"{company}  ·  Utlägg  ·  {datetime.now(timezone.utc).strftime('%Y-%m-%d')}", ps("ft", fontSize=7, textColor=colors.HexColor("#94a3b8"))))

    doc.build(elements)
    return Response(content=buf.getvalue(), media_type="application/pdf", headers={"Content-Disposition": 'inline; filename="utlagg.pdf"'})


# ── Utlägg Excel Export ───────────────────────────────────────────────────────
@api_router.get("/expenses/export-xlsx")
async def export_expenses_xlsx(start: str = None, end: str = None, current=Depends(get_current_user)):
    """Export expenses/utlägg to Excel"""
    query = {}
    if start and end:
        query["date"] = {"$gte": start, "$lte": end}
    expenses = await db.expenses.find(query).sort("date", -1).to_list(5000)
    employees = await db.employees.find().to_list(100)
    emp_map = {str(e["_id"]): e["name"] for e in employees}

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Utlägg"

    headers = ["Datum", "Anställd", "Kategori", "Beskrivning", "Moms %", "Ingående moms (kr)", "Belopp (kr)", "Status"]
    ws.append(headers)

    header_fill = PatternFill(start_color="141414", end_color="141414", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    status_map = {"approved":"Godkänd","pending":"Väntar","rejected":"Avvisad","paid":"Utbetald"}
    status_colors = {"approved":"DCFCE7","pending":"FEF3C7","rejected":"FEE2E2","paid":"DBEAFE"}

    total = total_moms = 0
    for e in expenses:
        eid = e.get("employee_id","")
        emp_name = emp_map.get(eid) or e.get("employee_name","") or "Okänd"
        amt = e.get("amount",0)
        mrate = e.get("moms_rate",0)
        moms = round(amt - amt/(1+mrate/100), 2) if mrate > 0 else 0
        total += amt
        total_moms += moms
        status = e.get("status","pending")

        ws.append([
            e.get("date",""),
            emp_name,
            e.get("category","") or "-",
            e.get("description","") or "-",
            f'{mrate}%',
            moms,
            amt,
            status_map.get(status, status),
        ])
        fill_color = status_colors.get(status, "FFFFFF")
        for cell in ws[ws.max_row]:
            cell.font = Font(size=9)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Totals
    ws.append(["TOTALT","","","","", round(total_moms,2), round(total,2),""])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    # Summary sheet
    ws2 = wb.create_sheet("Per kategori")
    ws2.append(["Kategori", "Antal", "Belopp (kr)", "Ingående moms (kr)"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = header_fill

    by_cat = {}
    for e in expenses:
        cat = e.get("category","Övrigt")
        if cat not in by_cat: by_cat[cat] = {"count":0,"total":0,"moms":0}
        amt = e.get("amount",0)
        mrate = e.get("moms_rate",0)
        moms = round(amt - amt/(1+mrate/100), 2) if mrate > 0 else 0
        by_cat[cat]["count"] += 1
        by_cat[cat]["total"] += amt
        by_cat[cat]["moms"] += moms

    for cat, vals in sorted(by_cat.items()):
        ws2.append([cat, vals["count"], round(vals["total"],2), round(vals["moms"],2)])
        for cell in ws2[ws2.max_row]: cell.font = Font(size=9)

    col_widths = [14, 22, 16, 30, 10, 20, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i, w in enumerate([20,10,16,18], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="utlagg.xlsx"'}
    )


# ── Website Settings ──────────────────────────────────────────────────────────
class WebsiteSettings(BaseModel):
    # Hero
    hero_title: Optional[str] = "Renhet med norrländsk precision i Umeå."
    hero_subtitle: Optional[str] = "Vi definierar premiumstädning genom certifierad expertis och hållbara metoder."
    hero_badge: Optional[str] = "Svanenmärkt & miljöcertifierat"
    hero_image: Optional[str] = ""
    cta_text: Optional[str] = "Boka tid online"
    badge1: Optional[str] = "SRY-kvalifikation"
    badge2: Optional[str] = "Pur-Eco produkter"
    # Kontakt
    phone: Optional[str] = "070-624 04 03"
    email: Optional[str] = ""
    address: Optional[str] = ""
    opening_hours: Optional[str] = "Mån–Fre: 08:00–18:00"
    # Företag
    company_name: Optional[str] = "PureNorth Städ"
    logo_url: Optional[str] = ""
    about_text: Optional[str] = ""
    # Sociala medier
    facebook_url: Optional[str] = ""
    instagram_url: Optional[str] = ""
    # Kontaktsektion
    contact_title: Optional[str] = ""
    contact_subtitle: Optional[str] = ""
    contact_description: Optional[str] = ""
    contact_box_title: Optional[str] = ""
    contact_box_text: Optional[str] = ""
    contact_box_btn: Optional[str] = ""
    # Varför oss
    whyus_label: Optional[str] = ""
    whyus_title: Optional[str] = ""
    sry_title: Optional[str] = ""
    sry_text: Optional[str] = ""
    sry_tags: Optional[str] = ""
    eco_title: Optional[str] = ""
    eco_text: Optional[str] = ""
    rut_title: Optional[str] = ""
    rut_text: Optional[str] = ""
    # Tjänster
    services_label: Optional[str] = ""
    services_title: Optional[str] = ""
    services: Optional[list] = []
    # Navbar
    nav_company: Optional[str] = ""
    nav_boka_text: Optional[str] = ""
    nav_links: Optional[list] = []
    # SEO
    seo_title: Optional[str] = ""
    seo_description: Optional[str] = ""
    seo_keywords: Optional[str] = ""
    seo_og_image: Optional[str] = ""
    # Färger
    primary_color: Optional[str] = "#166534"
    secondary_color: Optional[str] = "#141414"
    # Omdömen
    testimonials_label: Optional[str] = ""
    testimonials_title: Optional[str] = ""
    featured_reviews: Optional[list] = []

@api_router.get("/settings/website")
async def get_website_settings():
    defaults = WebsiteSettings().dict()
    doc = await db.website_settings.find_one({"_id": "main"})
    if not doc:
        return defaults
    doc.pop("_id", None)
    return {**defaults, **doc}

@api_router.patch("/settings/website")
async def update_website_settings(payload: WebsiteSettings, current=Depends(get_current_user)):
    data = {k: v for k, v in payload.dict().items() if v is not None or isinstance(v, list)}
    await db.website_settings.update_one(
        {"_id": "main"},
        {"$set": data},
        upsert=True
    )
    return {"success": True}

@api_router.post("/settings/website/upload-image")
async def upload_website_image(file: UploadFile = File(...), current=Depends(get_current_user)):
    """Upload image and return base64 URL"""
    import base64
    content = await file.read()
    b64 = base64.b64encode(content).decode()
    media_type = file.content_type or "image/jpeg"
    data_url = f"data:{media_type};base64,{b64}"
    return {"url": data_url}


# ── Årsredovisning Underlag ───────────────────────────────────────────────────
@api_router.get("/reports/arsredovisning")
async def arsredovisning_pdf(year: str, current=Depends(get_current_user)):
    """Annual report for accountant/konsult"""
    import calendar as cal_mod

    start = f"{year}-01-01"
    end = f"{year}-12-31"
    start_dt = f"{start}T00:00:00"
    end_dt = f"{end}T23:59:59"

    inv_settings = await get_invoice_settings_obj()
    company = inv_settings.company_name or "PureNorth Städ"
    orgnr = inv_settings.company_orgnr or ""

    # ── Fetch all data ────────────────────────────────────────────────────────
    invoices = await db.invoices.find({"created_at": {"$gte": start_dt, "$lte": end_dt}}).to_list(5000)
    paid_invoices = [i for i in invoices if i.get("status") == "paid"]
    costs = await db.costs.find({"date": {"$gte": start, "$lte": end}}).to_list(5000)
    expenses = await db.expenses.find({"date": {"$gte": start, "$lte": end}}).to_list(5000)
    employees = await db.employees.find().to_list(100)

    # ── Revenue calculations ──────────────────────────────────────────────────
    total_revenue_excl = sum(i.get("subtotal", 0) for i in invoices)
    total_vat_out = sum(i.get("vat_amount", 0) for i in invoices)
    total_rut = sum(i.get("rut_deduction", 0) for i in invoices)
    pam_fees = sum(sum(it.get("quantity",1)*it.get("unit_price",0) for it in i.get("items",[]) if "Påminnelseavgift" in it.get("service","")) for i in invoices)
    total_paid = sum(i.get("customer_pays", 0) for i in paid_invoices)
    total_unpaid = sum(i.get("customer_pays", 0) for i in invoices if i.get("status") != "paid")

    # ── Cost calculations ─────────────────────────────────────────────────────
    material_total = sum(c.get("amount", 0) for c in costs)
    material_vat_in = sum(c.get("amount",0) - c.get("amount",0)/(1+c.get("moms_rate",25)/100) for c in costs)
    expense_total = sum(e.get("amount", 0) for e in expenses)
    expense_vat_in = sum(e.get("amount",0) - e.get("amount",0)/(1+e.get("moms_rate",0)/100) for e in expenses if e.get("moms_rate",0) > 0)
    total_vat_in = round(material_vat_in + expense_vat_in, 2)

    # ── Payroll calculations ──────────────────────────────────────────────────
    PRELSKATT = 0.30
    ARBETSGIVARAVGIFT = 0.3142
    SEMESTER = 0.12
    payroll_summary, payroll_settings = await build_payroll_summary(start, end)
    total_brutto = sum(r["normal_h"]*r["hourly_rate"] + r["ob1_h"]*payroll_settings.ob1_extra + r["ob2_h"]*payroll_settings.ob2_extra + r.get("sjuklon_net",0) for r in payroll_summary.values())
    total_ag = round(total_brutto * ARBETSGIVARAVGIFT, 2)
    total_prel = round(total_brutto * PRELSKATT, 2)
    total_sem = round(total_brutto * SEMESTER, 2)
    total_netto = round(total_brutto - total_prel, 2)

    # ── Monthly breakdown ─────────────────────────────────────────────────────
    monthly = {}
    MONTHS_SV = ["","Jan","Feb","Mar","Apr","Maj","Jun","Jul","Aug","Sep","Okt","Nov","Dec"]
    for i in range(1, 13):
        monthly[i] = {"revenue": 0, "costs": 0, "invoices": 0}
    for inv in invoices:
        try:
            m = int(inv.get("created_at","")[:7].split("-")[1])
            monthly[m]["revenue"] += inv.get("subtotal", 0)
            monthly[m]["invoices"] += 1
        except: pass
    for c in costs:
        try:
            m = int(c.get("date","")[:7].split("-")[1])
            monthly[m]["costs"] += c.get("amount", 0)
        except: pass

    # ── Build PDF ─────────────────────────────────────────────────────────────
    from reportlab.lib.pagesizes import A4
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        topMargin=15*mm, bottomMargin=15*mm, leftMargin=20*mm, rightMargin=20*mm,
        title=f"Årsredovisning {year}",
        author=company)
    styles = getSampleStyleSheet()
    def ps(name, **kw): return ParagraphStyle(name, parent=styles["Normal"], **kw)
    elements = []
    page_w = A4[0] - 40*mm

    # Header
    hdr = Table([[
        Paragraph(f"<b>{company}</b>", ps("h", fontSize=16, fontName="Helvetica-Bold", textColor=colors.white)),
        Paragraph(f"<b>ÅRSREDOVISNING UNDERLAG</b><br/>Räkenskapsår: {year}<br/>Org.nr: {orgnr}", ps("s", fontSize=10, fontName="Helvetica-Bold", textColor=colors.white, alignment=2))
    ]], colWidths=[100*mm, None])
    hdr.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#141414")),("LEFTPADDING",(0,0),(-1,-1),12),("RIGHTPADDING",(0,0),(-1,-1),12),("TOPPADDING",(0,0),(-1,-1),12),("BOTTOMPADDING",(0,0),(-1,-1),12),("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
    elements.append(hdr)
    elements.append(Spacer(1,5*mm))

    # ── Section 1: Intäkter ───────────────────────────────────────────────────
    elements.append(Paragraph("1. INTÄKTER", ps("h2", fontSize=11, fontName="Helvetica-Bold", spaceBefore=4, spaceAfter=3)))
    int_data = [
        ["Post", "Konto", "Belopp (kr)"],
        ["Försäljning tjänster (exkl. moms)", "3000", f"{total_revenue_excl:.2f}"],
        ["Påminnelseavgifter (ingen moms)", "3590", f"{pam_fees:.2f}"],
        ["Utgående moms 25%", "2610", f"{total_vat_out:.2f}"],
        ["RUT-avdrag (betalas av Skatteverket)", "3001", f"{total_rut:.2f}"],
        ["TOTALA INTÄKTER", "", f"{total_revenue_excl + pam_fees:.2f}"],
        ["Varav betalda", "", f"{total_paid:.2f}"],
        ["Varav obetalda (kundfordringar)", "1510", f"{total_unpaid:.2f}"],
    ]
    t1 = Table(int_data, colWidths=[100*mm, 25*mm, None])
    t1.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1e3a5f")),("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),9),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
        ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("ALIGN",(2,0),(-1,-1),"RIGHT"),
        ("BACKGROUND",(0,5),(-1,5),colors.HexColor("#F1F5F9")),("FONTNAME",(0,5),(-1,5),"Helvetica-Bold"),
    ]))
    elements.append(t1)
    elements.append(Spacer(1,4*mm))

    # ── Section 2: Kostnader ──────────────────────────────────────────────────
    elements.append(Paragraph("2. KOSTNADER", ps("h2", fontSize=11, fontName="Helvetica-Bold", spaceBefore=4, spaceAfter=3)))
    kost_data = [
        ["Post", "Konto", "Belopp (kr)"],
        ["Bruttolöner", "7210", f"{total_brutto:.2f}"],
        ["Semesterersättning (12%)", "7290", f"{total_sem:.2f}"],
        ["Arbetsgivaravgift (31.42%)", "7510", f"{total_ag:.2f}"],
        ["Materialkostnader (inkl. moms)", "5410", f"{material_total:.2f}"],
        ["Utlägg anställda", "7331", f"{expense_total:.2f}"],
        ["TOTALA KOSTNADER", "", f"{total_brutto + total_sem + total_ag + material_total + expense_total:.2f}"],
    ]
    t2 = Table(kost_data, colWidths=[100*mm, 25*mm, None])
    t2.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#7c3aed")),("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),9),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
        ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("ALIGN",(2,0),(-1,-1),"RIGHT"),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#F1F5F9")),("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
    ]))
    elements.append(t2)
    elements.append(Spacer(1,4*mm))

    # ── Section 3: Moms ───────────────────────────────────────────────────────
    elements.append(Paragraph("3. MOMSREDOVISNING", ps("h2", fontSize=11, fontName="Helvetica-Bold", spaceBefore=4, spaceAfter=3)))
    moms_att_betala = round(total_vat_out - total_vat_in, 2)
    moms_data = [
        ["Post", "Konto", "Belopp (kr)"],
        ["Utgående moms (25% på försäljning)", "2610", f"{total_vat_out:.2f}"],
        ["Ingående moms (material + utlägg)", "2640", f"-{total_vat_in:.2f}"],
        ["NETTOMOMS ATT BETALA SKATTEVERKET", "", f"{moms_att_betala:.2f}"],
    ]
    t3 = Table(moms_data, colWidths=[100*mm, 25*mm, None])
    t3.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#0f766e")),("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),9),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
        ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("ALIGN",(2,0),(-1,-1),"RIGHT"),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#F1F5F9")),("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
    ]))
    elements.append(t3)
    elements.append(Spacer(1,4*mm))

    # ── Section 4: Lön ────────────────────────────────────────────────────────
    elements.append(Paragraph("4. LÖNESAMMANSTÄLLNING", ps("h2", fontSize=11, fontName="Helvetica-Bold", spaceBefore=4, spaceAfter=3)))
    lon_data = [["Anställd", "Brutto (kr)", "Prelskatt (kr)", "Netto (kr)", "AG-avgift (kr)"]]
    for r in payroll_summary.values():
        brutto = r["normal_h"]*r["hourly_rate"] + r["ob1_h"]*payroll_settings.ob1_extra + r["ob2_h"]*payroll_settings.ob2_extra + r.get("sjuklon_net",0)
        prel = round(brutto * PRELSKATT, 2)
        netto = round(brutto - prel, 2)
        ag = round(brutto * ARBETSGIVARAVGIFT, 2)
        lon_data.append([r["name"], f"{brutto:.2f}", f"{prel:.2f}", f"{netto:.2f}", f"{ag:.2f}"])
    lon_data.append(["TOTALT", f"{total_brutto:.2f}", f"{total_prel:.2f}", f"{total_netto:.2f}", f"{total_ag:.2f}"])
    t4 = Table(lon_data, colWidths=[60*mm, 28*mm, 28*mm, 28*mm, 28*mm])
    t4.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#b45309")),("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),9),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
        ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("ALIGN",(1,0),(-1,-1),"RIGHT"),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#F1F5F9")),("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
    ]))
    elements.append(t4)
    elements.append(Spacer(1,4*mm))

    # ── Section 5: Månadsöversikt ─────────────────────────────────────────────
    elements.append(Paragraph("5. MÅNADSÖVERSIKT", ps("h2", fontSize=11, fontName="Helvetica-Bold", spaceBefore=4, spaceAfter=3)))
    mån_data = [["Månad", "Fakturor", "Intäkter (kr)", "Kostnader (kr)", "Resultat (kr)"]]
    for m in range(1, 13):
        rev = monthly[m]["revenue"]
        cost = monthly[m]["costs"]
        res = round(rev - cost, 2)
        mån_data.append([MONTHS_SV[m], str(monthly[m]["invoices"]), f"{rev:.2f}", f"{cost:.2f}", f"{res:.2f}"])
    total_res = sum(monthly[m]["revenue"] - monthly[m]["costs"] for m in range(1,13))
    mån_data.append(["TOTALT", str(len(invoices)), f"{total_revenue_excl:.2f}", f"{material_total:.2f}", f"{total_res:.2f}"])
    t5 = Table(mån_data, colWidths=[20*mm, 20*mm, 38*mm, 38*mm, 38*mm])
    t5.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#166534")),("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8.5),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
        ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("ALIGN",(1,0),(-1,-1),"RIGHT"),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#F1F5F9")),("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
    ]))
    elements.append(t5)
    elements.append(Spacer(1,5*mm))

    # ── Resultaträkning ───────────────────────────────────────────────────────
    total_costs_all = total_brutto + total_sem + total_ag + material_total + expense_total
    resultat = round(total_revenue_excl + pam_fees - total_costs_all, 2)
    elements.append(Paragraph("RESULTATRÄKNING", ps("h2", fontSize=11, fontName="Helvetica-Bold", spaceBefore=4, spaceAfter=3)))
    res_data = [
        ["Post", "Belopp (kr)"],
        ["Nettoomsättning (exkl. moms)", f"{total_revenue_excl + pam_fees:.2f}"],
        ["Personalkostnader", f"-{total_brutto + total_sem + total_ag:.2f}"],
        ["Övriga externa kostnader", f"-{material_total + expense_total:.2f}"],
        ["RÖRELSERESULTAT (EBIT)", f"{resultat:.2f}"],
    ]
    t6 = Table(res_data, colWidths=[120*mm, None])
    t6.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#141414")),("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),9),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
        ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("ALIGN",(1,0),(-1,-1),"RIGHT"),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#DCFCE7") if resultat >= 0 else colors.HexColor("#FEE2E2")),
        ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
        ("TEXTCOLOR",(1,-1),(1,-1),colors.HexColor("#15803d") if resultat >= 0 else colors.HexColor("#dc2626")),
    ]))
    elements.append(t6)

    elements.append(Spacer(1,5*mm))
    elements.append(Paragraph(
        f"{company}  ·  Årsredovisning underlag {year}  ·  Skapad {datetime.now(timezone.utc).strftime('%Y-%m-%d')}  ·  Konfidentiellt",
        ps("ft", fontSize=7, textColor=colors.HexColor("#94a3b8"))
    ))

    doc.build(elements)
    return Response(content=buf.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="arsredovisning_{year}.pdf"'})


# ── F-skatt Kontroll ──────────────────────────────────────────────────────────
@api_router.get("/employees/fskatt-kontroll")
async def fskatt_kontroll(current=Depends(get_current_user)):
    """Get F-skatt status for all employees"""
    employees = await db.employees.find().to_list(100)
    result = []
    for e in employees:
        eid = str(e["_id"])
        skatt_typ = e.get("skatt_typ", "A-skatt")
        fskatt_nr = e.get("fskatt_nr", "")
        verified = e.get("fskatt_verified", False)
        
        # Determine status
        if skatt_typ == "F-skatt":
            if verified and fskatt_nr:
                status = "ok"
                status_text = "F-skatt verifierad"
            elif fskatt_nr and not verified:
                status = "warning"
                status_text = "F-skatt ej verifierad"
            else:
                status = "error"
                status_text = "F-skattsedel saknas"
        else:
            status = "ok"
            status_text = "A-skatt (prelskatt dras av arbetsgivare)"
        
        pnr = e.get("personnummer", "")
        pnr_result = validate_personnummer(pnr) if pnr else {"valid": None, "error": "Personnummer saknas"}
        pnr_valid = pnr_result["valid"]
        pnr_error = pnr_result.get("error")

        # Override status if personnummer is invalid
        if pnr_valid == False:
            status = "error"
            status_text = f"Ogiltigt personnummer: {pnr_error}"
        elif pnr_valid is None:
            status = "warning"
            status_text = "Personnummer saknas"

        result.append({
            "id": eid,
            "name": e.get("name", ""),
            "personnummer": pnr,
            "employment_type": e.get("employment_type", ""),
            "skatt_typ": skatt_typ,
            "fskatt_nr": fskatt_nr,
            "fskatt_verified": verified,
            "status": status,
            "status_text": status_text,
            "pnr_valid": pnr_valid,
            "pnr_error": pnr_error,
        })
    
    return {
        "employees": result,
        "total": len(result),
        "ok": sum(1 for r in result if r["status"] == "ok"),
        "warnings": sum(1 for r in result if r["status"] == "warning"),
        "errors": sum(1 for r in result if r["status"] == "error"),
    }

@api_router.patch("/employees/{employee_id}/fskatt")
async def update_fskatt(employee_id: str, payload: dict, current=Depends(get_current_user)):
    """Update F-skatt info for employee"""
    allowed = {"skatt_typ", "fskatt_nr", "fskatt_verified"}
    updates = {k: v for k, v in payload.items() if k in allowed}
    await db.employees.update_one(
        {"_id": to_object_id(employee_id)},
        {"$set": updates}
    )
    return {"success": True}


# ── Personnummer Validering ───────────────────────────────────────────────────
def validate_personnummer(pnr: str) -> dict:
    """Validate Swedish personnummer using Luhn algorithm"""
    import re
    
    # Clean input
    clean = re.sub(r'[-\s]', '', pnr)
    
    # Handle 12-digit (YYYYMMDD + 4)
    if len(clean) == 12:
        clean = clean[2:]  # Remove century
    
    if len(clean) != 10:
        return {"valid": False, "error": "Personnummer måste ha 10 siffror (YYMMDD-XXXX)"}
    
    if not clean.isdigit():
        return {"valid": False, "error": "Personnummer får bara innehålla siffror"}
    
    # Validate date part
    yy, mm, dd = int(clean[0:2]), int(clean[2:4]), int(clean[4:6])
    if mm < 1 or mm > 12:
        return {"valid": False, "error": f"Ogiltigt månadstal: {mm:02d}"}
    if dd < 1 or dd > 31:
        return {"valid": False, "error": f"Ogiltigt dagtal: {dd:02d}"}
    
    # Luhn algorithm
    digits = [int(d) for d in clean[:9]]
    total = 0
    for i, d in enumerate(digits):
        if i % 2 == 0:
            v = d * 2
            total += v - 9 if v > 9 else v
        else:
            total += d
    
    check = (10 - (total % 10)) % 10
    if check != int(clean[9]):
        return {"valid": False, "error": f"Kontrollsiffran är fel (ska vara {check}, fick {clean[9]})"}
    
    return {"valid": True, "error": None, "formatted": f"{clean[:6]}-{clean[6:]}"}

@api_router.post("/employees/validate-personnummer")
async def validate_pnr(payload: dict, current=Depends(get_current_user)):
    pnr = payload.get("personnummer", "")
    return validate_personnummer(pnr)


@api_router.patch("/invoices/by-booking/{booking_id}")
async def update_invoice_by_booking(booking_id: str, payload: dict, current=Depends(get_current_user)):
    """Update invoice linked to a booking - tries booking_id first, then customer_name"""
    
    updates = {k: v for k, v in payload.items() 
               if k in {"customer_name","customer_email","customer_phone","customer_address","notes"} and v}
    
    if not updates:
        return {"updated": False, "message": "No valid fields"}

    # Try by booking_id first
    invoice = await db.invoices.find_one({"booking_id": booking_id})
    
    # Fallback: match by customer_name from payload
    if not invoice and payload.get("customer_name"):
        invoice = await db.invoices.find_one(
            {"customer_name": payload["customer_name"]}
        )
    
    # Fallback: match by original_name (booking name before edit)
    if not invoice and payload.get("original_name"):
        invoice = await db.invoices.find_one(
            {"customer_name": payload["original_name"]}
        )

    if not invoice:
        return {"updated": False, "message": "No invoice found"}
    
    # Also save booking_id for future updates
    updates["booking_id"] = booking_id
    
    await db.invoices.update_many(
        {"_id": invoice["_id"]},
        {"$set": updates}
    )
    return {"updated": True}


@api_router.patch("/bookings/{booking_id}/link-invoice")
async def link_invoice_to_booking(booking_id: str, payload: dict, current=Depends(get_current_user)):
    """Save invoice_id to booking after invoice creation"""
    invoice_id = payload.get("invoice_id")
    await db.bookings.update_one(
        {"_id": to_object_id(booking_id)},
        {"$set": {"invoice_id": invoice_id}}
    )
    return {"success": True}

@api_router.patch("/bookings/{booking_id}/update-invoice")
async def update_invoice_from_booking(booking_id: str, payload: dict, current=Depends(get_current_user)):
    """Update linked invoice when booking is edited"""
    booking = await db.bookings.find_one({"_id": to_object_id(booking_id)})
    if not booking or not booking.get("invoice_id"):
        return {"updated": False, "message": "No linked invoice"}
    
    invoice_id = booking["invoice_id"]
    updates = {}
    
    if payload.get("name"): updates["customer_name"] = payload["name"]
    if payload.get("email"): updates["customer_email"] = payload["email"]
    if payload.get("phone"): updates["customer_phone"] = payload["phone"]
    if payload.get("address"): updates["customer_address"] = payload["address"]
    if payload.get("notes"): updates["notes"] = payload["notes"]
    
    # Update prices if provided
    if payload.get("subtotal"): updates["subtotal"] = payload["subtotal"]
    if payload.get("rut_deduction"): updates["rut_deduction"] = payload["rut_deduction"]
    if payload.get("vat_amount"): updates["vat_amount"] = payload["vat_amount"]
    if payload.get("total_amount"): updates["total_amount"] = payload["total_amount"]
    if payload.get("customer_pays"): updates["customer_pays"] = payload["customer_pays"]
    if payload.get("items"): updates["items"] = payload["items"]
    
    await db.invoices.update_one(
        {"_id": to_object_id(invoice_id)},
        {"$set": updates}
    )
    return {"updated": True, "invoice_id": invoice_id}


@api_router.post("/bookings/{booking_id}/sync-invoice")
async def sync_invoice_from_booking(booking_id: str, payload: dict, current=Depends(get_current_user)):
    """Sync invoice with booking changes"""
    booking = await db.bookings.find_one({"_id": to_object_id(booking_id)})
    if not booking or not booking.get("invoice_id"):
        return {"updated": False, "message": "No linked invoice"}
    
    invoice_id = booking["invoice_id"]
    invoice = await db.invoices.find_one({"_id": to_object_id(invoice_id)})
    if not invoice:
        return {"updated": False, "message": "Invoice not found"}
    
    updates = {}
    
    # Contact info updates
    if "name" in payload: updates["customer_name"] = payload["name"]
    if "email" in payload: updates["customer_email"] = payload["email"]
    if "phone" in payload: updates["customer_phone"] = payload["phone"]
    if "address" in payload: updates["customer_address"] = payload["address"]
    if "notes" in payload: updates["notes"] = payload["notes"]
    
    # Price recalculation if kvm or services changed
    if "kvm" in payload or "services" in payload:
        kvm = float(payload.get("kvm") or booking.get("kvm") or 0)
        services = payload.get("services") or booking.get("services") or []
        
        pricelist = await db.price_lists.find_one({}) or {}
        items_db = pricelist.get("items", [])
        active_items = [i for i in items_db if i.get("is_active")]
        
        KVM_KEYWORDS = ["städning", "storstäd", "flytt", "kontors", "fönster"]
        
        new_items = []
        labor_total = 0
        
        for svc in services:
            svc_lower = svc.lower()
            match = next((i for i in active_items if 
                i["service"].lower() in svc_lower or 
                svc_lower in i["service"].lower()), None)
            if not match:
                continue
            
            unit = match.get("unit", "tim")
            rate = match.get("price", 0)
            is_rut = match.get("is_rut_eligible", True)
            
            if unit == "kvm":
                qty = kvm or 1
                line_total = qty * rate
            elif unit == "tim":
                qty = max(1, round(kvm / 20)) if kvm else 1
                line_total = qty * rate
            else:
                qty = 1
                line_total = rate
            
            labor_total += line_total
            new_items.append({
                "service": match["service"],
                "description": match["service"],
                "quantity": qty,
                "unit_price": rate,
                "is_material": False,
                "is_rut_eligible": is_rut,
            })
        
        if new_items:
            inv_settings = await get_invoice_settings_obj()
            vat = labor_total * inv_settings.vat_rate / 100
            total = labor_total + vat
            rut = round(labor_total * 0.5, 2) if invoice.get("rut_eligible") else 0
            customer_pays = total - rut
            
            updates["items"] = new_items
            updates["labor_total"] = round(labor_total, 2)
            updates["subtotal"] = round(labor_total, 2)
            updates["vat_amount"] = round(vat, 2)
            updates["total_amount"] = round(total, 2)
            updates["rut_deduction"] = rut
            updates["customer_pays"] = round(customer_pays, 2)
    
    if updates:
        await db.invoices.update_one(
            {"_id": to_object_id(invoice_id)},
            {"$set": updates}
        )
        return {"updated": True, "invoice_id": invoice_id}
    
    return {"updated": False, "message": "Nothing to update"}

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@example.com").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        await db.users.insert_one({
            "email": admin_email,
            "password_hash": hash_password(admin_password),
            "name": "Admin",
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        logger.info("Admin seeded: %s", admin_email)
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one(
            {"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}}
        )
        logger.info("Admin password updated")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
# auto-remind deploy Fri Jun 26 22:44:29 UTC 2026
# force deploy 1782606093
